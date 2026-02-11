import os
import time
import pandas
import sqlite3
import streamlit
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import numbers
from dateutil.relativedelta import relativedelta

Execution_Path = os.getcwd()
Database_Folder = Execution_Path + '/Database/'
Database_Name = 'Betting_Database.db'
Database_Path = Database_Folder + Database_Name
Output_Folder = Execution_Path + '/Output/'

streamlit.set_page_config(layout="wide")

if 'PAGE' not in streamlit.session_state:
    streamlit.session_state['PAGE'] = 'PAGE_04'
else:
    if streamlit.session_state['PAGE'] != 'PAGE_04':
        streamlit.session_state.clear()
    else:
        pass

if 'Bulletin_Match_Selected_Index' not in streamlit.session_state:
    streamlit.session_state['Bulletin_Match_Selected_Index'] = 0

if 'Bulletin_Match_Maximum_Index' not in streamlit.session_state:
    streamlit.session_state['Bulletin_Match_Maximum_Index'] = 0

if 'Bulletin_Match_List' not in streamlit.session_state:
    streamlit.session_state['Bulletin_Match_List'] = []

if 'Bulletin_Match_DataFrame' not in streamlit.session_state:
    streamlit.session_state['Bulletin_Match_DataFrame'] = None

if 'Filters_Dictionary' not in streamlit.session_state:
    streamlit.session_state['Filters_Dictionary'] = {}

    # Initialize Filters_Dictionary
    streamlit.session_state.Filters_Dictionary['filter_start_date'] = datetime(2019, 1, 1)
    streamlit.session_state.Filters_Dictionary['filter_end_date'] = datetime(datetime.today().year, datetime.today().month, datetime.today().day)
    streamlit.session_state.Filters_Dictionary['filter_country'] = False
    streamlit.session_state.Filters_Dictionary['filter_league'] = False
    streamlit.session_state.Filters_Dictionary['filter_week'] = False
    streamlit.session_state.Filters_Dictionary['filter_hour'] = False
    streamlit.session_state.Filters_Dictionary['filter_home_team'] = False
    streamlit.session_state.Filters_Dictionary['filter_away_team'] = False
    streamlit.session_state.Filters_Dictionary['filter_referee'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_ms_1'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_ms_x'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_ms_2'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_ms_0_5_a'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_ms_0_5_ü'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_ms_1_5_a'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_ms_1_5_ü'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_ms_2_5_a'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_ms_2_5_ü'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_ms_3_5_a'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_ms_3_5_ü'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_iy_1'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_iy_x'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_iy_2'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_iy_0_5_a'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_iy_0_5_ü'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_iy_1_5_a'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_iy_1_5_ü'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_ms_1_x_cs'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_ms_1_2_cs'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_ms_x_2_cs'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_iy_1_x_cs'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_iy_1_2_cs'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_iy_x_2_cs'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_home_win'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_away_win'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_iy_kg_var'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_iy_kg_yok'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_ms_kg_var'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_ms_kg_yok'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_tek'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_cift'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_1den1'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_1denx'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_1den2'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_xden1'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_xdenx'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_xden2'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_2den1'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_2denx'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_2den2'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_10'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_20'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_21'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_30'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_31'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_32'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_40'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_41'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_42'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_43'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_00'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_11'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_22'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_33'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_44'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_01'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_02'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_12'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_03'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_13'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_23'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_04'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_14'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_24'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_34'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_10'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_20'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_21'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_30'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_31'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_32'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_40'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_41'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_42'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_43'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_00'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_11'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_22'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_33'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_44'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_01'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_02'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_12'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_03'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_13'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_23'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_04'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_14'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_24'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_34'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_neg_2'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_neg_2'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_neg_2_neg_1_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_neg_2_neg_1_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_neg_1_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_neg_1_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_neg_1_5_neg_1'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_neg_1_5_neg_1'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_neg_1'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_neg_1'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_neg_1_neg_0_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_neg_1_neg_0_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_neg_0_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_neg_0_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_neg_0_5_pos_0'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_neg_0_5_pos_0'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_pos_0_pos_0_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_pos_0_pos_0_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_pos_0_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_pos_0_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_pos_0_5_pos_1'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_pos_0_5_pos_1'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_pos_1'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_pos_1'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_pos_1_pos_1_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_pos_1_pos_1_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_pos_1_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_pos_1_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_pos_1_5_pos_2'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_pos_1_5_pos_2'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_pos_2'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_pos_2'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_neg_2'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_neg_2'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_neg_2_neg_1_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_neg_2_neg_1_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_neg_1_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_neg_1_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_neg_1_5_neg_1'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_neg_1_5_neg_1'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_neg_1'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_neg_1'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_neg_1_neg_0_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_neg_1_neg_0_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_neg_0_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_neg_0_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_neg_0_5_pos_0'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_neg_0_5_pos_0'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_pos_0_pos_0_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_pos_0_pos_0_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_pos_0_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_pos_0_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_pos_0_5_pos_1'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_pos_0_5_pos_1'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_pos_1'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_pos_1'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_pos_1_pos_1_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_pos_1_pos_1_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_pos_1_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_pos_1_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_pos_1_5_pos_2'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_pos_1_5_pos_2'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_pos_2'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_pos_2'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ev_neg_1'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ber_neg_1'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_dep_neg_1'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ev_neg_1_neg_0_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ber_neg_1_neg_0_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_dep_neg_1_neg_0_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ev_neg_0_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ber_neg_0_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_dep_neg_0_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ev_neg_0_5_pos_0'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ber_neg_0_5_pos_0'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_dep_neg_0_5_pos_0'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ev_pos_0_pos_0_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ber_pos_0_pos_0_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_dep_pos_0_pos_0_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ev_pos_0_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ber_pos_0_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_dep_pos_0_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ev_pos_0_5_pos_1'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ber_pos_0_5_pos_1'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_dep_pos_0_5_pos_1'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ev_pos_1'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ber_pos_1'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_dep_pos_1'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ev_neg_1'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ber_neg_1'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_dep_neg_1'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ev_neg_1_neg_0_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ber_neg_1_neg_0_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_dep_neg_1_neg_0_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ev_neg_0_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ber_neg_0_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_dep_neg_0_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ev_neg_0_5_pos_0'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ber_neg_0_5_pos_0'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_dep_neg_0_5_pos_0'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ev_pos_0_pos_0_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ber_pos_0_pos_0_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_dep_pos_0_pos_0_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ev_pos_0_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ber_pos_0_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_dep_pos_0_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ev_pos_0_5_pos_1'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ber_pos_0_5_pos_1'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_dep_pos_0_5_pos_1'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ev_pos_1'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ber_pos_1'] = False
    streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_dep_pos_1'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_ms_1'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_ms_x'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_ms_2'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_ms_0_5_a'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_ms_0_5_ü'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_ms_1_5_a'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_ms_1_5_ü'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_ms_2_5_a'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_ms_2_5_ü'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_ms_3_5_a'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_ms_3_5_ü'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_iy_1'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_iy_x'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_iy_2'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_iy_0_5_a'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_iy_0_5_ü'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_iy_1_5_a'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_iy_1_5_ü'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_ms_1_x_cs'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_ms_1_2_cs'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_ms_x_2_cs'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_iy_1_x_cs'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_iy_1_2_cs'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_iy_x_2_cs'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_home_win'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_away_win'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_iy_kg_var'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_iy_kg_yok'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_ms_kg_var'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_ms_kg_yok'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_tek'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_cift'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_1den1'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_1denx'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_1den2'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_xden1'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_xdenx'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_xden2'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_2den1'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_2denx'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_2den2'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_10'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_20'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_21'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_30'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_31'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_32'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_40'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_41'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_42'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_43'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_00'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_11'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_22'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_33'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_44'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_01'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_02'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_12'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_03'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_13'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_23'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_04'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_14'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_24'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_34'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_10'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_20'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_21'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_30'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_31'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_32'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_40'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_41'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_42'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_43'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_00'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_11'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_22'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_33'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_44'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_01'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_02'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_12'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_03'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_13'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_23'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_04'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_14'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_24'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_34'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_neg_2'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_neg_2'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_neg_2_neg_1_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_neg_2_neg_1_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_neg_1_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_neg_1_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_neg_1_5_neg_1'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_neg_1_5_neg_1'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_neg_1'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_neg_1'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_neg_1_neg_0_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_neg_1_neg_0_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_neg_0_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_neg_0_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_neg_0_5_pos_0'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_neg_0_5_pos_0'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_pos_0_pos_0_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_pos_0_pos_0_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_pos_0_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_pos_0_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_pos_0_5_pos_1'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_pos_0_5_pos_1'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_pos_1'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_pos_1'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_pos_1_pos_1_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_pos_1_pos_1_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_pos_1_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_pos_1_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_pos_1_5_pos_2'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_pos_1_5_pos_2'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_pos_2'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_pos_2'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_neg_2'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_neg_2'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_neg_2_neg_1_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_neg_2_neg_1_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_neg_1_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_neg_1_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_neg_1_5_neg_1'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_neg_1_5_neg_1'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_neg_1'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_neg_1'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_neg_1_neg_0_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_neg_1_neg_0_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_neg_0_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_neg_0_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_neg_0_5_pos_0'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_neg_0_5_pos_0'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_pos_0_pos_0_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_pos_0_pos_0_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_pos_0_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_pos_0_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_pos_0_5_pos_1'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_pos_0_5_pos_1'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_pos_1'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_pos_1'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_pos_1_pos_1_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_pos_1_pos_1_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_pos_1_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_pos_1_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_pos_1_5_pos_2'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_pos_1_5_pos_2'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_pos_2'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_pos_2'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ev_neg_1'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ber_neg_1'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_dep_neg_1'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ev_neg_1_neg_0_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ber_neg_1_neg_0_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_dep_neg_1_neg_0_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ev_neg_0_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ber_neg_0_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_dep_neg_0_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ev_neg_0_5_pos_0'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ber_neg_0_5_pos_0'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_dep_neg_0_5_pos_0'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ev_pos_0_pos_0_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ber_pos_0_pos_0_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_dep_pos_0_pos_0_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ev_pos_0_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ber_pos_0_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_dep_pos_0_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ev_pos_0_5_pos_1'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ber_pos_0_5_pos_1'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_dep_pos_0_5_pos_1'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ev_pos_1'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ber_pos_1'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_dep_pos_1'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ev_neg_1'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ber_neg_1'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_dep_neg_1'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ev_neg_1_neg_0_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ber_neg_1_neg_0_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_dep_neg_1_neg_0_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ev_neg_0_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ber_neg_0_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_dep_neg_0_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ev_neg_0_5_pos_0'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ber_neg_0_5_pos_0'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_dep_neg_0_5_pos_0'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ev_pos_0_pos_0_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ber_pos_0_pos_0_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_dep_pos_0_pos_0_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ev_pos_0_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ber_pos_0_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_dep_pos_0_5'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ev_pos_0_5_pos_1'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ber_pos_0_5_pos_1'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_dep_pos_0_5_pos_1'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ev_pos_1'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ber_pos_1'] = False
    streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_dep_pos_1'] = False

Filter_SQLQuery_Cross_Dictionary = {
    'filter_start_date': 'date',
    'filter_end_date': 'date',
    'filter_country': 'country',
    'filter_league': 'league',
    'filter_week': 'round',
    'filter_hour': 'hour',
    'filter_home_team': 'home_team_name',
    'filter_away_team': 'away_team_name',
    'filter_referee': 'referee',
    'filter_a_ms_1': 'fulltime_home_open',
    'filter_a_ms_x': 'fulltime_tie_open',
    'filter_a_ms_2': 'fulltime_away_open',
    'filter_a_ms_0_5_a': 'match_time_under_0_5_open',
    'filter_a_ms_0_5_ü': 'match_time_over_0_5_open',
    'filter_a_ms_1_5_a': 'match_time_under_1_5_open',
    'filter_a_ms_1_5_ü': 'match_time_over_1_5_open',
    'filter_a_ms_2_5_a': 'match_time_under_2_5_open',
    'filter_a_ms_2_5_ü': 'match_time_over_2_5_open',
    'filter_a_ms_3_5_a': 'match_time_under_3_5_open',
    'filter_a_ms_3_5_ü': 'match_time_over_3_5_open',
    'filter_a_iy_1': 'firsthalf_home_open',
    'filter_a_iy_x': 'firsthalf_tie_open',
    'filter_a_iy_2': 'firsthalf_away_open',
    'filter_a_iy_0_5_a': 'first_half_under_0_5_open',
    'filter_a_iy_0_5_ü': 'first_half_over_0_5_open',
    'filter_a_iy_1_5_a': 'first_half_under_1_5_open',
    'filter_a_iy_1_5_ü': 'first_half_over_1_5_open',
    'filter_a_ms_1_x_cs': 'match_time_home_and_tie_open',
    'filter_a_ms_1_2_cs': 'match_time_home_and_away_open',
    'filter_a_ms_x_2_cs': 'match_time_tie_and_away_open',
    'filter_a_iy_1_x_cs': 'first_half_home_and_tie_open',
    'filter_a_iy_1_2_cs': 'first_half_home_and_away_open',
    'filter_a_iy_x_2_cs': 'first_half_tie_and_away_open',
    'filter_a_home_win': 'draw_no_bet_home_open',
    'filter_a_away_win': 'draw_no_bet_away_open',
    'filter_a_iy_kg_var': 'firsthalf_both_teams_to_score_yes_open',
    'filter_a_iy_kg_yok': 'firsthalf_both_teams_to_score_no_open',
    'filter_a_ms_kg_var': 'both_teams_to_score_yes_open',
    'filter_a_ms_kg_yok': 'both_teams_to_score_no_open',
    'filter_a_tek': 'odd_open',
    'filter_a_cift': 'even_open',
    'filter_a_1den1': 'home_home_open',
    'filter_a_1denx': 'home_tie_open',
    'filter_a_1den2': 'home_away_open',
    'filter_a_xden1': 'tie_home_open',
    'filter_a_xdenx': 'tie_tie_open',
    'filter_a_xden2': 'tie_away_open',
    'filter_a_2den1': 'away_home_open',
    'filter_a_2denx': 'away_tie_open',
    'filter_a_2den2': 'away_away_open',
    'filter_a_iy_skor_10': 'first_half_1_0_open',
    'filter_a_iy_skor_20': 'first_half_2_0_open',
    'filter_a_iy_skor_21': 'first_half_2_1_open',
    'filter_a_iy_skor_30': 'first_half_3_0_open',
    'filter_a_iy_skor_31': 'first_half_3_1_open',
    'filter_a_iy_skor_32': 'first_half_3_2_open',
    'filter_a_iy_skor_40': 'first_half_4_0_open',
    'filter_a_iy_skor_41': 'first_half_4_1_open',
    'filter_a_iy_skor_42': 'first_half_4_2_open',
    'filter_a_iy_skor_43': 'first_half_4_3_open',
    'filter_a_iy_skor_00': 'first_half_0_0_open',
    'filter_a_iy_skor_11': 'first_half_1_1_open',
    'filter_a_iy_skor_22': 'first_half_2_2_open',
    'filter_a_iy_skor_33': 'first_half_3_3_open',
    'filter_a_iy_skor_44': 'first_half_4_4_open',
    'filter_a_iy_skor_01': 'first_half_0_1_open',
    'filter_a_iy_skor_02': 'first_half_0_2_open',
    'filter_a_iy_skor_12': 'first_half_1_2_open',
    'filter_a_iy_skor_03': 'first_half_0_3_open',
    'filter_a_iy_skor_13': 'first_half_1_3_open',
    'filter_a_iy_skor_23': 'first_half_2_3_open',
    'filter_a_iy_skor_04': 'first_half_0_4_open',
    'filter_a_iy_skor_14': 'first_half_1_4_open',
    'filter_a_iy_skor_24': 'first_half_2_4_open',
    'filter_a_iy_skor_34': 'first_half_3_4_open',
    'filter_a_ms_skor_10': 'match_time_1_0_open',
    'filter_a_ms_skor_20': 'match_time_2_0_open',
    'filter_a_ms_skor_21': 'match_time_2_1_open',
    'filter_a_ms_skor_30': 'match_time_3_0_open',
    'filter_a_ms_skor_31': 'match_time_3_1_open',
    'filter_a_ms_skor_32': 'match_time_3_2_open',
    'filter_a_ms_skor_40': 'match_time_4_0_open',
    'filter_a_ms_skor_41': 'match_time_4_1_open',
    'filter_a_ms_skor_42': 'match_time_4_2_open',
    'filter_a_ms_skor_43': 'match_time_4_3_open',
    'filter_a_ms_skor_00': 'match_time_0_0_open',
    'filter_a_ms_skor_11': 'match_time_1_1_open',
    'filter_a_ms_skor_22': 'match_time_2_2_open',
    'filter_a_ms_skor_33': 'match_time_3_3_open',
    'filter_a_ms_skor_44': 'match_time_4_4_open',
    'filter_a_ms_skor_01': 'match_time_0_1_open',
    'filter_a_ms_skor_02': 'match_time_0_2_open',
    'filter_a_ms_skor_12': 'match_time_1_2_open',
    'filter_a_ms_skor_03': 'match_time_0_3_open',
    'filter_a_ms_skor_13': 'match_time_1_3_open',
    'filter_a_ms_skor_23': 'match_time_2_3_open',
    'filter_a_ms_skor_04': 'match_time_0_4_open',
    'filter_a_ms_skor_14': 'match_time_1_4_open',
    'filter_a_ms_skor_24': 'match_time_2_4_open',
    'filter_a_ms_skor_34': 'match_time_3_4_open',
    'filter_a_iy_asya_ev_neg_2': 'halftime_asian_handicap_neg2_home_open',
    'filter_a_iy_asya_dep_neg_2': 'halftime_asian_handicap_neg2_away_open',
    'filter_a_iy_asya_ev_neg_2_neg_1_5': 'halftime_asian_handicap_neg2_neg1_5_home_open',
    'filter_a_iy_asya_dep_neg_2_neg_1_5': 'halftime_asian_handicap_neg2_neg1_5_away_open',
    'filter_a_iy_asya_ev_neg_1_5': 'halftime_asian_handicap_neg1_5_home_open',
    'filter_a_iy_asya_dep_neg_1_5': 'halftime_asian_handicap_neg1_5_away_open',
    'filter_a_iy_asya_ev_neg_1_5_neg_1': 'halftime_asian_handicap_neg1_5_neg_1_home_open',
    'filter_a_iy_asya_dep_neg_1_5_neg_1': 'halftime_asian_handicap_neg1_5_neg_1_away_open',
    'filter_a_iy_asya_ev_neg_1': 'halftime_asian_handicap_neg_1_home_open',
    'filter_a_iy_asya_dep_neg_1': 'halftime_asian_handicap_neg_1_away_open',
    'filter_a_iy_asya_ev_neg_1_neg_0_5': 'halftime_asian_handicap_neg_1_neg_0_5_home_open',
    'filter_a_iy_asya_dep_neg_1_neg_0_5': 'halftime_asian_handicap_neg_1_neg_0_5_away_open',
    'filter_a_iy_asya_ev_neg_0_5': 'halftime_asian_handicap_neg_0_5_home_open',
    'filter_a_iy_asya_dep_neg_0_5': 'halftime_asian_handicap_neg_0_5_away_open',
    'filter_a_iy_asya_ev_neg_0_5_pos_0': 'halftime_asian_handicap_neg_0_5_0_0_home_open',
    'filter_a_iy_asya_dep_neg_0_5_pos_0': 'halftime_asian_handicap_neg_0_5_0_0_away_open',
    'filter_a_iy_asya_ev_pos_0_pos_0_5': 'halftime_asian_handicap_0_0_pos_0_5_home_open',
    'filter_a_iy_asya_dep_pos_0_pos_0_5': 'halftime_asian_handicap_0_0_pos_0_5_away_open',
    'filter_a_iy_asya_ev_pos_0_5': 'halftime_asian_handicap_pos_0_5_home_open',
    'filter_a_iy_asya_dep_pos_0_5': 'halftime_asian_handicap_pos_0_5_away_open',
    'filter_a_iy_asya_ev_pos_0_5_pos_1': 'halftime_asian_handicap_pos_0_5_pos_1_0_home_open',
    'filter_a_iy_asya_dep_pos_0_5_pos_1': 'halftime_asian_handicap_pos_0_5_pos_1_0_away_open',
    'filter_a_iy_asya_ev_pos_1': 'halftime_asian_handicap_pos_1_home_open',
    'filter_a_iy_asya_dep_pos_1': 'halftime_asian_handicap_pos_1_away_open',
    'filter_a_iy_asya_ev_pos_1_pos_1_5': 'halftime_asian_handicap_pos_1_0_pos_1_5_home_open',
    'filter_a_iy_asya_dep_pos_1_pos_1_5': 'halftime_asian_handicap_pos_1_0_pos_1_5_away_open',
    'filter_a_iy_asya_ev_pos_1_5': 'halftime_asian_handicap_pos_1_5_home_open',
    'filter_a_iy_asya_dep_pos_1_5': 'halftime_asian_handicap_pos_1_5_away_open',
    'filter_a_iy_asya_ev_pos_1_5_pos_2': 'halftime_asian_handicap_pos_1_5_pos_2_0_home_open',
    'filter_a_iy_asya_dep_pos_1_5_pos_2': 'halftime_asian_handicap_pos_1_5_pos_2_0_away_open',
    'filter_a_iy_asya_ev_pos_2': 'halftime_asian_handicap_pos_2_home_open',
    'filter_a_iy_asya_dep_pos_2': 'halftime_asian_handicap_pos_2_away_open',
    'filter_a_ms_asya_ev_neg_2': 'fulltime_asian_handicap_neg2_home_open',
    'filter_a_ms_asya_dep_neg_2': 'fulltime_asian_handicap_neg2_away_open',
    'filter_a_ms_asya_ev_neg_2_neg_1_5': 'fulltime_asian_handicap_neg2_neg1_5_home_open',
    'filter_a_ms_asya_dep_neg_2_neg_1_5': 'fulltime_asian_handicap_neg2_neg1_5_away_open',
    'filter_a_ms_asya_ev_neg_1_5': 'fulltime_asian_handicap_neg1_5_home_open',
    'filter_a_ms_asya_dep_neg_1_5': 'fulltime_asian_handicap_neg1_5_away_open',
    'filter_a_ms_asya_ev_neg_1_5_neg_1': 'fulltime_asian_handicap_neg1_5_neg_1_home_open',
    'filter_a_ms_asya_dep_neg_1_5_neg_1': 'fulltime_asian_handicap_neg1_5_neg_1_away_open',
    'filter_a_ms_asya_ev_neg_1': 'fulltime_asian_handicap_neg_1_home_open',
    'filter_a_ms_asya_dep_neg_1': 'fulltime_asian_handicap_neg_1_away_open',
    'filter_a_ms_asya_ev_neg_1_neg_0_5': 'fulltime_asian_handicap_neg_1_neg_0_5_home_open',
    'filter_a_ms_asya_dep_neg_1_neg_0_5': 'fulltime_asian_handicap_neg_1_neg_0_5_away_open',
    'filter_a_ms_asya_ev_neg_0_5': 'fulltime_asian_handicap_neg_0_5_home_open',
    'filter_a_ms_asya_dep_neg_0_5': 'fulltime_asian_handicap_neg_0_5_away_open',
    'filter_a_ms_asya_ev_neg_0_5_pos_0': 'fulltime_asian_handicap_neg_0_5_0_0_home_open',
    'filter_a_ms_asya_dep_neg_0_5_pos_0': 'fulltime_asian_handicap_neg_0_5_0_0_away_open',
    'filter_a_ms_asya_ev_pos_0_pos_0_5': 'fulltime_asian_handicap_0_0_pos_0_5_home_open',
    'filter_a_ms_asya_dep_pos_0_pos_0_5': 'fulltime_asian_handicap_0_0_pos_0_5_away_open',
    'filter_a_ms_asya_ev_pos_0_5': 'fulltime_asian_handicap_pos_0_5_home_open',
    'filter_a_ms_asya_dep_pos_0_5': 'fulltime_asian_handicap_pos_0_5_away_open',
    'filter_a_ms_asya_ev_pos_0_5_pos_1': 'fulltime_asian_handicap_pos_0_5_pos_1_0_home_open',
    'filter_a_ms_asya_dep_pos_0_5_pos_1': 'fulltime_asian_handicap_pos_0_5_pos_1_0_away_open',
    'filter_a_ms_asya_ev_pos_1': 'fulltime_asian_handicap_pos_1_home_open',
    'filter_a_ms_asya_dep_pos_1': 'fulltime_asian_handicap_pos_1_away_open',
    'filter_a_ms_asya_ev_pos_1_pos_1_5': 'fulltime_asian_handicap_pos_1_0_pos_1_5_home_open',
    'filter_a_ms_asya_dep_pos_1_pos_1_5': 'fulltime_asian_handicap_pos_1_0_pos_1_5_away_open',
    'filter_a_ms_asya_ev_pos_1_5': 'fulltime_asian_handicap_pos_1_5_home_open',
    'filter_a_ms_asya_dep_pos_1_5': 'fulltime_asian_handicap_pos_1_5_away_open',
    'filter_a_ms_asya_ev_pos_1_5_pos_2': 'fulltime_asian_handicap_pos_1_5_pos_2_0_home_open',
    'filter_a_ms_asya_dep_pos_1_5_pos_2': 'fulltime_asian_handicap_pos_1_5_pos_2_0_away_open',
    'filter_a_ms_asya_ev_pos_2': 'fulltime_asian_handicap_pos_2_0_home_open',
    'filter_a_ms_asya_dep_pos_2': 'fulltime_asian_handicap_pos_2_0_away_open',
    'filter_a_iy_avrupa_ev_neg_1': 'firsthalf_european_handicap_neg_1_home_open',
    'filter_a_iy_avrupa_ber_neg_1': 'firsthalf_european_handicap_neg_1_tie_open',
    'filter_a_iy_avrupa_dep_neg_1': 'firsthalf_european_handicap_neg_1_away_open',
    'filter_a_iy_avrupa_ev_neg_1_neg_0_5': 'firsthalf_european_handicap_neg_1_neg_0_5_home_open',
    'filter_a_iy_avrupa_ber_neg_1_neg_0_5': 'firsthalf_european_handicap_neg_1_neg_0_5_tie_open',
    'filter_a_iy_avrupa_dep_neg_1_neg_0_5': 'firsthalf_european_handicap_neg_1_neg_0_5_away_open',
    'filter_a_iy_avrupa_ev_neg_0_5': 'firsthalf_european_handicap_neg_0_5_home_open',
    'filter_a_iy_avrupa_ber_neg_0_5': 'firsthalf_european_handicap_neg_0_5_tie_open',
    'filter_a_iy_avrupa_dep_neg_0_5': 'firsthalf_european_handicap_neg_0_5_away_open',
    'filter_a_iy_avrupa_ev_neg_0_5_pos_0': 'firsthalf_european_handicap_neg_0_5_0_0_home_open',
    'filter_a_iy_avrupa_ber_neg_0_5_pos_0': 'firsthalf_european_handicap_neg_0_5_0_0_tie_open',
    'filter_a_iy_avrupa_dep_neg_0_5_pos_0': 'firsthalf_european_handicap_neg_0_5_0_0_away_open',
    'filter_a_iy_avrupa_ev_pos_0_pos_0_5': 'firsthalf_european_handicap_0_0_pos_0_5_home_open',
    'filter_a_iy_avrupa_ber_pos_0_pos_0_5': 'firsthalf_european_handicap_0_0_pos_0_5_tie_open',
    'filter_a_iy_avrupa_dep_pos_0_pos_0_5': 'firsthalf_european_handicap_0_0_pos_0_5_away_open',
    'filter_a_iy_avrupa_ev_pos_0_5': 'firsthalf_european_handicap_pos_0_5_home_open',
    'filter_a_iy_avrupa_ber_pos_0_5': 'firsthalf_european_handicap_pos_0_5_tie_open',
    'filter_a_iy_avrupa_dep_pos_0_5': 'firsthalf_european_handicap_pos_0_5_away_open',
    'filter_a_iy_avrupa_ev_pos_0_5_pos_1': 'firsthalf_european_handicap_pos_0_5_pos_1_0_home_open',
    'filter_a_iy_avrupa_ber_pos_0_5_pos_1': 'firsthalf_european_handicap_pos_0_5_pos_1_0_tie_open',
    'filter_a_iy_avrupa_dep_pos_0_5_pos_1': 'firsthalf_european_handicap_pos_0_5_pos_1_0_away_open',
    'filter_a_iy_avrupa_ev_pos_1': 'firsthalf_european_handicap_pos_1_home_open',
    'filter_a_iy_avrupa_ber_pos_1': 'firsthalf_european_handicap_pos_1_tie_open',
    'filter_a_iy_avrupa_dep_pos_1': 'firsthalf_european_handicap_pos_1_away_open',
    'filter_a_ms_avrupa_ev_neg_1': 'fulltime_european_handicap_neg_1_home_open',
    'filter_a_ms_avrupa_ber_neg_1': 'fulltime_european_handicap_neg_1_tie_open',
    'filter_a_ms_avrupa_dep_neg_1': 'fulltime_european_handicap_neg_1_away_open',
    'filter_a_ms_avrupa_ev_neg_1_neg_0_5': 'fulltime_european_handicap_neg_1_neg_0_5_home_open',
    'filter_a_ms_avrupa_ber_neg_1_neg_0_5': 'fulltime_european_handicap_neg_1_neg_0_5_tie_open',
    'filter_a_ms_avrupa_dep_neg_1_neg_0_5': 'fulltime_european_handicap_neg_1_neg_0_5_away_open',
    'filter_a_ms_avrupa_ev_neg_0_5': 'fulltime_european_handicap_neg_0_5_home_open',
    'filter_a_ms_avrupa_ber_neg_0_5': 'fulltime_european_handicap_neg_0_5_tie_open',
    'filter_a_ms_avrupa_dep_neg_0_5': 'fulltime_european_handicap_neg_0_5_away_open',
    'filter_a_ms_avrupa_ev_neg_0_5_pos_0': 'fulltime_european_handicap_neg_0_5_0_0_home_open',
    'filter_a_ms_avrupa_ber_neg_0_5_pos_0': 'fulltime_european_handicap_neg_0_5_0_0_tie_open',
    'filter_a_ms_avrupa_dep_neg_0_5_pos_0': 'fulltime_european_handicap_neg_0_5_0_0_away_open',
    'filter_a_ms_avrupa_ev_pos_0_pos_0_5': 'fulltime_european_handicap_0_0_pos_0_5_home_open',
    'filter_a_ms_avrupa_ber_pos_0_pos_0_5': 'fulltime_european_handicap_0_0_pos_0_5_tie_open',
    'filter_a_ms_avrupa_dep_pos_0_pos_0_5': 'fulltime_european_handicap_0_0_pos_0_5_away_open',
    'filter_a_ms_avrupa_ev_pos_0_5': 'fulltime_european_handicap_pos_0_5_home_open',
    'filter_a_ms_avrupa_ber_pos_0_5': 'fulltime_european_handicap_pos_0_5_tie_open',
    'filter_a_ms_avrupa_dep_pos_0_5': 'fulltime_european_handicap_pos_0_5_away_open',
    'filter_a_ms_avrupa_ev_pos_0_5_pos_1': 'fulltime_european_handicap_pos_0_5_pos_1_0_home_open',
    'filter_a_ms_avrupa_ber_pos_0_5_pos_1': 'fulltime_european_handicap_pos_0_5_pos_1_0_tie_open',
    'filter_a_ms_avrupa_dep_pos_0_5_pos_1': 'fulltime_european_handicap_pos_0_5_pos_1_0_away_open',
    'filter_a_ms_avrupa_ev_pos_1': 'fulltime_european_handicap_pos_1_home_open',
    'filter_a_ms_avrupa_ber_pos_1': 'fulltime_european_handicap_pos_1_tie_open',
    'filter_a_ms_avrupa_dep_pos_1': 'fulltime_european_handicap_pos_1_away_open',
    'filter_k_ms_1': 'fulltime_home_close',
    'filter_k_ms_x': 'fulltime_tie_close',
    'filter_k_ms_2': 'fulltime_away_close',
    'filter_k_ms_0_5_a': 'match_time_under_0_5_close',
    'filter_k_ms_0_5_ü': 'match_time_over_0_5_close',
    'filter_k_ms_1_5_a': 'match_time_under_1_5_close',
    'filter_k_ms_1_5_ü': 'match_time_over_1_5_close',
    'filter_k_ms_2_5_a': 'match_time_under_2_5_close',
    'filter_k_ms_2_5_ü': 'match_time_over_2_5_close',
    'filter_k_ms_3_5_a': 'match_time_under_3_5_close',
    'filter_k_ms_3_5_ü': 'match_time_over_3_5_close',
    'filter_k_iy_1': 'firsthalf_home_close',
    'filter_k_iy_x': 'firsthalf_tie_close',
    'filter_k_iy_2': 'firsthalf_away_close',
    'filter_k_iy_0_5_a': 'first_half_under_0_5_close',
    'filter_k_iy_0_5_ü': 'first_half_over_0_5_close',
    'filter_k_iy_1_5_a': 'first_half_under_1_5_close',
    'filter_k_iy_1_5_ü': 'first_half_over_1_5_close',
    'filter_k_ms_1_x_cs': 'match_time_home_and_tie_close',
    'filter_k_ms_1_2_cs': 'match_time_home_and_away_close',
    'filter_k_ms_x_2_cs': 'match_time_tie_and_away_close',
    'filter_k_iy_1_x_cs': 'first_half_home_and_tie_close',
    'filter_k_iy_1_2_cs': 'first_half_home_and_away_close',
    'filter_k_iy_x_2_cs': 'first_half_tie_and_away_close',
    'filter_k_home_win': 'draw_no_bet_home_close',
    'filter_k_away_win': 'draw_no_bet_away_close',
    'filter_k_iy_kg_var': 'firsthalf_both_teams_to_score_yes_close',
    'filter_k_iy_kg_yok': 'firsthalf_both_teams_to_score_no_close',
    'filter_k_ms_kg_var': 'both_teams_to_score_yes_close',
    'filter_k_ms_kg_yok': 'both_teams_to_score_no_close',
    'filter_k_tek': 'odd_close',
    'filter_k_cift': 'even_close',
    'filter_k_1den1': 'home_home_close',
    'filter_k_1denx': 'home_tie_close',
    'filter_k_1den2': 'home_away_close',
    'filter_k_xden1': 'tie_home_close',
    'filter_k_xdenx': 'tie_tie_close',
    'filter_k_xden2': 'tie_away_close',
    'filter_k_2den1': 'away_home_close',
    'filter_k_2denx': 'away_tie_close',
    'filter_k_2den2': 'away_away_close',
    'filter_k_iy_skor_10': 'first_half_1_0_close',
    'filter_k_iy_skor_20': 'first_half_2_0_close',
    'filter_k_iy_skor_21': 'first_half_2_1_close',
    'filter_k_iy_skor_30': 'first_half_3_0_close',
    'filter_k_iy_skor_31': 'first_half_3_1_close',
    'filter_k_iy_skor_32': 'first_half_3_2_close',
    'filter_k_iy_skor_40': 'first_half_4_0_close',
    'filter_k_iy_skor_41': 'first_half_4_1_close',
    'filter_k_iy_skor_42': 'first_half_4_2_close',
    'filter_k_iy_skor_43': 'first_half_4_3_close',
    'filter_k_iy_skor_00': 'first_half_0_0_close',
    'filter_k_iy_skor_11': 'first_half_1_1_close',
    'filter_k_iy_skor_22': 'first_half_2_2_close',
    'filter_k_iy_skor_33': 'first_half_3_3_close',
    'filter_k_iy_skor_44': 'first_half_4_4_close',
    'filter_k_iy_skor_01': 'first_half_0_1_close',
    'filter_k_iy_skor_02': 'first_half_0_2_close',
    'filter_k_iy_skor_12': 'first_half_1_2_close',
    'filter_k_iy_skor_03': 'first_half_0_3_close',
    'filter_k_iy_skor_13': 'first_half_1_3_close',
    'filter_k_iy_skor_23': 'first_half_2_3_close',
    'filter_k_iy_skor_04': 'first_half_0_4_close',
    'filter_k_iy_skor_14': 'first_half_1_4_close',
    'filter_k_iy_skor_24': 'first_half_2_4_close',
    'filter_k_iy_skor_34': 'first_half_3_4_close',
    'filter_k_ms_skor_10': 'match_time_1_0_close',
    'filter_k_ms_skor_20': 'match_time_2_0_close',
    'filter_k_ms_skor_21': 'match_time_2_1_close',
    'filter_k_ms_skor_30': 'match_time_3_0_close',
    'filter_k_ms_skor_31': 'match_time_3_1_close',
    'filter_k_ms_skor_32': 'match_time_3_2_close',
    'filter_k_ms_skor_40': 'match_time_4_0_close',
    'filter_k_ms_skor_41': 'match_time_4_1_close',
    'filter_k_ms_skor_42': 'match_time_4_2_close',
    'filter_k_ms_skor_43': 'match_time_4_3_close',
    'filter_k_ms_skor_00': 'match_time_0_0_close',
    'filter_k_ms_skor_11': 'match_time_1_1_close',
    'filter_k_ms_skor_22': 'match_time_2_2_close',
    'filter_k_ms_skor_33': 'match_time_3_3_close',
    'filter_k_ms_skor_44': 'match_time_4_4_close',
    'filter_k_ms_skor_01': 'match_time_0_1_close',
    'filter_k_ms_skor_02': 'match_time_0_2_close',
    'filter_k_ms_skor_12': 'match_time_1_2_close',
    'filter_k_ms_skor_03': 'match_time_0_3_close',
    'filter_k_ms_skor_13': 'match_time_1_3_close',
    'filter_k_ms_skor_23': 'match_time_2_3_close',
    'filter_k_ms_skor_04': 'match_time_0_4_close',
    'filter_k_ms_skor_14': 'match_time_1_4_close',
    'filter_k_ms_skor_24': 'match_time_2_4_close',
    'filter_k_ms_skor_34': 'match_time_3_4_close',
    'filter_k_iy_asya_ev_neg_2': 'halftime_asian_handicap_neg2_home_close',
    'filter_k_iy_asya_dep_neg_2': 'halftime_asian_handicap_neg2_away_close',
    'filter_k_iy_asya_ev_neg_2_neg_1_5': 'halftime_asian_handicap_neg2_neg1_5_home_close',
    'filter_k_iy_asya_dep_neg_2_neg_1_5': 'halftime_asian_handicap_neg2_neg1_5_away_close',
    'filter_k_iy_asya_ev_neg_1_5': 'halftime_asian_handicap_neg1_5_home_close',
    'filter_k_iy_asya_dep_neg_1_5': 'halftime_asian_handicap_neg1_5_away_close',
    'filter_k_iy_asya_ev_neg_1_5_neg_1': 'halftime_asian_handicap_neg1_5_neg_1_home_close',
    'filter_k_iy_asya_dep_neg_1_5_neg_1': 'halftime_asian_handicap_neg1_5_neg_1_away_close',
    'filter_k_iy_asya_ev_neg_1': 'halftime_asian_handicap_neg_1_home_close',
    'filter_k_iy_asya_dep_neg_1': 'halftime_asian_handicap_neg_1_away_close',
    'filter_k_iy_asya_ev_neg_1_neg_0_5': 'halftime_asian_handicap_neg_1_neg_0_5_home_close',
    'filter_k_iy_asya_dep_neg_1_neg_0_5': 'halftime_asian_handicap_neg_1_neg_0_5_away_close',
    'filter_k_iy_asya_ev_neg_0_5': 'halftime_asian_handicap_neg_0_5_home_close',
    'filter_k_iy_asya_dep_neg_0_5': 'halftime_asian_handicap_neg_0_5_away_close',
    'filter_k_iy_asya_ev_neg_0_5_pos_0': 'halftime_asian_handicap_neg_0_5_0_0_home_close',
    'filter_k_iy_asya_dep_neg_0_5_pos_0': 'halftime_asian_handicap_neg_0_5_0_0_away_close',
    'filter_k_iy_asya_ev_pos_0_pos_0_5': 'halftime_asian_handicap_0_0_pos_0_5_home_close',
    'filter_k_iy_asya_dep_pos_0_pos_0_5': 'halftime_asian_handicap_0_0_pos_0_5_away_close',
    'filter_k_iy_asya_ev_pos_0_5': 'halftime_asian_handicap_pos_0_5_home_close',
    'filter_k_iy_asya_dep_pos_0_5': 'halftime_asian_handicap_pos_0_5_away_close',
    'filter_k_iy_asya_ev_pos_0_5_pos_1': 'halftime_asian_handicap_pos_0_5_pos_1_0_home_close',
    'filter_k_iy_asya_dep_pos_0_5_pos_1': 'halftime_asian_handicap_pos_0_5_pos_1_0_away_close',
    'filter_k_iy_asya_ev_pos_1': 'halftime_asian_handicap_pos_1_home_close',
    'filter_k_iy_asya_dep_pos_1': 'halftime_asian_handicap_pos_1_away_close',
    'filter_k_iy_asya_ev_pos_1_pos_1_5': 'halftime_asian_handicap_pos_1_0_pos_1_5_home_close',
    'filter_k_iy_asya_dep_pos_1_pos_1_5': 'halftime_asian_handicap_pos_1_0_pos_1_5_away_close',
    'filter_k_iy_asya_ev_pos_1_5': 'halftime_asian_handicap_pos_1_5_home_close',
    'filter_k_iy_asya_dep_pos_1_5': 'halftime_asian_handicap_pos_1_5_away_close',
    'filter_k_iy_asya_ev_pos_1_5_pos_2': 'halftime_asian_handicap_pos_1_5_pos_2_0_home_close',
    'filter_k_iy_asya_dep_pos_1_5_pos_2': 'halftime_asian_handicap_pos_1_5_pos_2_0_away_close',
    'filter_k_iy_asya_ev_pos_2': 'halftime_asian_handicap_pos_2_home_close',
    'filter_k_iy_asya_dep_pos_2': 'halftime_asian_handicap_pos_2_away_close',
    'filter_k_ms_asya_ev_neg_2': 'fulltime_asian_handicap_neg2_home_close',
    'filter_k_ms_asya_dep_neg_2': 'fulltime_asian_handicap_neg2_away_close',
    'filter_k_ms_asya_ev_neg_2_neg_1_5': 'fulltime_asian_handicap_neg2_neg1_5_home_close',
    'filter_k_ms_asya_dep_neg_2_neg_1_5': 'fulltime_asian_handicap_neg2_neg1_5_away_close',
    'filter_k_ms_asya_ev_neg_1_5': 'fulltime_asian_handicap_neg1_5_home_close',
    'filter_k_ms_asya_dep_neg_1_5': 'fulltime_asian_handicap_neg1_5_away_close',
    'filter_k_ms_asya_ev_neg_1_5_neg_1': 'fulltime_asian_handicap_neg1_5_neg_1_home_close',
    'filter_k_ms_asya_dep_neg_1_5_neg_1': 'fulltime_asian_handicap_neg1_5_neg_1_away_close',
    'filter_k_ms_asya_ev_neg_1': 'fulltime_asian_handicap_neg_1_home_close',
    'filter_k_ms_asya_dep_neg_1': 'fulltime_asian_handicap_neg_1_away_close',
    'filter_k_ms_asya_ev_neg_1_neg_0_5': 'fulltime_asian_handicap_neg_1_neg_0_5_home_close',
    'filter_k_ms_asya_dep_neg_1_neg_0_5': 'fulltime_asian_handicap_neg_1_neg_0_5_away_close',
    'filter_k_ms_asya_ev_neg_0_5': 'fulltime_asian_handicap_neg_0_5_home_close',
    'filter_k_ms_asya_dep_neg_0_5': 'fulltime_asian_handicap_neg_0_5_away_close',
    'filter_k_ms_asya_ev_neg_0_5_pos_0': 'fulltime_asian_handicap_neg_0_5_0_0_home_close',
    'filter_k_ms_asya_dep_neg_0_5_pos_0': 'fulltime_asian_handicap_neg_0_5_0_0_away_close',
    'filter_k_ms_asya_ev_pos_0_pos_0_5': 'fulltime_asian_handicap_0_0_pos_0_5_home_close',
    'filter_k_ms_asya_dep_pos_0_pos_0_5': 'fulltime_asian_handicap_0_0_pos_0_5_away_close',
    'filter_k_ms_asya_ev_pos_0_5': 'fulltime_asian_handicap_pos_0_5_home_close',
    'filter_k_ms_asya_dep_pos_0_5': 'fulltime_asian_handicap_pos_0_5_away_close',
    'filter_k_ms_asya_ev_pos_0_5_pos_1': 'fulltime_asian_handicap_pos_0_5_pos_1_0_home_close',
    'filter_k_ms_asya_dep_pos_0_5_pos_1': 'fulltime_asian_handicap_pos_0_5_pos_1_0_away_close',
    'filter_k_ms_asya_ev_pos_1': 'fulltime_asian_handicap_pos_1_home_close',
    'filter_k_ms_asya_dep_pos_1': 'fulltime_asian_handicap_pos_1_away_close',
    'filter_k_ms_asya_ev_pos_1_pos_1_5': 'fulltime_asian_handicap_pos_1_0_pos_1_5_home_close',
    'filter_k_ms_asya_dep_pos_1_pos_1_5': 'fulltime_asian_handicap_pos_1_0_pos_1_5_away_close',
    'filter_k_ms_asya_ev_pos_1_5': 'fulltime_asian_handicap_pos_1_5_home_close',
    'filter_k_ms_asya_dep_pos_1_5': 'fulltime_asian_handicap_pos_1_5_away_close',
    'filter_k_ms_asya_ev_pos_1_5_pos_2': 'fulltime_asian_handicap_pos_1_5_pos_2_0_home_close',
    'filter_k_ms_asya_dep_pos_1_5_pos_2': 'fulltime_asian_handicap_pos_1_5_pos_2_0_away_close',
    'filter_k_ms_asya_ev_pos_2': 'fulltime_asian_handicap_pos_2_0_home_close',
    'filter_k_ms_asya_dep_pos_2': 'fulltime_asian_handicap_pos_2_0_away_close',
    'filter_k_iy_avrupa_ev_neg_1': 'firsthalf_european_handicap_neg_1_home_close',
    'filter_k_iy_avrupa_ber_neg_1': 'firsthalf_european_handicap_neg_1_tie_close',
    'filter_k_iy_avrupa_dep_neg_1': 'firsthalf_european_handicap_neg_1_away_close',
    'filter_k_iy_avrupa_ev_neg_1_neg_0_5': 'firsthalf_european_handicap_neg_1_neg_0_5_home_close',
    'filter_k_iy_avrupa_ber_neg_1_neg_0_5': 'firsthalf_european_handicap_neg_1_neg_0_5_tie_close',
    'filter_k_iy_avrupa_dep_neg_1_neg_0_5': 'firsthalf_european_handicap_neg_1_neg_0_5_away_close',
    'filter_k_iy_avrupa_ev_neg_0_5': 'firsthalf_european_handicap_neg_0_5_home_close',
    'filter_k_iy_avrupa_ber_neg_0_5': 'firsthalf_european_handicap_neg_0_5_tie_close',
    'filter_k_iy_avrupa_dep_neg_0_5': 'firsthalf_european_handicap_neg_0_5_away_close',
    'filter_k_iy_avrupa_ev_neg_0_5_pos_0': 'firsthalf_european_handicap_neg_0_5_0_0_home_close',
    'filter_k_iy_avrupa_ber_neg_0_5_pos_0': 'firsthalf_european_handicap_neg_0_5_0_0_tie_close',
    'filter_k_iy_avrupa_dep_neg_0_5_pos_0': 'firsthalf_european_handicap_neg_0_5_0_0_away_close',
    'filter_k_iy_avrupa_ev_pos_0_pos_0_5': 'firsthalf_european_handicap_0_0_pos_0_5_home_close',
    'filter_k_iy_avrupa_ber_pos_0_pos_0_5': 'firsthalf_european_handicap_0_0_pos_0_5_tie_close',
    'filter_k_iy_avrupa_dep_pos_0_pos_0_5': 'firsthalf_european_handicap_0_0_pos_0_5_away_close',
    'filter_k_iy_avrupa_ev_pos_0_5': 'firsthalf_european_handicap_pos_0_5_home_close',
    'filter_k_iy_avrupa_ber_pos_0_5': 'firsthalf_european_handicap_pos_0_5_tie_close',
    'filter_k_iy_avrupa_dep_pos_0_5': 'firsthalf_european_handicap_pos_0_5_away_close',
    'filter_k_iy_avrupa_ev_pos_0_5_pos_1': 'firsthalf_european_handicap_pos_0_5_pos_1_0_home_close',
    'filter_k_iy_avrupa_ber_pos_0_5_pos_1': 'firsthalf_european_handicap_pos_0_5_pos_1_0_tie_close',
    'filter_k_iy_avrupa_dep_pos_0_5_pos_1': 'firsthalf_european_handicap_pos_0_5_pos_1_0_away_close',
    'filter_k_iy_avrupa_ev_pos_1': 'firsthalf_european_handicap_pos_1_home_close',
    'filter_k_iy_avrupa_ber_pos_1': 'firsthalf_european_handicap_pos_1_tie_close',
    'filter_k_iy_avrupa_dep_pos_1': 'firsthalf_european_handicap_pos_1_away_close',
    'filter_k_ms_avrupa_ev_neg_1': 'fulltime_european_handicap_neg_1_home_close',
    'filter_k_ms_avrupa_ber_neg_1': 'fulltime_european_handicap_neg_1_tie_close',
    'filter_k_ms_avrupa_dep_neg_1': 'fulltime_european_handicap_neg_1_away_close',
    'filter_k_ms_avrupa_ev_neg_1_neg_0_5': 'fulltime_european_handicap_neg_1_neg_0_5_home_close',
    'filter_k_ms_avrupa_ber_neg_1_neg_0_5': 'fulltime_european_handicap_neg_1_neg_0_5_tie_close',
    'filter_k_ms_avrupa_dep_neg_1_neg_0_5': 'fulltime_european_handicap_neg_1_neg_0_5_away_close',
    'filter_k_ms_avrupa_ev_neg_0_5': 'fulltime_european_handicap_neg_0_5_home_close',
    'filter_k_ms_avrupa_ber_neg_0_5': 'fulltime_european_handicap_neg_0_5_tie_close',
    'filter_k_ms_avrupa_dep_neg_0_5': 'fulltime_european_handicap_neg_0_5_away_close',
    'filter_k_ms_avrupa_ev_neg_0_5_pos_0': 'fulltime_european_handicap_neg_0_5_0_0_home_close',
    'filter_k_ms_avrupa_ber_neg_0_5_pos_0': 'fulltime_european_handicap_neg_0_5_0_0_tie_close',
    'filter_k_ms_avrupa_dep_neg_0_5_pos_0': 'fulltime_european_handicap_neg_0_5_0_0_away_close',
    'filter_k_ms_avrupa_ev_pos_0_pos_0_5': 'fulltime_european_handicap_0_0_pos_0_5_home_close',
    'filter_k_ms_avrupa_ber_pos_0_pos_0_5': 'fulltime_european_handicap_0_0_pos_0_5_tie_close',
    'filter_k_ms_avrupa_dep_pos_0_pos_0_5': 'fulltime_european_handicap_0_0_pos_0_5_away_close',
    'filter_k_ms_avrupa_ev_pos_0_5': 'fulltime_european_handicap_pos_0_5_home_close',
    'filter_k_ms_avrupa_ber_pos_0_5': 'fulltime_european_handicap_pos_0_5_tie_close',
    'filter_k_ms_avrupa_dep_pos_0_5': 'fulltime_european_handicap_pos_0_5_away_close',
    'filter_k_ms_avrupa_ev_pos_0_5_pos_1': 'fulltime_european_handicap_pos_0_5_pos_1_0_home_close',
    'filter_k_ms_avrupa_ber_pos_0_5_pos_1': 'fulltime_european_handicap_pos_0_5_pos_1_0_tie_close',
    'filter_k_ms_avrupa_dep_pos_0_5_pos_1': 'fulltime_european_handicap_pos_0_5_pos_1_0_away_close',
    'filter_k_ms_avrupa_ev_pos_1': 'fulltime_european_handicap_pos_1_home_close',
    'filter_k_ms_avrupa_ber_pos_1': 'fulltime_european_handicap_pos_1_tie_close',
    'filter_k_ms_avrupa_dep_pos_1': 'fulltime_european_handicap_pos_1_away_close'
}

streamlit.write('## Oran Analiz Ekranı ##')

### ROW 1 CODE START ###
Row_1 = streamlit.columns([0.35, 0.30, 0.35])
with Row_1[1]:
    streamlit.date_input(label='Bülten Tarihi Seçiniz', value=datetime.today(), min_value=datetime.today() - relativedelta(days=7), max_value=datetime.today() + relativedelta(days=7), label_visibility="collapsed", key='bulletin_date')
### ROW 1 CODE END ###

### ROW 2 CODE START ###
Row_2 = streamlit.columns([0.14, 0.06, 0.60, 0.06, 0.14])
with Row_2[1]:
    streamlit.button(label='<', use_container_width=True, key='button_bulletin_previous')
with Row_2[3]:
    streamlit.button(label='\>', use_container_width=True, key='button_bulletin_next')

if streamlit.session_state.button_bulletin_previous:
    if streamlit.session_state.Bulletin_Match_Selected_Index > 0:
        streamlit.session_state.Bulletin_Match_Selected_Index -= 1

if streamlit.session_state.button_bulletin_next:
    if streamlit.session_state.Bulletin_Match_Selected_Index < streamlit.session_state.Bulletin_Match_Maximum_Index:
        streamlit.session_state.Bulletin_Match_Selected_Index += 1


def Bulletin_Match_Selected_Index_Changer():
    if len(streamlit.session_state.Bulletin_Match_List) > 0:
        Index = streamlit.session_state.Bulletin_Match_List.index(streamlit.session_state.bulletin_match_selectbox)
        streamlit.session_state.Bulletin_Match_Selected_Index = Index
    else:
        pass


with Row_2[2]:
    # Create a connection to the database
    connection = sqlite3.connect(Database_Path)
    cursor = connection.cursor()

    # Execute the query
    query = "SELECT * FROM bulletin_bet_data WHERE date=?"
    cursor.execute(query, (streamlit.session_state.bulletin_date.strftime('%d.%m.%Y'),))

    # Fetch all data
    data = cursor.fetchall()

    # Get column names from the cursor description
    columns = [col[0] for col in cursor.description]

    # Convert to Pandas DataFrame
    Temporal_DataFrame = pandas.DataFrame(data, columns=columns)

    # Initialize and populate the showing list
    streamlit.session_state.Bulletin_Match_List = []
    for index, row in Temporal_DataFrame.iterrows():
        streamlit.session_state.Bulletin_Match_List.append(f"{index + 1} | {row['hour']} | {row['home_team_name']} - {row['away_team_name']}")

    # Set the bulletin match dataframe
    streamlit.session_state.Bulletin_Match_DataFrame = Temporal_DataFrame

    # Get the maximum index
    streamlit.session_state.Bulletin_Match_Maximum_Index = len(Temporal_DataFrame) - 1

    try:
        streamlit.session_state.bulletin_match_selectbox = streamlit.session_state.Bulletin_Match_List[streamlit.session_state.Bulletin_Match_Selected_Index]
    except:
        pass

    if len(streamlit.session_state.Bulletin_Match_List) > 0:
        streamlit.selectbox(label='Maçlar Listesi', options=streamlit.session_state.Bulletin_Match_List, index=streamlit.session_state.Bulletin_Match_Selected_Index, label_visibility='collapsed', key='bulletin_match_selectbox', on_change=Bulletin_Match_Selected_Index_Changer)
    else:
        streamlit.selectbox(label='Maçlar Listesi', options=None, index=None, placeholder="İlgili güne ait bülten verisi bulunamadı!", label_visibility='collapsed', key='bulletin_match_selectbox', disabled=True)


### ROW 2 CODE END ###

def Load_Filters():
    # Veritabanı bağlantısı
    Connection = sqlite3.connect(Database_Path)

    # Bir cursor oluşturalım
    Cursor = Connection.cursor()

    # Sorgu
    Query = "SELECT * FROM saved_filters WHERE filter_name = ?"

    # Sorguyu parametre ile çalıştırma
    Cursor.execute(Query, (streamlit.session_state.selectbox_saved_filters,))

    # Veri alma
    Data = Cursor.fetchall()

    # Eğer veri geldiyse
    if Data:
        # Alınan satır
        Fetched_Data = Data[0]

        # Gerisi kolon işi
        streamlit.session_state.Filters_Dictionary['filter_start_date'] = datetime.strptime(Fetched_Data[1], '%Y-%m-%d')
        streamlit.session_state.Filters_Dictionary['filter_end_date'] = datetime.strptime(Fetched_Data[2], '%Y-%m-%d')
        streamlit.session_state.Filters_Dictionary['filter_country'] = Fetched_Data[3]
        streamlit.session_state.Filters_Dictionary['filter_league'] = Fetched_Data[4]
        streamlit.session_state.Filters_Dictionary['filter_week'] = Fetched_Data[5]
        streamlit.session_state.Filters_Dictionary['filter_hour'] = Fetched_Data[6]
        streamlit.session_state.Filters_Dictionary['filter_home_team'] = Fetched_Data[7]
        streamlit.session_state.Filters_Dictionary['filter_away_team'] = Fetched_Data[8]
        streamlit.session_state.Filters_Dictionary['filter_referee'] = Fetched_Data[9]
        streamlit.session_state.Filters_Dictionary['filter_a_ms_1'] = Fetched_Data[10]
        streamlit.session_state.Filters_Dictionary['filter_a_ms_x'] = Fetched_Data[11]
        streamlit.session_state.Filters_Dictionary['filter_a_ms_2'] = Fetched_Data[12]
        streamlit.session_state.Filters_Dictionary['filter_a_ms_0_5_a'] = Fetched_Data[13]
        streamlit.session_state.Filters_Dictionary['filter_a_ms_0_5_ü'] = Fetched_Data[14]
        streamlit.session_state.Filters_Dictionary['filter_a_ms_1_5_a'] = Fetched_Data[15]
        streamlit.session_state.Filters_Dictionary['filter_a_ms_1_5_ü'] = Fetched_Data[16]
        streamlit.session_state.Filters_Dictionary['filter_a_ms_2_5_a'] = Fetched_Data[17]
        streamlit.session_state.Filters_Dictionary['filter_a_ms_2_5_ü'] = Fetched_Data[18]
        streamlit.session_state.Filters_Dictionary['filter_a_ms_3_5_a'] = Fetched_Data[19]
        streamlit.session_state.Filters_Dictionary['filter_a_ms_3_5_ü'] = Fetched_Data[20]
        streamlit.session_state.Filters_Dictionary['filter_a_iy_1'] = Fetched_Data[21]
        streamlit.session_state.Filters_Dictionary['filter_a_iy_x'] = Fetched_Data[22]
        streamlit.session_state.Filters_Dictionary['filter_a_iy_2'] = Fetched_Data[23]
        streamlit.session_state.Filters_Dictionary['filter_a_iy_0_5_a'] = Fetched_Data[24]
        streamlit.session_state.Filters_Dictionary['filter_a_iy_0_5_ü'] = Fetched_Data[25]
        streamlit.session_state.Filters_Dictionary['filter_a_iy_1_5_a'] = Fetched_Data[26]
        streamlit.session_state.Filters_Dictionary['filter_a_iy_1_5_ü'] = Fetched_Data[27]
        streamlit.session_state.Filters_Dictionary['filter_a_ms_1_x_cs'] = Fetched_Data[28]
        streamlit.session_state.Filters_Dictionary['filter_a_ms_1_2_cs'] = Fetched_Data[29]
        streamlit.session_state.Filters_Dictionary['filter_a_ms_x_2_cs'] = Fetched_Data[30]
        streamlit.session_state.Filters_Dictionary['filter_a_iy_1_x_cs'] = Fetched_Data[31]
        streamlit.session_state.Filters_Dictionary['filter_a_iy_1_2_cs'] = Fetched_Data[32]
        streamlit.session_state.Filters_Dictionary['filter_a_iy_x_2_cs'] = Fetched_Data[33]
        streamlit.session_state.Filters_Dictionary['filter_a_home_win'] = Fetched_Data[34]
        streamlit.session_state.Filters_Dictionary['filter_a_away_win'] = Fetched_Data[35]
        streamlit.session_state.Filters_Dictionary['filter_a_iy_kg_var'] = Fetched_Data[36]
        streamlit.session_state.Filters_Dictionary['filter_a_iy_kg_yok'] = Fetched_Data[37]
        streamlit.session_state.Filters_Dictionary['filter_a_ms_kg_var'] = Fetched_Data[38]
        streamlit.session_state.Filters_Dictionary['filter_a_ms_kg_yok'] = Fetched_Data[39]
        streamlit.session_state.Filters_Dictionary['filter_a_tek'] = Fetched_Data[40]
        streamlit.session_state.Filters_Dictionary['filter_a_cift'] = Fetched_Data[41]
        streamlit.session_state.Filters_Dictionary['filter_a_1den1'] = Fetched_Data[42]
        streamlit.session_state.Filters_Dictionary['filter_a_1denx'] = Fetched_Data[43]
        streamlit.session_state.Filters_Dictionary['filter_a_1den2'] = Fetched_Data[44]
        streamlit.session_state.Filters_Dictionary['filter_a_xden1'] = Fetched_Data[45]
        streamlit.session_state.Filters_Dictionary['filter_a_xdenx'] = Fetched_Data[46]
        streamlit.session_state.Filters_Dictionary['filter_a_xden2'] = Fetched_Data[47]
        streamlit.session_state.Filters_Dictionary['filter_a_2den1'] = Fetched_Data[48]
        streamlit.session_state.Filters_Dictionary['filter_a_2denx'] = Fetched_Data[49]
        streamlit.session_state.Filters_Dictionary['filter_a_2den2'] = Fetched_Data[50]
        streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_10'] = Fetched_Data[51]
        streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_20'] = Fetched_Data[52]
        streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_21'] = Fetched_Data[53]
        streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_30'] = Fetched_Data[54]
        streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_31'] = Fetched_Data[55]
        streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_32'] = Fetched_Data[56]
        streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_40'] = Fetched_Data[57]
        streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_41'] = Fetched_Data[58]
        streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_42'] = Fetched_Data[59]
        streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_43'] = Fetched_Data[60]
        streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_00'] = Fetched_Data[61]
        streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_11'] = Fetched_Data[62]
        streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_22'] = Fetched_Data[63]
        streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_33'] = Fetched_Data[64]
        streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_44'] = Fetched_Data[65]
        streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_01'] = Fetched_Data[66]
        streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_02'] = Fetched_Data[67]
        streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_12'] = Fetched_Data[68]
        streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_03'] = Fetched_Data[69]
        streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_13'] = Fetched_Data[70]
        streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_23'] = Fetched_Data[71]
        streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_04'] = Fetched_Data[72]
        streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_14'] = Fetched_Data[73]
        streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_24'] = Fetched_Data[74]
        streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_34'] = Fetched_Data[75]
        streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_10'] = Fetched_Data[76]
        streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_20'] = Fetched_Data[77]
        streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_21'] = Fetched_Data[78]
        streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_30'] = Fetched_Data[79]
        streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_31'] = Fetched_Data[80]
        streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_32'] = Fetched_Data[81]
        streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_40'] = Fetched_Data[82]
        streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_41'] = Fetched_Data[83]
        streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_42'] = Fetched_Data[84]
        streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_43'] = Fetched_Data[85]
        streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_00'] = Fetched_Data[86]
        streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_11'] = Fetched_Data[87]
        streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_22'] = Fetched_Data[88]
        streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_33'] = Fetched_Data[89]
        streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_44'] = Fetched_Data[90]
        streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_01'] = Fetched_Data[91]
        streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_02'] = Fetched_Data[92]
        streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_12'] = Fetched_Data[93]
        streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_03'] = Fetched_Data[94]
        streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_13'] = Fetched_Data[95]
        streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_23'] = Fetched_Data[96]
        streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_04'] = Fetched_Data[97]
        streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_14'] = Fetched_Data[98]
        streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_24'] = Fetched_Data[99]
        streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_34'] = Fetched_Data[100]
        streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_neg_2'] = Fetched_Data[101]
        streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_neg_2'] = Fetched_Data[102]
        streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_neg_2_neg_1_5'] = Fetched_Data[103]
        streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_neg_2_neg_1_5'] = Fetched_Data[104]
        streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_neg_1_5'] = Fetched_Data[105]
        streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_neg_1_5'] = Fetched_Data[106]
        streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_neg_1_5_neg_1'] = Fetched_Data[107]
        streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_neg_1_5_neg_1'] = Fetched_Data[108]
        streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_neg_1'] = Fetched_Data[109]
        streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_neg_1'] = Fetched_Data[110]
        streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_neg_1_neg_0_5'] = Fetched_Data[111]
        streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_neg_1_neg_0_5'] = Fetched_Data[112]
        streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_neg_0_5'] = Fetched_Data[113]
        streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_neg_0_5'] = Fetched_Data[114]
        streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_neg_0_5_pos_0'] = Fetched_Data[115]
        streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_neg_0_5_pos_0'] = Fetched_Data[116]
        streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_pos_0_pos_0_5'] = Fetched_Data[117]
        streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_pos_0_pos_0_5'] = Fetched_Data[118]
        streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_pos_0_5'] = Fetched_Data[119]
        streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_pos_0_5'] = Fetched_Data[120]
        streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_pos_0_5_pos_1'] = Fetched_Data[121]
        streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_pos_0_5_pos_1'] = Fetched_Data[122]
        streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_pos_1'] = Fetched_Data[123]
        streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_pos_1'] = Fetched_Data[124]
        streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_pos_1_pos_1_5'] = Fetched_Data[125]
        streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_pos_1_pos_1_5'] = Fetched_Data[126]
        streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_pos_1_5'] = Fetched_Data[127]
        streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_pos_1_5'] = Fetched_Data[128]
        streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_pos_1_5_pos_2'] = Fetched_Data[129]
        streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_pos_1_5_pos_2'] = Fetched_Data[130]
        streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_pos_2'] = Fetched_Data[131]
        streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_pos_2'] = Fetched_Data[132]
        streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_neg_2'] = Fetched_Data[133]
        streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_neg_2'] = Fetched_Data[134]
        streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_neg_2_neg_1_5'] = Fetched_Data[135]
        streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_neg_2_neg_1_5'] = Fetched_Data[136]
        streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_neg_1_5'] = Fetched_Data[137]
        streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_neg_1_5'] = Fetched_Data[138]
        streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_neg_1_5_neg_1'] = Fetched_Data[139]
        streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_neg_1_5_neg_1'] = Fetched_Data[140]
        streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_neg_1'] = Fetched_Data[141]
        streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_neg_1'] = Fetched_Data[142]
        streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_neg_1_neg_0_5'] = Fetched_Data[143]
        streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_neg_1_neg_0_5'] = Fetched_Data[144]
        streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_neg_0_5'] = Fetched_Data[145]
        streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_neg_0_5'] = Fetched_Data[146]
        streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_neg_0_5_pos_0'] = Fetched_Data[147]
        streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_neg_0_5_pos_0'] = Fetched_Data[148]
        streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_pos_0_pos_0_5'] = Fetched_Data[149]
        streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_pos_0_pos_0_5'] = Fetched_Data[150]
        streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_pos_0_5'] = Fetched_Data[151]
        streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_pos_0_5'] = Fetched_Data[152]
        streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_pos_0_5_pos_1'] = Fetched_Data[153]
        streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_pos_0_5_pos_1'] = Fetched_Data[154]
        streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_pos_1'] = Fetched_Data[155]
        streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_pos_1'] = Fetched_Data[156]
        streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_pos_1_pos_1_5'] = Fetched_Data[157]
        streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_pos_1_pos_1_5'] = Fetched_Data[158]
        streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_pos_1_5'] = Fetched_Data[159]
        streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_pos_1_5'] = Fetched_Data[160]
        streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_pos_1_5_pos_2'] = Fetched_Data[161]
        streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_pos_1_5_pos_2'] = Fetched_Data[162]
        streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_pos_2'] = Fetched_Data[163]
        streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_pos_2'] = Fetched_Data[164]
        streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ev_neg_1'] = Fetched_Data[165]
        streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ber_neg_1'] = Fetched_Data[166]
        streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_dep_neg_1'] = Fetched_Data[167]
        streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ev_neg_1_neg_0_5'] = Fetched_Data[168]
        streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ber_neg_1_neg_0_5'] = Fetched_Data[169]
        streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_dep_neg_1_neg_0_5'] = Fetched_Data[170]
        streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ev_neg_0_5'] = Fetched_Data[171]
        streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ber_neg_0_5'] = Fetched_Data[172]
        streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_dep_neg_0_5'] = Fetched_Data[173]
        streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ev_neg_0_5_pos_0'] = Fetched_Data[174]
        streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ber_neg_0_5_pos_0'] = Fetched_Data[175]
        streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_dep_neg_0_5_pos_0'] = Fetched_Data[176]
        streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ev_pos_0_pos_0_5'] = Fetched_Data[177]
        streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ber_pos_0_pos_0_5'] = Fetched_Data[178]
        streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_dep_pos_0_pos_0_5'] = Fetched_Data[179]
        streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ev_pos_0_5'] = Fetched_Data[180]
        streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ber_pos_0_5'] = Fetched_Data[181]
        streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_dep_pos_0_5'] = Fetched_Data[182]
        streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ev_pos_0_5_pos_1'] = Fetched_Data[183]
        streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ber_pos_0_5_pos_1'] = Fetched_Data[184]
        streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_dep_pos_0_5_pos_1'] = Fetched_Data[185]
        streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ev_pos_1'] = Fetched_Data[186]
        streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ber_pos_1'] = Fetched_Data[187]
        streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_dep_pos_1'] = Fetched_Data[188]
        streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ev_neg_1'] = Fetched_Data[189]
        streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ber_neg_1'] = Fetched_Data[190]
        streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_dep_neg_1'] = Fetched_Data[191]
        streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ev_neg_1_neg_0_5'] = Fetched_Data[192]
        streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ber_neg_1_neg_0_5'] = Fetched_Data[193]
        streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_dep_neg_1_neg_0_5'] = Fetched_Data[194]
        streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ev_neg_0_5'] = Fetched_Data[195]
        streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ber_neg_0_5'] = Fetched_Data[196]
        streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_dep_neg_0_5'] = Fetched_Data[197]
        streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ev_neg_0_5_pos_0'] = Fetched_Data[198]
        streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ber_neg_0_5_pos_0'] = Fetched_Data[199]
        streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_dep_neg_0_5_pos_0'] = Fetched_Data[200]
        streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ev_pos_0_pos_0_5'] = Fetched_Data[201]
        streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ber_pos_0_pos_0_5'] = Fetched_Data[202]
        streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_dep_pos_0_pos_0_5'] = Fetched_Data[203]
        streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ev_pos_0_5'] = Fetched_Data[204]
        streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ber_pos_0_5'] = Fetched_Data[205]
        streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_dep_pos_0_5'] = Fetched_Data[206]
        streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ev_pos_0_5_pos_1'] = Fetched_Data[207]
        streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ber_pos_0_5_pos_1'] = Fetched_Data[208]
        streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_dep_pos_0_5_pos_1'] = Fetched_Data[209]
        streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ev_pos_1'] = Fetched_Data[210]
        streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ber_pos_1'] = Fetched_Data[211]
        streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_dep_pos_1'] = Fetched_Data[212]
        streamlit.session_state.Filters_Dictionary['filter_k_ms_1'] = Fetched_Data[213]
        streamlit.session_state.Filters_Dictionary['filter_k_ms_x'] = Fetched_Data[214]
        streamlit.session_state.Filters_Dictionary['filter_k_ms_2'] = Fetched_Data[215]
        streamlit.session_state.Filters_Dictionary['filter_k_ms_0_5_a'] = Fetched_Data[216]
        streamlit.session_state.Filters_Dictionary['filter_k_ms_0_5_ü'] = Fetched_Data[217]
        streamlit.session_state.Filters_Dictionary['filter_k_ms_1_5_a'] = Fetched_Data[218]
        streamlit.session_state.Filters_Dictionary['filter_k_ms_1_5_ü'] = Fetched_Data[219]
        streamlit.session_state.Filters_Dictionary['filter_k_ms_2_5_a'] = Fetched_Data[220]
        streamlit.session_state.Filters_Dictionary['filter_k_ms_2_5_ü'] = Fetched_Data[221]
        streamlit.session_state.Filters_Dictionary['filter_k_ms_3_5_a'] = Fetched_Data[222]
        streamlit.session_state.Filters_Dictionary['filter_k_ms_3_5_ü'] = Fetched_Data[223]
        streamlit.session_state.Filters_Dictionary['filter_k_iy_1'] = Fetched_Data[224]
        streamlit.session_state.Filters_Dictionary['filter_k_iy_x'] = Fetched_Data[225]
        streamlit.session_state.Filters_Dictionary['filter_k_iy_2'] = Fetched_Data[226]
        streamlit.session_state.Filters_Dictionary['filter_k_iy_0_5_a'] = Fetched_Data[227]
        streamlit.session_state.Filters_Dictionary['filter_k_iy_0_5_ü'] = Fetched_Data[228]
        streamlit.session_state.Filters_Dictionary['filter_k_iy_1_5_a'] = Fetched_Data[229]
        streamlit.session_state.Filters_Dictionary['filter_k_iy_1_5_ü'] = Fetched_Data[230]
        streamlit.session_state.Filters_Dictionary['filter_k_ms_1_x_cs'] = Fetched_Data[231]
        streamlit.session_state.Filters_Dictionary['filter_k_ms_1_2_cs'] = Fetched_Data[232]
        streamlit.session_state.Filters_Dictionary['filter_k_ms_x_2_cs'] = Fetched_Data[233]
        streamlit.session_state.Filters_Dictionary['filter_k_iy_1_x_cs'] = Fetched_Data[234]
        streamlit.session_state.Filters_Dictionary['filter_k_iy_1_2_cs'] = Fetched_Data[235]
        streamlit.session_state.Filters_Dictionary['filter_k_iy_x_2_cs'] = Fetched_Data[236]
        streamlit.session_state.Filters_Dictionary['filter_k_home_win'] = Fetched_Data[237]
        streamlit.session_state.Filters_Dictionary['filter_k_away_win'] = Fetched_Data[238]
        streamlit.session_state.Filters_Dictionary['filter_k_iy_kg_var'] = Fetched_Data[239]
        streamlit.session_state.Filters_Dictionary['filter_k_iy_kg_yok'] = Fetched_Data[240]
        streamlit.session_state.Filters_Dictionary['filter_k_ms_kg_var'] = Fetched_Data[241]
        streamlit.session_state.Filters_Dictionary['filter_k_ms_kg_yok'] = Fetched_Data[242]
        streamlit.session_state.Filters_Dictionary['filter_k_tek'] = Fetched_Data[243]
        streamlit.session_state.Filters_Dictionary['filter_k_cift'] = Fetched_Data[244]
        streamlit.session_state.Filters_Dictionary['filter_k_1den1'] = Fetched_Data[245]
        streamlit.session_state.Filters_Dictionary['filter_k_1denx'] = Fetched_Data[246]
        streamlit.session_state.Filters_Dictionary['filter_k_1den2'] = Fetched_Data[247]
        streamlit.session_state.Filters_Dictionary['filter_k_xden1'] = Fetched_Data[248]
        streamlit.session_state.Filters_Dictionary['filter_k_xdenx'] = Fetched_Data[249]
        streamlit.session_state.Filters_Dictionary['filter_k_xden2'] = Fetched_Data[250]
        streamlit.session_state.Filters_Dictionary['filter_k_2den1'] = Fetched_Data[251]
        streamlit.session_state.Filters_Dictionary['filter_k_2denx'] = Fetched_Data[252]
        streamlit.session_state.Filters_Dictionary['filter_k_2den2'] = Fetched_Data[253]
        streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_10'] = Fetched_Data[254]
        streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_20'] = Fetched_Data[255]
        streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_21'] = Fetched_Data[256]
        streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_30'] = Fetched_Data[257]
        streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_31'] = Fetched_Data[258]
        streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_32'] = Fetched_Data[259]
        streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_40'] = Fetched_Data[260]
        streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_41'] = Fetched_Data[261]
        streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_42'] = Fetched_Data[262]
        streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_43'] = Fetched_Data[263]
        streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_00'] = Fetched_Data[264]
        streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_11'] = Fetched_Data[265]
        streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_22'] = Fetched_Data[266]
        streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_33'] = Fetched_Data[267]
        streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_44'] = Fetched_Data[268]
        streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_01'] = Fetched_Data[269]
        streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_02'] = Fetched_Data[270]
        streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_12'] = Fetched_Data[271]
        streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_03'] = Fetched_Data[272]
        streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_13'] = Fetched_Data[273]
        streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_23'] = Fetched_Data[274]
        streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_04'] = Fetched_Data[275]
        streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_14'] = Fetched_Data[276]
        streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_24'] = Fetched_Data[277]
        streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_34'] = Fetched_Data[278]
        streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_10'] = Fetched_Data[279]
        streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_20'] = Fetched_Data[280]
        streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_21'] = Fetched_Data[281]
        streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_30'] = Fetched_Data[282]
        streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_31'] = Fetched_Data[283]
        streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_32'] = Fetched_Data[284]
        streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_40'] = Fetched_Data[285]
        streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_41'] = Fetched_Data[286]
        streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_42'] = Fetched_Data[287]
        streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_43'] = Fetched_Data[288]
        streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_00'] = Fetched_Data[289]
        streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_11'] = Fetched_Data[290]
        streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_22'] = Fetched_Data[291]
        streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_33'] = Fetched_Data[292]
        streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_44'] = Fetched_Data[293]
        streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_01'] = Fetched_Data[294]
        streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_02'] = Fetched_Data[295]
        streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_12'] = Fetched_Data[296]
        streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_03'] = Fetched_Data[297]
        streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_13'] = Fetched_Data[298]
        streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_23'] = Fetched_Data[299]
        streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_04'] = Fetched_Data[300]
        streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_14'] = Fetched_Data[301]
        streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_24'] = Fetched_Data[302]
        streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_34'] = Fetched_Data[303]
        streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_neg_2'] = Fetched_Data[304]
        streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_neg_2'] = Fetched_Data[305]
        streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_neg_2_neg_1_5'] = Fetched_Data[306]
        streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_neg_2_neg_1_5'] = Fetched_Data[307]
        streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_neg_1_5'] = Fetched_Data[308]
        streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_neg_1_5'] = Fetched_Data[309]
        streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_neg_1_5_neg_1'] = Fetched_Data[310]
        streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_neg_1_5_neg_1'] = Fetched_Data[311]
        streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_neg_1'] = Fetched_Data[312]
        streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_neg_1'] = Fetched_Data[313]
        streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_neg_1_neg_0_5'] = Fetched_Data[314]
        streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_neg_1_neg_0_5'] = Fetched_Data[315]
        streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_neg_0_5'] = Fetched_Data[316]
        streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_neg_0_5'] = Fetched_Data[317]
        streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_neg_0_5_pos_0'] = Fetched_Data[318]
        streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_neg_0_5_pos_0'] = Fetched_Data[319]
        streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_pos_0_pos_0_5'] = Fetched_Data[320]
        streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_pos_0_pos_0_5'] = Fetched_Data[321]
        streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_pos_0_5'] = Fetched_Data[322]
        streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_pos_0_5'] = Fetched_Data[323]
        streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_pos_0_5_pos_1'] = Fetched_Data[324]
        streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_pos_0_5_pos_1'] = Fetched_Data[325]
        streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_pos_1'] = Fetched_Data[326]
        streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_pos_1'] = Fetched_Data[327]
        streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_pos_1_pos_1_5'] = Fetched_Data[328]
        streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_pos_1_pos_1_5'] = Fetched_Data[329]
        streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_pos_1_5'] = Fetched_Data[330]
        streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_pos_1_5'] = Fetched_Data[331]
        streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_pos_1_5_pos_2'] = Fetched_Data[332]
        streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_pos_1_5_pos_2'] = Fetched_Data[333]
        streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_pos_2'] = Fetched_Data[334]
        streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_pos_2'] = Fetched_Data[335]
        streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_neg_2'] = Fetched_Data[336]
        streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_neg_2'] = Fetched_Data[337]
        streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_neg_2_neg_1_5'] = Fetched_Data[338]
        streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_neg_2_neg_1_5'] = Fetched_Data[339]
        streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_neg_1_5'] = Fetched_Data[340]
        streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_neg_1_5'] = Fetched_Data[341]
        streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_neg_1_5_neg_1'] = Fetched_Data[342]
        streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_neg_1_5_neg_1'] = Fetched_Data[343]
        streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_neg_1'] = Fetched_Data[344]
        streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_neg_1'] = Fetched_Data[345]
        streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_neg_1_neg_0_5'] = Fetched_Data[346]
        streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_neg_1_neg_0_5'] = Fetched_Data[347]
        streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_neg_0_5'] = Fetched_Data[348]
        streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_neg_0_5'] = Fetched_Data[349]
        streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_neg_0_5_pos_0'] = Fetched_Data[350]
        streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_neg_0_5_pos_0'] = Fetched_Data[351]
        streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_pos_0_pos_0_5'] = Fetched_Data[352]
        streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_pos_0_pos_0_5'] = Fetched_Data[353]
        streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_pos_0_5'] = Fetched_Data[354]
        streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_pos_0_5'] = Fetched_Data[355]
        streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_pos_0_5_pos_1'] = Fetched_Data[356]
        streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_pos_0_5_pos_1'] = Fetched_Data[357]
        streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_pos_1'] = Fetched_Data[358]
        streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_pos_1'] = Fetched_Data[359]
        streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_pos_1_pos_1_5'] = Fetched_Data[360]
        streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_pos_1_pos_1_5'] = Fetched_Data[361]
        streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_pos_1_5'] = Fetched_Data[362]
        streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_pos_1_5'] = Fetched_Data[363]
        streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_pos_1_5_pos_2'] = Fetched_Data[364]
        streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_pos_1_5_pos_2'] = Fetched_Data[365]
        streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_pos_2'] = Fetched_Data[366]
        streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_pos_2'] = Fetched_Data[367]
        streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ev_neg_1'] = Fetched_Data[368]
        streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ber_neg_1'] = Fetched_Data[369]
        streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_dep_neg_1'] = Fetched_Data[370]
        streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ev_neg_1_neg_0_5'] = Fetched_Data[371]
        streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ber_neg_1_neg_0_5'] = Fetched_Data[372]
        streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_dep_neg_1_neg_0_5'] = Fetched_Data[373]
        streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ev_neg_0_5'] = Fetched_Data[374]
        streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ber_neg_0_5'] = Fetched_Data[375]
        streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_dep_neg_0_5'] = Fetched_Data[376]
        streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ev_neg_0_5_pos_0'] = Fetched_Data[377]
        streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ber_neg_0_5_pos_0'] = Fetched_Data[378]
        streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_dep_neg_0_5_pos_0'] = Fetched_Data[379]
        streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ev_pos_0_pos_0_5'] = Fetched_Data[380]
        streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ber_pos_0_pos_0_5'] = Fetched_Data[381]
        streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_dep_pos_0_pos_0_5'] = Fetched_Data[382]
        streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ev_pos_0_5'] = Fetched_Data[383]
        streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ber_pos_0_5'] = Fetched_Data[384]
        streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_dep_pos_0_5'] = Fetched_Data[385]
        streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ev_pos_0_5_pos_1'] = Fetched_Data[386]
        streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ber_pos_0_5_pos_1'] = Fetched_Data[387]
        streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_dep_pos_0_5_pos_1'] = Fetched_Data[388]
        streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ev_pos_1'] = Fetched_Data[389]
        streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ber_pos_1'] = Fetched_Data[390]
        streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_dep_pos_1'] = Fetched_Data[391]
        streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ev_neg_1'] = Fetched_Data[392]
        streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ber_neg_1'] = Fetched_Data[393]
        streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_dep_neg_1'] = Fetched_Data[394]
        streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ev_neg_1_neg_0_5'] = Fetched_Data[395]
        streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ber_neg_1_neg_0_5'] = Fetched_Data[396]
        streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_dep_neg_1_neg_0_5'] = Fetched_Data[397]
        streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ev_neg_0_5'] = Fetched_Data[398]
        streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ber_neg_0_5'] = Fetched_Data[399]
        streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_dep_neg_0_5'] = Fetched_Data[400]
        streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ev_neg_0_5_pos_0'] = Fetched_Data[401]
        streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ber_neg_0_5_pos_0'] = Fetched_Data[402]
        streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_dep_neg_0_5_pos_0'] = Fetched_Data[403]
        streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ev_pos_0_pos_0_5'] = Fetched_Data[404]
        streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ber_pos_0_pos_0_5'] = Fetched_Data[405]
        streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_dep_pos_0_pos_0_5'] = Fetched_Data[406]
        streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ev_pos_0_5'] = Fetched_Data[407]
        streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ber_pos_0_5'] = Fetched_Data[408]
        streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_dep_pos_0_5'] = Fetched_Data[409]
        streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ev_pos_0_5_pos_1'] = Fetched_Data[410]
        streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ber_pos_0_5_pos_1'] = Fetched_Data[411]
        streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_dep_pos_0_5_pos_1'] = Fetched_Data[412]
        streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ev_pos_1'] = Fetched_Data[413]
        streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ber_pos_1'] = Fetched_Data[414]
        streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_dep_pos_1'] = Fetched_Data[415]

    # Bağlantıyı kapat
    Connection.close()


### ROW 3 CODE START ###
Row_3 = streamlit.columns([0.14, 0.72, 0.14])
with Row_3[1]:
    if len(streamlit.session_state.Bulletin_Match_List) == 0:
        streamlit.selectbox(label='Filtre Listesi', options=None, index=None, placeholder="Kayıtlı Filtre Seçiniz...", label_visibility='collapsed', disabled=True)
    else:
        # Create a connection to the database
        connection = sqlite3.connect(Database_Path)
        cursor = connection.cursor()

        # Execute the query
        query = "SELECT * FROM saved_filters"
        cursor.execute(query, )

        # Fetch all data
        data = cursor.fetchall()

        # Get column names from the cursor description
        columns = [col[0] for col in cursor.description]

        # Convert to Pandas DataFrame
        Saved_Filters_DataFrame = pandas.DataFrame(data, columns=columns)

        # Get filters by name
        Saved_Filters_List = Saved_Filters_DataFrame['filter_name'].to_list()

        # Filtre selectbox
        streamlit.selectbox(label='Filtre Listesi', options=Saved_Filters_List, index=None, placeholder="Kayıtlı Filtre Seçiniz...", label_visibility='collapsed', key="selectbox_saved_filters", on_change=Load_Filters)
### ROW 3 CODE END ###

### ROW 4 CODE START ###
Row_4 = streamlit.columns([0.14, 0.60, 0.06, 0.06, 0.14])
with Row_4[1]:
    Loaded_Filter_Text = 'Filtre Seçiniz'
    if len(streamlit.session_state.Bulletin_Match_List) == 0:
        with streamlit.popover(label=Loaded_Filter_Text, use_container_width=True, disabled=True):
            tab1, tab2, tab3, tab4 = streamlit.tabs(['Veritabanı', 'Genel Filtre', 'Açılış', 'Kapanış'])
    elif len(streamlit.session_state.Bulletin_Match_List) > 0:
        # Selected Match Data
        MatchData_DataFrame = streamlit.session_state.Bulletin_Match_DataFrame.iloc[streamlit.session_state.Bulletin_Match_Selected_Index, :]
        MatchData_DataFrame = MatchData_DataFrame.fillna("-")

        with streamlit.popover(label=Loaded_Filter_Text, use_container_width=True):
            tab1, tab2, tab3, tab4 = streamlit.tabs(['Veritabanı', 'Genel Filtre', 'Açılış', 'Kapanış'])

            # General filter
            Ülke = MatchData_DataFrame['country']
            Lig = MatchData_DataFrame['league']
            Hafta = MatchData_DataFrame['round']
            Saat = MatchData_DataFrame['hour']
            Ev = MatchData_DataFrame['home_team_name']
            Dep = MatchData_DataFrame['away_team_name']
            Hakem = MatchData_DataFrame['referee']

            # OPEN ODDS START
            # Fulltime
            Fulltime_Home_Open = MatchData_DataFrame['fulltime_home_open']
            Fulltime_Tie_Open = MatchData_DataFrame['fulltime_tie_open']
            Fulltime_Away_Open = MatchData_DataFrame['fulltime_home_open']

            # Firsthalf
            Firsthalf_Home_Open = MatchData_DataFrame['firsthalf_home_open']
            Firsthalf_Tie_Open = MatchData_DataFrame['firsthalf_tie_open']
            Firsthalf_Away_Open = MatchData_DataFrame['firsthalf_away_open']

            # Firsthalf-Fulltime
            Home_Home_Open = MatchData_DataFrame['home_home_open']
            Home_Tie_Open = MatchData_DataFrame['home_tie_open']
            Home_Away_Open = MatchData_DataFrame['home_away_open']
            Tie_Home_Open = MatchData_DataFrame['tie_home_open']
            Tie_Tie_Open = MatchData_DataFrame['tie_tie_open']
            Tie_Away_Open = MatchData_DataFrame['tie_away_open']
            Away_Home_Open = MatchData_DataFrame['away_home_open']
            Away_Tie_Open = MatchData_DataFrame['away_tie_open']
            Away_Away_Open = MatchData_DataFrame['away_away_open']

            # Firsthalf Exact Score
            Firsthalf_1_0_Open = MatchData_DataFrame['first_half_1_0_open']
            Firsthalf_2_0_Open = MatchData_DataFrame['first_half_2_0_open']
            Firsthalf_2_1_Open = MatchData_DataFrame['first_half_2_1_open']
            Firsthalf_3_0_Open = MatchData_DataFrame['first_half_3_0_open']
            Firsthalf_3_1_Open = MatchData_DataFrame['first_half_3_1_open']
            Firsthalf_3_2_Open = MatchData_DataFrame['first_half_3_2_open']
            Firsthalf_4_0_Open = MatchData_DataFrame['first_half_4_0_open']
            Firsthalf_4_1_Open = MatchData_DataFrame['first_half_4_1_open']
            Firsthalf_4_2_Open = MatchData_DataFrame['first_half_4_2_open']
            Firsthalf_4_3_Open = MatchData_DataFrame['first_half_4_3_open']
            Firsthalf_0_0_Open = MatchData_DataFrame['first_half_0_0_open']
            Firsthalf_1_1_Open = MatchData_DataFrame['first_half_1_1_open']
            Firsthalf_2_2_Open = MatchData_DataFrame['first_half_2_2_open']
            Firsthalf_3_3_Open = MatchData_DataFrame['first_half_3_3_open']
            Firsthalf_4_4_Open = MatchData_DataFrame['first_half_4_4_open']
            Firsthalf_0_1_Open = MatchData_DataFrame['first_half_0_1_open']
            Firsthalf_0_2_Open = MatchData_DataFrame['first_half_0_2_open']
            Firsthalf_1_2_Open = MatchData_DataFrame['first_half_1_2_open']
            Firsthalf_0_3_Open = MatchData_DataFrame['first_half_0_3_open']
            Firsthalf_1_3_Open = MatchData_DataFrame['first_half_1_3_open']
            Firsthalf_2_3_Open = MatchData_DataFrame['first_half_2_3_open']
            Firsthalf_0_4_Open = MatchData_DataFrame['first_half_0_4_open']
            Firsthalf_1_4_Open = MatchData_DataFrame['first_half_1_4_open']
            Firsthalf_2_4_Open = MatchData_DataFrame['first_half_2_4_open']
            Firsthalf_3_4_Open = MatchData_DataFrame['first_half_3_4_open']

            # Matchtime Exact Score
            Matchtime_1_0_Open = MatchData_DataFrame['match_time_1_0_open']
            Matchtime_2_0_Open = MatchData_DataFrame['match_time_2_0_open']
            Matchtime_2_1_Open = MatchData_DataFrame['match_time_2_1_open']
            Matchtime_3_0_Open = MatchData_DataFrame['match_time_3_0_open']
            Matchtime_3_1_Open = MatchData_DataFrame['match_time_3_1_open']
            Matchtime_3_2_Open = MatchData_DataFrame['match_time_3_2_open']
            Matchtime_4_0_Open = MatchData_DataFrame['match_time_4_0_open']
            Matchtime_4_1_Open = MatchData_DataFrame['match_time_4_1_open']
            Matchtime_4_2_Open = MatchData_DataFrame['match_time_4_2_open']
            Matchtime_4_3_Open = MatchData_DataFrame['match_time_4_3_open']
            Matchtime_0_0_Open = MatchData_DataFrame['match_time_0_0_open']
            Matchtime_1_1_Open = MatchData_DataFrame['match_time_1_1_open']
            Matchtime_2_2_Open = MatchData_DataFrame['match_time_2_2_open']
            Matchtime_3_3_Open = MatchData_DataFrame['match_time_3_3_open']
            Matchtime_4_4_Open = MatchData_DataFrame['match_time_4_4_open']
            Matchtime_0_1_Open = MatchData_DataFrame['match_time_0_1_open']
            Matchtime_0_2_Open = MatchData_DataFrame['match_time_0_2_open']
            Matchtime_1_2_Open = MatchData_DataFrame['match_time_1_2_open']
            Matchtime_0_3_Open = MatchData_DataFrame['match_time_0_3_open']
            Matchtime_1_3_Open = MatchData_DataFrame['match_time_1_3_open']
            Matchtime_2_3_Open = MatchData_DataFrame['match_time_2_3_open']
            Matchtime_0_4_Open = MatchData_DataFrame['match_time_0_4_open']
            Matchtime_1_4_Open = MatchData_DataFrame['match_time_1_4_open']
            Matchtime_2_4_Open = MatchData_DataFrame['match_time_2_4_open']
            Matchtime_3_4_Open = MatchData_DataFrame['match_time_3_4_open']

            # Double Chance
            Firsthalf_Home_And_Tie_Open = MatchData_DataFrame['first_half_home_and_tie_open']
            Firsthalf_Home_And_Away_Open = MatchData_DataFrame['first_half_home_and_away_open']
            Firsthalf_Tie_And_Away_Open = MatchData_DataFrame['first_half_tie_and_away_open']
            Matchtime_Home_And_Tie_Open = MatchData_DataFrame['match_time_home_and_tie_open']
            Matchtime_Home_And_Away_Open = MatchData_DataFrame['match_time_home_and_away_open']
            Matchtime_Tie_And_Away_Open = MatchData_DataFrame['match_time_tie_and_away_open']

            # Firsthalf Over/Under
            Firsthalf_Over_0_5_Open = MatchData_DataFrame['first_half_over_0_5_open']
            Firsthalf_Under_0_5_Open = MatchData_DataFrame['first_half_under_0_5_open']
            Firsthalf_Over_1_5_Open = MatchData_DataFrame['first_half_over_1_5_open']
            Firsthalf_Under_1_5_Open = MatchData_DataFrame['first_half_under_1_5_open']

            # Matchtime Over/Under
            Matchtime_Over_0_5_Open = MatchData_DataFrame['match_time_over_0_5_open']
            Matchtime_Under_0_5_Open = MatchData_DataFrame['match_time_under_0_5_open']
            Matchtime_Over_1_5_Open = MatchData_DataFrame['match_time_over_1_5_open']
            Matchtime_Under_1_5_Open = MatchData_DataFrame['match_time_under_1_5_open']
            Matchtime_Over_2_5_Open = MatchData_DataFrame['match_time_over_2_5_open']
            Matchtime_Under_2_5_Open = MatchData_DataFrame['match_time_under_2_5_open']
            Matchtime_Over_3_5_Open = MatchData_DataFrame['match_time_over_3_5_open']
            Matchtime_Under_3_5_Open = MatchData_DataFrame['match_time_under_3_5_open']

            # Draw No Bet
            Draw_No_Bet_Home_Open = MatchData_DataFrame['draw_no_bet_home_open']
            Draw_No_Bet_Away_Open = MatchData_DataFrame['draw_no_bet_away_open']

            # Both Teams To Score
            Firsthalf_Both_Teams_To_Score_Yes_Open = MatchData_DataFrame['firsthalf_both_teams_to_score_yes_open']
            Firsthalf_Both_Teams_To_Score_No_Open = MatchData_DataFrame['firsthalf_both_teams_to_score_no_open']
            Matchtime_Both_Teams_To_Score_Yes_Open = MatchData_DataFrame['both_teams_to_score_yes_open']
            Matchtime_Both_Teams_To_Score_No_Open = MatchData_DataFrame['both_teams_to_score_no_open']

            # Odd/Even
            Odd_Open = MatchData_DataFrame['odd_open']
            Even_Open = MatchData_DataFrame['even_open']

            # Firsthalf Asian Handicap
            Firsthalf_Asian_Neg2_Home_Open = MatchData_DataFrame['halftime_asian_handicap_neg2_home_open']
            Firsthalf_Asian_Neg2_Away_Open = MatchData_DataFrame['halftime_asian_handicap_neg2_away_open']
            Firsthalf_Asian_Neg2_Neg1_5_Home_Open = MatchData_DataFrame['halftime_asian_handicap_neg2_neg1_5_home_open']
            Firsthalf_Asian_Neg2_Neg1_5_Away_Open = MatchData_DataFrame['halftime_asian_handicap_neg2_neg1_5_away_open']
            Firsthalf_Asian_Neg1_5_Home_Open = MatchData_DataFrame['halftime_asian_handicap_neg1_5_home_open']
            Firsthalf_Asian_Neg1_5_Away_Open = MatchData_DataFrame['halftime_asian_handicap_neg1_5_away_open']
            Firsthalf_Asian_Neg1_5_Neg1_Home_Open = MatchData_DataFrame['halftime_asian_handicap_neg1_5_neg_1_home_open']
            Firsthalf_Asian_Neg1_5_Neg1_Away_Open = MatchData_DataFrame['halftime_asian_handicap_neg1_5_neg_1_away_open']
            Firsthalf_Asian_Neg1_Home_Open = MatchData_DataFrame['halftime_asian_handicap_neg_1_home_open']
            Firsthalf_Asian_Neg1_Away_Open = MatchData_DataFrame['halftime_asian_handicap_neg_1_away_open']
            Firsthalf_Asian_Neg1_Neg0_5_Home_Open = MatchData_DataFrame['halftime_asian_handicap_neg_1_neg_0_5_home_open']
            Firsthalf_Asian_Neg1_Neg0_5_Away_Open = MatchData_DataFrame['halftime_asian_handicap_neg_1_neg_0_5_away_open']
            Firsthalf_Asian_Neg0_5_Home_Open = MatchData_DataFrame['halftime_asian_handicap_neg_0_5_home_open']
            Firsthalf_Asian_Neg0_5_Away_Open = MatchData_DataFrame['halftime_asian_handicap_neg_0_5_away_open']
            Firsthalf_Asian_Neg0_5_Neg0_0_Home_Open = MatchData_DataFrame['halftime_asian_handicap_neg_0_5_0_0_home_open']
            Firsthalf_Asian_Neg0_5_Neg0_0_Away_Open = MatchData_DataFrame['halftime_asian_handicap_neg_0_5_0_0_away_open']
            Firsthalf_Asian_0_0_Pos0_5_Home_Open = MatchData_DataFrame['halftime_asian_handicap_0_0_pos_0_5_home_open']
            Firsthalf_Asian_0_0_Pos0_5_Away_Open = MatchData_DataFrame['halftime_asian_handicap_0_0_pos_0_5_away_open']
            Firsthalf_Asian_Pos0_5_Home_Open = MatchData_DataFrame['halftime_asian_handicap_pos_0_5_home_open']
            Firsthalf_Asian_Pos0_5_Away_Open = MatchData_DataFrame['halftime_asian_handicap_pos_0_5_away_open']
            Firsthalf_Asian_Pos0_5_Pos1_0_Home_Open = MatchData_DataFrame['halftime_asian_handicap_pos_0_5_pos_1_0_home_open']
            Firsthalf_Asian_Pos0_5_Pos1_0_Away_Open = MatchData_DataFrame['halftime_asian_handicap_pos_0_5_pos_1_0_away_open']
            Firsthalf_Asian_Pos1_Home_Open = MatchData_DataFrame['halftime_asian_handicap_pos_1_home_open']
            Firsthalf_Asian_Pos1_Away_Open = MatchData_DataFrame['halftime_asian_handicap_pos_1_away_open']
            Firsthalf_Asian_Pos1_Pos1_5_Home_Open = MatchData_DataFrame['halftime_asian_handicap_pos_1_0_pos_1_5_home_open']
            Firsthalf_Asian_Pos1_Pos1_5_Away_Open = MatchData_DataFrame['halftime_asian_handicap_pos_1_0_pos_1_5_away_open']
            Firsthalf_Asian_Pos1_5_Home_Open = MatchData_DataFrame['halftime_asian_handicap_pos_1_5_home_open']
            Firsthalf_Asian_Pos1_5_Away_Open = MatchData_DataFrame['halftime_asian_handicap_pos_1_5_away_open']
            Firsthalf_Asian_Pos1_5_Pos2_Home_Open = MatchData_DataFrame['halftime_asian_handicap_pos_1_5_pos_2_0_home_open']
            Firsthalf_Asian_Pos1_5_Pos2_Away_Open = MatchData_DataFrame['halftime_asian_handicap_pos_1_5_pos_2_0_away_open']
            Firsthalf_Asian_Pos2_Home_Open = MatchData_DataFrame['halftime_asian_handicap_pos_2_home_open']
            Firsthalf_Asian_Pos2_Away_Open = MatchData_DataFrame['halftime_asian_handicap_pos_2_away_open']

            # Matchtime Asian Handicap
            Matchtime_Asian_Neg2_Home_Open = MatchData_DataFrame['fulltime_asian_handicap_neg2_home_open']
            Matchtime_Asian_Neg2_Away_Open = MatchData_DataFrame['fulltime_asian_handicap_neg2_away_open']
            Matchtime_Asian_Neg2_Neg1_5_Home_Open = MatchData_DataFrame['fulltime_asian_handicap_neg2_neg1_5_home_open']
            Matchtime_Asian_Neg2_Neg1_5_Away_Open = MatchData_DataFrame['fulltime_asian_handicap_neg2_neg1_5_away_open']
            Matchtime_Asian_Neg1_5_Home_Open = MatchData_DataFrame['fulltime_asian_handicap_neg1_5_home_open']
            Matchtime_Asian_Neg1_5_Away_Open = MatchData_DataFrame['fulltime_asian_handicap_neg1_5_away_open']
            Matchtime_Asian_Neg1_5_Neg1_Home_Open = MatchData_DataFrame['fulltime_asian_handicap_neg1_5_neg_1_home_open']
            Matchtime_Asian_Neg1_5_Neg1_Away_Open = MatchData_DataFrame['fulltime_asian_handicap_neg1_5_neg_1_away_open']
            Matchtime_Asian_Neg1_Home_Open = MatchData_DataFrame['fulltime_asian_handicap_neg_1_home_open']
            Matchtime_Asian_Neg1_Away_Open = MatchData_DataFrame['fulltime_asian_handicap_neg_1_away_open']
            Matchtime_Asian_Neg1_Neg0_5_Home_Open = MatchData_DataFrame['fulltime_asian_handicap_neg_1_neg_0_5_home_open']
            Matchtime_Asian_Neg1_Neg0_5_Away_Open = MatchData_DataFrame['fulltime_asian_handicap_neg_1_neg_0_5_away_open']
            Matchtime_Asian_Neg0_5_Home_Open = MatchData_DataFrame['fulltime_asian_handicap_neg_0_5_home_open']
            Matchtime_Asian_Neg0_5_Away_Open = MatchData_DataFrame['fulltime_asian_handicap_neg_0_5_away_open']
            Matchtime_Asian_Neg0_5_Neg0_0_Home_Open = MatchData_DataFrame['fulltime_asian_handicap_neg_0_5_0_0_home_open']
            Matchtime_Asian_Neg0_5_Neg0_0_Away_Open = MatchData_DataFrame['fulltime_asian_handicap_neg_0_5_0_0_away_open']
            Matchtime_Asian_0_0_Pos0_5_Home_Open = MatchData_DataFrame['fulltime_asian_handicap_0_0_pos_0_5_home_open']
            Matchtime_Asian_0_0_Pos0_5_Away_Open = MatchData_DataFrame['fulltime_asian_handicap_0_0_pos_0_5_away_open']
            Matchtime_Asian_Pos0_5_Home_Open = MatchData_DataFrame['fulltime_asian_handicap_pos_0_5_home_open']
            Matchtime_Asian_Pos0_5_Away_Open = MatchData_DataFrame['fulltime_asian_handicap_pos_0_5_away_open']
            Matchtime_Asian_Pos0_5_Pos1_0_Home_Open = MatchData_DataFrame['fulltime_asian_handicap_pos_0_5_pos_1_0_home_open']
            Matchtime_Asian_Pos0_5_Pos1_0_Away_Open = MatchData_DataFrame['fulltime_asian_handicap_pos_0_5_pos_1_0_away_open']
            Matchtime_Asian_Pos1_Home_Open = MatchData_DataFrame['fulltime_asian_handicap_pos_1_home_open']
            Matchtime_Asian_Pos1_Away_Open = MatchData_DataFrame['fulltime_asian_handicap_pos_1_away_open']
            Matchtime_Asian_Pos1_Pos1_5_Home_Open = MatchData_DataFrame['fulltime_asian_handicap_pos_1_0_pos_1_5_home_open']
            Matchtime_Asian_Pos1_Pos1_5_Away_Open = MatchData_DataFrame['fulltime_asian_handicap_pos_1_0_pos_1_5_away_open']
            Matchtime_Asian_Pos1_5_Home_Open = MatchData_DataFrame['fulltime_asian_handicap_pos_1_5_home_open']
            Matchtime_Asian_Pos1_5_Away_Open = MatchData_DataFrame['fulltime_asian_handicap_pos_1_5_away_open']
            Matchtime_Asian_Pos1_5_Pos2_Home_Open = MatchData_DataFrame['fulltime_asian_handicap_pos_1_5_pos_2_0_home_open']
            Matchtime_Asian_Pos1_5_Pos2_Away_Open = MatchData_DataFrame['fulltime_asian_handicap_pos_1_5_pos_2_0_away_open']
            Matchtime_Asian_Pos2_Home_Open = MatchData_DataFrame['fulltime_asian_handicap_pos_2_home_open']
            Matchtime_Asian_Pos2_Away_Open = MatchData_DataFrame['fulltime_asian_handicap_pos_2_away_open']

            # Firsthalf European Handicap
            Firsthalf_European_Neg1_Home_Open = MatchData_DataFrame['firsthalf_european_handicap_neg_1_home_open']
            Firsthalf_European_Neg1_Tie_Open = MatchData_DataFrame['firsthalf_european_handicap_neg_1_tie_open']
            Firsthalf_European_Neg1_Away_Open = MatchData_DataFrame['firsthalf_european_handicap_neg_1_away_open']
            Firsthalf_European_Neg1_Neg0_5_Home_Open = MatchData_DataFrame['firsthalf_european_handicap_neg_1_neg_0_5_home_open']
            Firsthalf_European_Neg1_Neg0_5_Tie_Open = MatchData_DataFrame['firsthalf_european_handicap_neg_1_neg_0_5_tie_open']
            Firsthalf_European_Neg1_Neg0_5_Away_Open = MatchData_DataFrame['firsthalf_european_handicap_neg_1_neg_0_5_away_open']
            Firsthalf_European_Neg0_5_Home_Open = MatchData_DataFrame['firsthalf_european_handicap_neg_0_5_home_open']
            Firsthalf_European_Neg0_5_Tie_Open = MatchData_DataFrame['firsthalf_european_handicap_neg_0_5_tie_open']
            Firsthalf_European_Neg0_5_Away_Open = MatchData_DataFrame['firsthalf_european_handicap_neg_0_5_away_open']
            Firsthalf_European_Neg0_5_0_0_Home_Open = MatchData_DataFrame['firsthalf_european_handicap_neg_0_5_0_0_home_open']
            Firsthalf_European_Neg0_5_0_0_Tie_Open = MatchData_DataFrame['firsthalf_european_handicap_neg_0_5_0_0_tie_open']
            Firsthalf_European_Neg0_5_0_0_Away_Open = MatchData_DataFrame['firsthalf_european_handicap_neg_0_5_0_0_away_open']
            Firsthalf_European_0_0_Pos0_5_Home_Open = MatchData_DataFrame['firsthalf_european_handicap_0_0_pos_0_5_home_open']
            Firsthalf_European_0_0_Pos0_5_Tie_Open = MatchData_DataFrame['firsthalf_european_handicap_0_0_pos_0_5_tie_open']
            Firsthalf_European_0_0_Pos0_5_Away_Open = MatchData_DataFrame['firsthalf_european_handicap_0_0_pos_0_5_away_open']
            Firsthalf_European_Pos0_5_Home_Open = MatchData_DataFrame['firsthalf_european_handicap_pos_0_5_home_open']
            Firsthalf_European_Pos0_5_Tie_Open = MatchData_DataFrame['firsthalf_european_handicap_pos_0_5_tie_open']
            Firsthalf_European_Pos0_5_Away_Open = MatchData_DataFrame['firsthalf_european_handicap_pos_0_5_away_open']
            Firsthalf_European_Pos0_5_Pos1_0_Home_Open = MatchData_DataFrame['firsthalf_european_handicap_pos_0_5_pos_1_0_home_open']
            Firsthalf_European_Pos0_5_Pos1_0_Tie_Open = MatchData_DataFrame['firsthalf_european_handicap_pos_0_5_pos_1_0_tie_open']
            Firsthalf_European_Pos0_5_Pos1_0_Away_Open = MatchData_DataFrame['firsthalf_european_handicap_pos_0_5_pos_1_0_away_open']
            Firsthalf_European_Pos1_0_Home_Open = MatchData_DataFrame['firsthalf_european_handicap_pos_1_home_open']
            Firsthalf_European_Pos1_0_Tie_Open = MatchData_DataFrame['firsthalf_european_handicap_pos_1_tie_open']
            Firsthalf_European_Pos1_0_Away_Open = MatchData_DataFrame['firsthalf_european_handicap_pos_1_away_open']

            # Matchtime European Handicap
            Matchtime_European_Neg1_Home_Open = MatchData_DataFrame['fulltime_european_handicap_neg_1_home_open']
            Matchtime_European_Neg1_Tie_Open = MatchData_DataFrame['fulltime_european_handicap_neg_1_tie_open']
            Matchtime_European_Neg1_Away_Open = MatchData_DataFrame['fulltime_european_handicap_neg_1_away_open']
            Matchtime_European_Neg1_Neg0_5_Home_Open = MatchData_DataFrame['fulltime_european_handicap_neg_1_neg_0_5_home_open']
            Matchtime_European_Neg1_Neg0_5_Tie_Open = MatchData_DataFrame['fulltime_european_handicap_neg_1_neg_0_5_tie_open']
            Matchtime_European_Neg1_Neg0_5_Away_Open = MatchData_DataFrame['fulltime_european_handicap_neg_1_neg_0_5_away_open']
            Matchtime_European_Neg0_5_Home_Open = MatchData_DataFrame['fulltime_european_handicap_neg_0_5_home_open']
            Matchtime_European_Neg0_5_Tie_Open = MatchData_DataFrame['fulltime_european_handicap_neg_0_5_tie_open']
            Matchtime_European_Neg0_5_Away_Open = MatchData_DataFrame['fulltime_european_handicap_neg_0_5_away_open']
            Matchtime_European_Neg0_5_0_0_Home_Open = MatchData_DataFrame['fulltime_european_handicap_neg_0_5_0_0_home_open']
            Matchtime_European_Neg0_5_0_0_Tie_Open = MatchData_DataFrame['fulltime_european_handicap_neg_0_5_0_0_tie_open']
            Matchtime_European_Neg0_5_0_0_Away_Open = MatchData_DataFrame['fulltime_european_handicap_neg_0_5_0_0_away_open']
            Matchtime_European_0_0_Pos0_5_Home_Open = MatchData_DataFrame['fulltime_european_handicap_0_0_pos_0_5_home_open']
            Matchtime_European_0_0_Pos0_5_Tie_Open = MatchData_DataFrame['fulltime_european_handicap_0_0_pos_0_5_tie_open']
            Matchtime_European_0_0_Pos0_5_Away_Open = MatchData_DataFrame['fulltime_european_handicap_0_0_pos_0_5_away_open']
            Matchtime_European_Pos0_5_Home_Open = MatchData_DataFrame['fulltime_european_handicap_pos_0_5_home_open']
            Matchtime_European_Pos0_5_Tie_Open = MatchData_DataFrame['fulltime_european_handicap_pos_0_5_tie_open']
            Matchtime_European_Pos0_5_Away_Open = MatchData_DataFrame['fulltime_european_handicap_pos_0_5_away_open']
            Matchtime_European_Pos0_5_Pos1_0_Home_Open = MatchData_DataFrame['fulltime_european_handicap_pos_0_5_pos_1_0_home_open']
            Matchtime_European_Pos0_5_Pos1_0_Tie_Open = MatchData_DataFrame['fulltime_european_handicap_pos_0_5_pos_1_0_tie_open']
            Matchtime_European_Pos0_5_Pos1_0_Away_Open = MatchData_DataFrame['fulltime_european_handicap_pos_0_5_pos_1_0_away_open']
            Matchtime_European_Pos1_0_Home_Open = MatchData_DataFrame['fulltime_european_handicap_pos_1_home_open']
            Matchtime_European_Pos1_0_Tie_Open = MatchData_DataFrame['fulltime_european_handicap_pos_1_tie_open']
            Matchtime_European_Pos1_0_Away_Open = MatchData_DataFrame['fulltime_european_handicap_pos_1_away_open']
            # OPEN ODDS END

            # CLOSE ODDS START
            # Fulltime
            Fulltime_Home_Close = MatchData_DataFrame['fulltime_home_close']
            Fulltime_Tie_Close = MatchData_DataFrame['fulltime_tie_close']
            Fulltime_Away_Close = MatchData_DataFrame['fulltime_home_close']

            # Firsthalf
            Firsthalf_Home_Close = MatchData_DataFrame['firsthalf_home_close']
            Firsthalf_Tie_Close = MatchData_DataFrame['firsthalf_tie_close']
            Firsthalf_Away_Close = MatchData_DataFrame['firsthalf_away_close']

            # Firsthalf-Fulltime
            Home_Home_Close = MatchData_DataFrame['home_home_close']
            Home_Tie_Close = MatchData_DataFrame['home_tie_close']
            Home_Away_Close = MatchData_DataFrame['home_away_close']
            Tie_Home_Close = MatchData_DataFrame['tie_home_close']
            Tie_Tie_Close = MatchData_DataFrame['tie_tie_close']
            Tie_Away_Close = MatchData_DataFrame['tie_away_close']
            Away_Home_Close = MatchData_DataFrame['away_home_close']
            Away_Tie_Close = MatchData_DataFrame['away_tie_close']
            Away_Away_Close = MatchData_DataFrame['away_away_close']

            # Firsthalf Exact Score
            Firsthalf_1_0_Close = MatchData_DataFrame['first_half_1_0_close']
            Firsthalf_2_0_Close = MatchData_DataFrame['first_half_2_0_close']
            Firsthalf_2_1_Close = MatchData_DataFrame['first_half_2_1_close']
            Firsthalf_3_0_Close = MatchData_DataFrame['first_half_3_0_close']
            Firsthalf_3_1_Close = MatchData_DataFrame['first_half_3_1_close']
            Firsthalf_3_2_Close = MatchData_DataFrame['first_half_3_2_close']
            Firsthalf_4_0_Close = MatchData_DataFrame['first_half_4_0_close']
            Firsthalf_4_1_Close = MatchData_DataFrame['first_half_4_1_close']
            Firsthalf_4_2_Close = MatchData_DataFrame['first_half_4_2_close']
            Firsthalf_4_3_Close = MatchData_DataFrame['first_half_4_3_close']
            Firsthalf_0_0_Close = MatchData_DataFrame['first_half_0_0_close']
            Firsthalf_1_1_Close = MatchData_DataFrame['first_half_1_1_close']
            Firsthalf_2_2_Close = MatchData_DataFrame['first_half_2_2_close']
            Firsthalf_3_3_Close = MatchData_DataFrame['first_half_3_3_close']
            Firsthalf_4_4_Close = MatchData_DataFrame['first_half_4_4_close']
            Firsthalf_0_1_Close = MatchData_DataFrame['first_half_0_1_close']
            Firsthalf_0_2_Close = MatchData_DataFrame['first_half_0_2_close']
            Firsthalf_1_2_Close = MatchData_DataFrame['first_half_1_2_close']
            Firsthalf_0_3_Close = MatchData_DataFrame['first_half_0_3_close']
            Firsthalf_1_3_Close = MatchData_DataFrame['first_half_1_3_close']
            Firsthalf_2_3_Close = MatchData_DataFrame['first_half_2_3_close']
            Firsthalf_0_4_Close = MatchData_DataFrame['first_half_0_4_close']
            Firsthalf_1_4_Close = MatchData_DataFrame['first_half_1_4_close']
            Firsthalf_2_4_Close = MatchData_DataFrame['first_half_2_4_close']
            Firsthalf_3_4_Close = MatchData_DataFrame['first_half_3_4_close']

            # Matchtime Exact Score
            Matchtime_1_0_Close = MatchData_DataFrame['match_time_1_0_close']
            Matchtime_2_0_Close = MatchData_DataFrame['match_time_2_0_close']
            Matchtime_2_1_Close = MatchData_DataFrame['match_time_2_1_close']
            Matchtime_3_0_Close = MatchData_DataFrame['match_time_3_0_close']
            Matchtime_3_1_Close = MatchData_DataFrame['match_time_3_1_close']
            Matchtime_3_2_Close = MatchData_DataFrame['match_time_3_2_close']
            Matchtime_4_0_Close = MatchData_DataFrame['match_time_4_0_close']
            Matchtime_4_1_Close = MatchData_DataFrame['match_time_4_1_close']
            Matchtime_4_2_Close = MatchData_DataFrame['match_time_4_2_close']
            Matchtime_4_3_Close = MatchData_DataFrame['match_time_4_3_close']
            Matchtime_0_0_Close = MatchData_DataFrame['match_time_0_0_close']
            Matchtime_1_1_Close = MatchData_DataFrame['match_time_1_1_close']
            Matchtime_2_2_Close = MatchData_DataFrame['match_time_2_2_close']
            Matchtime_3_3_Close = MatchData_DataFrame['match_time_3_3_close']
            Matchtime_4_4_Close = MatchData_DataFrame['match_time_4_4_close']
            Matchtime_0_1_Close = MatchData_DataFrame['match_time_0_1_close']
            Matchtime_0_2_Close = MatchData_DataFrame['match_time_0_2_close']
            Matchtime_1_2_Close = MatchData_DataFrame['match_time_1_2_close']
            Matchtime_0_3_Close = MatchData_DataFrame['match_time_0_3_close']
            Matchtime_1_3_Close = MatchData_DataFrame['match_time_1_3_close']
            Matchtime_2_3_Close = MatchData_DataFrame['match_time_2_3_close']
            Matchtime_0_4_Close = MatchData_DataFrame['match_time_0_4_close']
            Matchtime_1_4_Close = MatchData_DataFrame['match_time_1_4_close']
            Matchtime_2_4_Close = MatchData_DataFrame['match_time_2_4_close']
            Matchtime_3_4_Close = MatchData_DataFrame['match_time_3_4_close']

            # Double Chance
            Firsthalf_Home_And_Tie_Close = MatchData_DataFrame['first_half_home_and_tie_close']
            Firsthalf_Home_And_Away_Close = MatchData_DataFrame['first_half_home_and_away_close']
            Firsthalf_Tie_And_Away_Close = MatchData_DataFrame['first_half_tie_and_away_close']
            Matchtime_Home_And_Tie_Close = MatchData_DataFrame['match_time_home_and_tie_close']
            Matchtime_Home_And_Away_Close = MatchData_DataFrame['match_time_home_and_away_close']
            Matchtime_Tie_And_Away_Close = MatchData_DataFrame['match_time_tie_and_away_close']

            # Firsthalf Over/Under
            Firsthalf_Over_0_5_Close = MatchData_DataFrame['first_half_over_0_5_close']
            Firsthalf_Under_0_5_Close = MatchData_DataFrame['first_half_under_0_5_close']
            Firsthalf_Over_1_5_Close = MatchData_DataFrame['first_half_over_1_5_close']
            Firsthalf_Under_1_5_Close = MatchData_DataFrame['first_half_under_1_5_close']

            # Matchtime Over/Under
            Matchtime_Over_0_5_Close = MatchData_DataFrame['match_time_over_0_5_close']
            Matchtime_Under_0_5_Close = MatchData_DataFrame['match_time_under_0_5_close']
            Matchtime_Over_1_5_Close = MatchData_DataFrame['match_time_over_1_5_close']
            Matchtime_Under_1_5_Close = MatchData_DataFrame['match_time_under_1_5_close']
            Matchtime_Over_2_5_Close = MatchData_DataFrame['match_time_over_2_5_close']
            Matchtime_Under_2_5_Close = MatchData_DataFrame['match_time_under_2_5_close']
            Matchtime_Over_3_5_Close = MatchData_DataFrame['match_time_over_3_5_close']
            Matchtime_Under_3_5_Close = MatchData_DataFrame['match_time_under_3_5_close']

            # Draw No Bet
            Draw_No_Bet_Home_Close = MatchData_DataFrame['draw_no_bet_home_close']
            Draw_No_Bet_Away_Close = MatchData_DataFrame['draw_no_bet_away_close']

            # Both Teams To Score
            Firsthalf_Both_Teams_To_Score_Yes_Close = MatchData_DataFrame['firsthalf_both_teams_to_score_yes_close']
            Firsthalf_Both_Teams_To_Score_No_Close = MatchData_DataFrame['firsthalf_both_teams_to_score_no_close']
            Matchtime_Both_Teams_To_Score_Yes_Close = MatchData_DataFrame['both_teams_to_score_yes_close']
            Matchtime_Both_Teams_To_Score_No_Close = MatchData_DataFrame['both_teams_to_score_no_close']

            # Odd/Even
            Odd_Close = MatchData_DataFrame['odd_close']
            Even_Close = MatchData_DataFrame['even_close']

            # Firsthalf Asian Handicap
            Firsthalf_Asian_Neg2_Home_Close = MatchData_DataFrame['halftime_asian_handicap_neg2_home_close']
            Firsthalf_Asian_Neg2_Away_Close = MatchData_DataFrame['halftime_asian_handicap_neg2_away_close']
            Firsthalf_Asian_Neg2_Neg1_5_Home_Close = MatchData_DataFrame['halftime_asian_handicap_neg2_neg1_5_home_close']
            Firsthalf_Asian_Neg2_Neg1_5_Away_Close = MatchData_DataFrame['halftime_asian_handicap_neg2_neg1_5_away_close']
            Firsthalf_Asian_Neg1_5_Home_Close = MatchData_DataFrame['halftime_asian_handicap_neg1_5_home_close']
            Firsthalf_Asian_Neg1_5_Away_Close = MatchData_DataFrame['halftime_asian_handicap_neg1_5_away_close']
            Firsthalf_Asian_Neg1_5_Neg1_Home_Close = MatchData_DataFrame['halftime_asian_handicap_neg1_5_neg_1_home_close']
            Firsthalf_Asian_Neg1_5_Neg1_Away_Close = MatchData_DataFrame['halftime_asian_handicap_neg1_5_neg_1_away_close']
            Firsthalf_Asian_Neg1_Home_Close = MatchData_DataFrame['halftime_asian_handicap_neg_1_home_close']
            Firsthalf_Asian_Neg1_Away_Close = MatchData_DataFrame['halftime_asian_handicap_neg_1_away_close']
            Firsthalf_Asian_Neg1_Neg0_5_Home_Close = MatchData_DataFrame['halftime_asian_handicap_neg_1_neg_0_5_home_close']
            Firsthalf_Asian_Neg1_Neg0_5_Away_Close = MatchData_DataFrame['halftime_asian_handicap_neg_1_neg_0_5_away_close']
            Firsthalf_Asian_Neg0_5_Home_Close = MatchData_DataFrame['halftime_asian_handicap_neg_0_5_home_close']
            Firsthalf_Asian_Neg0_5_Away_Close = MatchData_DataFrame['halftime_asian_handicap_neg_0_5_away_close']
            Firsthalf_Asian_Neg0_5_Neg0_0_Home_Close = MatchData_DataFrame['halftime_asian_handicap_neg_0_5_0_0_home_close']
            Firsthalf_Asian_Neg0_5_Neg0_0_Away_Close = MatchData_DataFrame['halftime_asian_handicap_neg_0_5_0_0_away_close']
            Firsthalf_Asian_0_0_Pos0_5_Home_Close = MatchData_DataFrame['halftime_asian_handicap_0_0_pos_0_5_home_close']
            Firsthalf_Asian_0_0_Pos0_5_Away_Close = MatchData_DataFrame['halftime_asian_handicap_0_0_pos_0_5_away_close']
            Firsthalf_Asian_Pos0_5_Home_Close = MatchData_DataFrame['halftime_asian_handicap_pos_0_5_home_close']
            Firsthalf_Asian_Pos0_5_Away_Close = MatchData_DataFrame['halftime_asian_handicap_pos_0_5_away_close']
            Firsthalf_Asian_Pos0_5_Pos1_0_Home_Close = MatchData_DataFrame['halftime_asian_handicap_pos_0_5_pos_1_0_home_close']
            Firsthalf_Asian_Pos0_5_Pos1_0_Away_Close = MatchData_DataFrame['halftime_asian_handicap_pos_0_5_pos_1_0_away_close']
            Firsthalf_Asian_Pos1_Home_Close = MatchData_DataFrame['halftime_asian_handicap_pos_1_home_close']
            Firsthalf_Asian_Pos1_Away_Close = MatchData_DataFrame['halftime_asian_handicap_pos_1_away_close']
            Firsthalf_Asian_Pos1_Pos1_5_Home_Close = MatchData_DataFrame['halftime_asian_handicap_pos_1_0_pos_1_5_home_close']
            Firsthalf_Asian_Pos1_Pos1_5_Away_Close = MatchData_DataFrame['halftime_asian_handicap_pos_1_0_pos_1_5_away_close']
            Firsthalf_Asian_Pos1_5_Home_Close = MatchData_DataFrame['halftime_asian_handicap_pos_1_5_home_close']
            Firsthalf_Asian_Pos1_5_Away_Close = MatchData_DataFrame['halftime_asian_handicap_pos_1_5_away_close']
            Firsthalf_Asian_Pos1_5_Pos2_Home_Close = MatchData_DataFrame['halftime_asian_handicap_pos_1_5_pos_2_0_home_close']
            Firsthalf_Asian_Pos1_5_Pos2_Away_Close = MatchData_DataFrame['halftime_asian_handicap_pos_1_5_pos_2_0_away_close']
            Firsthalf_Asian_Pos2_Home_Close = MatchData_DataFrame['halftime_asian_handicap_pos_2_home_close']
            Firsthalf_Asian_Pos2_Away_Close = MatchData_DataFrame['halftime_asian_handicap_pos_2_away_close']

            # Matchtime Asian Handicap
            Matchtime_Asian_Neg2_Home_Close = MatchData_DataFrame['fulltime_asian_handicap_neg2_home_close']
            Matchtime_Asian_Neg2_Away_Close = MatchData_DataFrame['fulltime_asian_handicap_neg2_away_close']
            Matchtime_Asian_Neg2_Neg1_5_Home_Close = MatchData_DataFrame['fulltime_asian_handicap_neg2_neg1_5_home_close']
            Matchtime_Asian_Neg2_Neg1_5_Away_Close = MatchData_DataFrame['fulltime_asian_handicap_neg2_neg1_5_away_close']
            Matchtime_Asian_Neg1_5_Home_Close = MatchData_DataFrame['fulltime_asian_handicap_neg1_5_home_close']
            Matchtime_Asian_Neg1_5_Away_Close = MatchData_DataFrame['fulltime_asian_handicap_neg1_5_away_close']
            Matchtime_Asian_Neg1_5_Neg1_Home_Close = MatchData_DataFrame['fulltime_asian_handicap_neg1_5_neg_1_home_close']
            Matchtime_Asian_Neg1_5_Neg1_Away_Close = MatchData_DataFrame['fulltime_asian_handicap_neg1_5_neg_1_away_close']
            Matchtime_Asian_Neg1_Home_Close = MatchData_DataFrame['fulltime_asian_handicap_neg_1_home_close']
            Matchtime_Asian_Neg1_Away_Close = MatchData_DataFrame['fulltime_asian_handicap_neg_1_away_close']
            Matchtime_Asian_Neg1_Neg0_5_Home_Close = MatchData_DataFrame['fulltime_asian_handicap_neg_1_neg_0_5_home_close']
            Matchtime_Asian_Neg1_Neg0_5_Away_Close = MatchData_DataFrame['fulltime_asian_handicap_neg_1_neg_0_5_away_close']
            Matchtime_Asian_Neg0_5_Home_Close = MatchData_DataFrame['fulltime_asian_handicap_neg_0_5_home_close']
            Matchtime_Asian_Neg0_5_Away_Close = MatchData_DataFrame['fulltime_asian_handicap_neg_0_5_away_close']
            Matchtime_Asian_Neg0_5_Neg0_0_Home_Close = MatchData_DataFrame['fulltime_asian_handicap_neg_0_5_0_0_home_close']
            Matchtime_Asian_Neg0_5_Neg0_0_Away_Close = MatchData_DataFrame['fulltime_asian_handicap_neg_0_5_0_0_away_close']
            Matchtime_Asian_0_0_Pos0_5_Home_Close = MatchData_DataFrame['fulltime_asian_handicap_0_0_pos_0_5_home_close']
            Matchtime_Asian_0_0_Pos0_5_Away_Close = MatchData_DataFrame['fulltime_asian_handicap_0_0_pos_0_5_away_close']
            Matchtime_Asian_Pos0_5_Home_Close = MatchData_DataFrame['fulltime_asian_handicap_pos_0_5_home_close']
            Matchtime_Asian_Pos0_5_Away_Close = MatchData_DataFrame['fulltime_asian_handicap_pos_0_5_away_close']
            Matchtime_Asian_Pos0_5_Pos1_0_Home_Close = MatchData_DataFrame['fulltime_asian_handicap_pos_0_5_pos_1_0_home_close']
            Matchtime_Asian_Pos0_5_Pos1_0_Away_Close = MatchData_DataFrame['fulltime_asian_handicap_pos_0_5_pos_1_0_away_close']
            Matchtime_Asian_Pos1_Home_Close = MatchData_DataFrame['fulltime_asian_handicap_pos_1_home_close']
            Matchtime_Asian_Pos1_Away_Close = MatchData_DataFrame['fulltime_asian_handicap_pos_1_away_close']
            Matchtime_Asian_Pos1_Pos1_5_Home_Close = MatchData_DataFrame['fulltime_asian_handicap_pos_1_0_pos_1_5_home_close']
            Matchtime_Asian_Pos1_Pos1_5_Away_Close = MatchData_DataFrame['fulltime_asian_handicap_pos_1_0_pos_1_5_away_close']
            Matchtime_Asian_Pos1_5_Home_Close = MatchData_DataFrame['fulltime_asian_handicap_pos_1_5_home_close']
            Matchtime_Asian_Pos1_5_Away_Close = MatchData_DataFrame['fulltime_asian_handicap_pos_1_5_away_close']
            Matchtime_Asian_Pos1_5_Pos2_Home_Close = MatchData_DataFrame['fulltime_asian_handicap_pos_1_5_pos_2_0_home_close']
            Matchtime_Asian_Pos1_5_Pos2_Away_Close = MatchData_DataFrame['fulltime_asian_handicap_pos_1_5_pos_2_0_away_close']
            Matchtime_Asian_Pos2_Home_Close = MatchData_DataFrame['fulltime_asian_handicap_pos_2_home_close']
            Matchtime_Asian_Pos2_Away_Close = MatchData_DataFrame['fulltime_asian_handicap_pos_2_away_close']

            # Firsthalf European Handicap
            Firsthalf_European_Neg1_Home_Close = MatchData_DataFrame['firsthalf_european_handicap_neg_1_home_close']
            Firsthalf_European_Neg1_Tie_Close = MatchData_DataFrame['firsthalf_european_handicap_neg_1_tie_close']
            Firsthalf_European_Neg1_Away_Close = MatchData_DataFrame['firsthalf_european_handicap_neg_1_away_close']
            Firsthalf_European_Neg1_Neg0_5_Home_Close = MatchData_DataFrame['firsthalf_european_handicap_neg_1_neg_0_5_home_close']
            Firsthalf_European_Neg1_Neg0_5_Tie_Close = MatchData_DataFrame['firsthalf_european_handicap_neg_1_neg_0_5_tie_close']
            Firsthalf_European_Neg1_Neg0_5_Away_Close = MatchData_DataFrame['firsthalf_european_handicap_neg_1_neg_0_5_away_close']
            Firsthalf_European_Neg0_5_Home_Close = MatchData_DataFrame['firsthalf_european_handicap_neg_0_5_home_close']
            Firsthalf_European_Neg0_5_Tie_Close = MatchData_DataFrame['firsthalf_european_handicap_neg_0_5_tie_close']
            Firsthalf_European_Neg0_5_Away_Close = MatchData_DataFrame['firsthalf_european_handicap_neg_0_5_away_close']
            Firsthalf_European_Neg0_5_0_0_Home_Close = MatchData_DataFrame['firsthalf_european_handicap_neg_0_5_0_0_home_close']
            Firsthalf_European_Neg0_5_0_0_Tie_Close = MatchData_DataFrame['firsthalf_european_handicap_neg_0_5_0_0_tie_close']
            Firsthalf_European_Neg0_5_0_0_Away_Close = MatchData_DataFrame['firsthalf_european_handicap_neg_0_5_0_0_away_close']
            Firsthalf_European_0_0_Pos0_5_Home_Close = MatchData_DataFrame['firsthalf_european_handicap_0_0_pos_0_5_home_close']
            Firsthalf_European_0_0_Pos0_5_Tie_Close = MatchData_DataFrame['firsthalf_european_handicap_0_0_pos_0_5_tie_close']
            Firsthalf_European_0_0_Pos0_5_Away_Close = MatchData_DataFrame['firsthalf_european_handicap_0_0_pos_0_5_away_close']
            Firsthalf_European_Pos0_5_Home_Close = MatchData_DataFrame['firsthalf_european_handicap_pos_0_5_home_close']
            Firsthalf_European_Pos0_5_Tie_Close = MatchData_DataFrame['firsthalf_european_handicap_pos_0_5_tie_close']
            Firsthalf_European_Pos0_5_Away_Close = MatchData_DataFrame['firsthalf_european_handicap_pos_0_5_away_close']
            Firsthalf_European_Pos0_5_Pos1_0_Home_Close = MatchData_DataFrame['firsthalf_european_handicap_pos_0_5_pos_1_0_home_close']
            Firsthalf_European_Pos0_5_Pos1_0_Tie_Close = MatchData_DataFrame['firsthalf_european_handicap_pos_0_5_pos_1_0_tie_close']
            Firsthalf_European_Pos0_5_Pos1_0_Away_Close = MatchData_DataFrame['firsthalf_european_handicap_pos_0_5_pos_1_0_away_close']
            Firsthalf_European_Pos1_0_Home_Close = MatchData_DataFrame['firsthalf_european_handicap_pos_1_home_close']
            Firsthalf_European_Pos1_0_Tie_Close = MatchData_DataFrame['firsthalf_european_handicap_pos_1_tie_close']
            Firsthalf_European_Pos1_0_Away_Close = MatchData_DataFrame['firsthalf_european_handicap_pos_1_away_close']

            # Matchtime European Handicap
            Matchtime_European_Neg1_Home_Close = MatchData_DataFrame['fulltime_european_handicap_neg_1_home_close']
            Matchtime_European_Neg1_Tie_Close = MatchData_DataFrame['fulltime_european_handicap_neg_1_tie_close']
            Matchtime_European_Neg1_Away_Close = MatchData_DataFrame['fulltime_european_handicap_neg_1_away_close']
            Matchtime_European_Neg1_Neg0_5_Home_Close = MatchData_DataFrame['fulltime_european_handicap_neg_1_neg_0_5_home_close']
            Matchtime_European_Neg1_Neg0_5_Tie_Close = MatchData_DataFrame['fulltime_european_handicap_neg_1_neg_0_5_tie_close']
            Matchtime_European_Neg1_Neg0_5_Away_Close = MatchData_DataFrame['fulltime_european_handicap_neg_1_neg_0_5_away_close']
            Matchtime_European_Neg0_5_Home_Close = MatchData_DataFrame['fulltime_european_handicap_neg_0_5_home_close']
            Matchtime_European_Neg0_5_Tie_Close = MatchData_DataFrame['fulltime_european_handicap_neg_0_5_tie_close']
            Matchtime_European_Neg0_5_Away_Close = MatchData_DataFrame['fulltime_european_handicap_neg_0_5_away_close']
            Matchtime_European_Neg0_5_0_0_Home_Close = MatchData_DataFrame['fulltime_european_handicap_neg_0_5_0_0_home_close']
            Matchtime_European_Neg0_5_0_0_Tie_Close = MatchData_DataFrame['fulltime_european_handicap_neg_0_5_0_0_tie_close']
            Matchtime_European_Neg0_5_0_0_Away_Close = MatchData_DataFrame['fulltime_european_handicap_neg_0_5_0_0_away_close']
            Matchtime_European_0_0_Pos0_5_Home_Close = MatchData_DataFrame['fulltime_european_handicap_0_0_pos_0_5_home_close']
            Matchtime_European_0_0_Pos0_5_Tie_Close = MatchData_DataFrame['fulltime_european_handicap_0_0_pos_0_5_tie_close']
            Matchtime_European_0_0_Pos0_5_Away_Close = MatchData_DataFrame['fulltime_european_handicap_0_0_pos_0_5_away_close']
            Matchtime_European_Pos0_5_Home_Close = MatchData_DataFrame['fulltime_european_handicap_pos_0_5_home_close']
            Matchtime_European_Pos0_5_Tie_Close = MatchData_DataFrame['fulltime_european_handicap_pos_0_5_tie_close']
            Matchtime_European_Pos0_5_Away_Close = MatchData_DataFrame['fulltime_european_handicap_pos_0_5_away_close']
            Matchtime_European_Pos0_5_Pos1_0_Home_Close = MatchData_DataFrame['fulltime_european_handicap_pos_0_5_pos_1_0_home_close']
            Matchtime_European_Pos0_5_Pos1_0_Tie_Close = MatchData_DataFrame['fulltime_european_handicap_pos_0_5_pos_1_0_tie_close']
            Matchtime_European_Pos0_5_Pos1_0_Away_Close = MatchData_DataFrame['fulltime_european_handicap_pos_0_5_pos_1_0_away_close']
            Matchtime_European_Pos1_0_Home_Close = MatchData_DataFrame['fulltime_european_handicap_pos_1_home_close']
            Matchtime_European_Pos1_0_Tie_Close = MatchData_DataFrame['fulltime_european_handicap_pos_1_tie_close']
            Matchtime_European_Pos1_0_Away_Close = MatchData_DataFrame['fulltime_european_handicap_pos_1_away_close']
            # CLOSE ODDS END

            with tab1:
                streamlit.markdown("Analiz başlangıç ve bitiş tarihlerini seçin")
                streamlit.date_input(label='Başlangıç Tarihi', min_value=datetime(2019, 1, 1), value=streamlit.session_state.Filters_Dictionary['filter_start_date'], max_value=datetime.today(), key='database_start_date')
                streamlit.date_input(label='Bitiş Tarihi', min_value=streamlit.session_state.database_start_date, value=streamlit.session_state.Filters_Dictionary['filter_end_date'], max_value=datetime.today(), key='database_end_date')
            with tab2:
                streamlit.markdown("Genel filtreler")
                streamlit.session_state.Filters_Dictionary['filter_country'] = streamlit.checkbox(f'Ülke | {Ülke}', value=streamlit.session_state.Filters_Dictionary['filter_country'])
                streamlit.session_state.Filters_Dictionary['filter_league'] = streamlit.checkbox(f'Lig | {Lig}', value=streamlit.session_state.Filters_Dictionary['filter_league'])
                streamlit.session_state.Filters_Dictionary['filter_week'] = streamlit.checkbox(f'Hafta | {Hafta}', value=streamlit.session_state.Filters_Dictionary['filter_week'])
                streamlit.session_state.Filters_Dictionary['filter_hour'] = streamlit.checkbox(f'Saat | {Saat}', value=streamlit.session_state.Filters_Dictionary['filter_hour'])
                streamlit.session_state.Filters_Dictionary['filter_home_team'] = streamlit.checkbox(f'Ev | {Ev}', value=streamlit.session_state.Filters_Dictionary['filter_home_team'])
                streamlit.session_state.Filters_Dictionary['filter_away_team'] = streamlit.checkbox(f'Dep | {Dep}', value=streamlit.session_state.Filters_Dictionary['filter_away_team'])
                streamlit.session_state.Filters_Dictionary['filter_referee'] = streamlit.checkbox(f'Hakem | {Hakem}', value=streamlit.session_state.Filters_Dictionary['filter_referee'])
            with tab3:
                with streamlit.expander("Maç sonucu filtreleri"):
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_1'] = streamlit.checkbox(f'A MS 1 | {Fulltime_Home_Open:.2f}' if isinstance(Fulltime_Home_Open, (int, float)) else f'A MS 1 | {Fulltime_Home_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_ms_1'])
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_x'] = streamlit.checkbox(f'A MS X | {Fulltime_Tie_Open:.2f}' if isinstance(Fulltime_Tie_Open, (int, float)) else f'A MS X | {Fulltime_Tie_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_ms_x'])
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_2'] = streamlit.checkbox(f'A MS 2 | {Fulltime_Away_Open:.2f}' if isinstance(Fulltime_Away_Open, (int, float)) else f'A MS 2 | {Fulltime_Away_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_ms_2'])
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_0_5_a'] = streamlit.checkbox(f'A MS 0.5 A | {Matchtime_Under_0_5_Open:.2f}' if isinstance(Matchtime_Under_0_5_Open, (int, float)) else f'A MS 0.5 A | {Matchtime_Under_0_5_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_ms_0_5_a'])
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_0_5_ü'] = streamlit.checkbox(f'A MS 0.5 Ü | {Matchtime_Over_0_5_Open:.2f}' if isinstance(Matchtime_Over_0_5_Open, (int, float)) else f'A MS 0.5 Ü | {Matchtime_Over_0_5_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_ms_0_5_ü'])
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_1_5_a'] = streamlit.checkbox(f'A MS 1.5 A | {Matchtime_Under_1_5_Open:.2f}' if isinstance(Matchtime_Under_1_5_Open, (int, float)) else f'A MS 1.5 A | {Matchtime_Under_1_5_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_ms_1_5_a'])
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_1_5_ü'] = streamlit.checkbox(f'A MS 1.5 Ü | {Matchtime_Over_1_5_Open:.2f}' if isinstance(Matchtime_Over_1_5_Open, (int, float)) else f'A MS 1.5 Ü | {Matchtime_Over_1_5_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_ms_1_5_ü'])
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_2_5_a'] = streamlit.checkbox(f'A MS 2.5 A | {Matchtime_Under_2_5_Open:.2f}' if isinstance(Matchtime_Under_2_5_Open, (int, float)) else f'A MS 2.5 A | {Matchtime_Under_2_5_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_ms_2_5_a'])
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_2_5_ü'] = streamlit.checkbox(f'A MS 2.5 Ü | {Matchtime_Over_2_5_Open:.2f}' if isinstance(Matchtime_Over_2_5_Open, (int, float)) else f'A MS 2.5 Ü | {Matchtime_Over_2_5_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_ms_2_5_ü'])
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_3_5_a'] = streamlit.checkbox(f'A MS 3.5 A | {Matchtime_Under_3_5_Open:.2f}' if isinstance(Matchtime_Under_3_5_Open, (int, float)) else f'A MS 3.5 A | {Matchtime_Under_3_5_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_ms_3_5_a'])
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_3_5_ü'] = streamlit.checkbox(f'A MS 3.5 Ü | {Matchtime_Over_3_5_Open:.2f}' if isinstance(Matchtime_Over_3_5_Open, (int, float)) else f'A MS 3.5 Ü | {Matchtime_Over_3_5_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_ms_3_5_ü'])

                with streamlit.expander("İlk yarı filtreleri"):
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_1'] = streamlit.checkbox(f'A İY 1 | {Firsthalf_Home_Open:.2f}' if isinstance(Firsthalf_Home_Open, (int, float)) else f'A İY 1 | {Firsthalf_Home_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_iy_1'])
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_x'] = streamlit.checkbox(f'A İY X | {Firsthalf_Tie_Open:.2f}' if isinstance(Firsthalf_Tie_Open, (int, float)) else f'A İY X | {Firsthalf_Tie_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_iy_x'])
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_2'] = streamlit.checkbox(f'A İY 2 | {Firsthalf_Away_Open:.2f}' if isinstance(Firsthalf_Away_Open, (int, float)) else f'A İY 2 | {Firsthalf_Away_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_iy_2'])
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_0_5_a'] = streamlit.checkbox(f'A İY 0.5 A | {Firsthalf_Under_0_5_Open:.2f}' if isinstance(Firsthalf_Under_0_5_Open, (int, float)) else f'A İY 0.5 A | {Firsthalf_Under_0_5_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_iy_0_5_a'])
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_0_5_ü'] = streamlit.checkbox(f'A İY 0.5 Ü | {Firsthalf_Over_0_5_Open:.2f}' if isinstance(Firsthalf_Over_0_5_Open, (int, float)) else f'A İY 0.5 Ü | {Firsthalf_Over_0_5_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_iy_0_5_ü'])
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_1_5_a'] = streamlit.checkbox(f'A İY 1.5 A | {Firsthalf_Under_1_5_Open:.2f}' if isinstance(Firsthalf_Under_1_5_Open, (int, float)) else f'A İY 1.5 A | {Firsthalf_Under_1_5_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_iy_1_5_a'])
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_1_5_ü'] = streamlit.checkbox(f'A İY 1.5 Ü | {Firsthalf_Over_1_5_Open:.2f}' if isinstance(Firsthalf_Over_1_5_Open, (int, float)) else f'A İY 1.5 Ü | {Firsthalf_Over_1_5_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_iy_1_5_ü'])

                with streamlit.expander("Çifte şans filtreleri"):
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_1_x_cs'] = streamlit.checkbox(f'A MS 1-X | {Matchtime_Home_And_Tie_Open:.2f}' if isinstance(Matchtime_Home_And_Tie_Open, (int, float)) else f'A MS 1-X | {Matchtime_Home_And_Tie_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_ms_1_x_cs'])
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_1_2_cs'] = streamlit.checkbox(f'A MS 1-2 | {Matchtime_Home_And_Away_Open:.2f}' if isinstance(Matchtime_Home_And_Away_Open, (int, float)) else f'A MS 1-2 | {Matchtime_Home_And_Away_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_ms_1_2_cs'])
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_x_2_cs'] = streamlit.checkbox(f'A MS X-2 | {Matchtime_Tie_And_Away_Open:.2f}' if isinstance(Matchtime_Tie_And_Away_Open, (int, float)) else f'A MS X-2 | {Matchtime_Tie_And_Away_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_ms_x_2_cs'])
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_1_x_cs'] = streamlit.checkbox(f'A İY 1-X | {Firsthalf_Home_And_Tie_Open:.2f}' if isinstance(Firsthalf_Home_And_Tie_Open, (int, float)) else f'A İY 1-X | {Firsthalf_Home_And_Tie_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_iy_1_x_cs'])
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_1_2_cs'] = streamlit.checkbox(f'A İY 1-2 | {Firsthalf_Home_And_Away_Open:.2f}' if isinstance(Firsthalf_Home_And_Away_Open, (int, float)) else f'A İY 1-2 | {Firsthalf_Home_And_Away_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_iy_1_2_cs'])
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_x_2_cs'] = streamlit.checkbox(f'A İY X-2 | {Firsthalf_Tie_And_Away_Open:.2f}' if isinstance(Firsthalf_Tie_And_Away_Open, (int, float)) else f'A İY X-2 | {Firsthalf_Tie_And_Away_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_iy_x_2_cs'])

                with streamlit.expander("Karşılıklı filtreleri"):
                    streamlit.session_state.Filters_Dictionary['filter_a_home_win'] = streamlit.checkbox(f'A EV Kazanır | {Draw_No_Bet_Home_Open:.2f}' if isinstance(Draw_No_Bet_Home_Open, (int, float)) else f'A EV Kazanır | {Draw_No_Bet_Home_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_home_win'])
                    streamlit.session_state.Filters_Dictionary['filter_a_away_win'] = streamlit.checkbox(f'A DEP Kazanır | {Draw_No_Bet_Away_Open:.2f}' if isinstance(Draw_No_Bet_Away_Open, (int, float)) else f'A DEP Kazanır | {Draw_No_Bet_Away_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_away_win'])
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_kg_var'] = streamlit.checkbox(f'A İY KG VAR | {Firsthalf_Both_Teams_To_Score_Yes_Open:.2f}' if isinstance(Firsthalf_Both_Teams_To_Score_Yes_Open, (int, float)) else f'A İY KG VAR | {Firsthalf_Both_Teams_To_Score_Yes_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_iy_kg_var'])
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_kg_yok'] = streamlit.checkbox(f'A İY KG YOK | {Firsthalf_Both_Teams_To_Score_No_Open:.2f}' if isinstance(Firsthalf_Both_Teams_To_Score_No_Open, (int, float)) else f'A İY KG YOK | {Firsthalf_Both_Teams_To_Score_No_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_iy_kg_yok'])
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_kg_var'] = streamlit.checkbox(f'A MS KG VAR | {Matchtime_Both_Teams_To_Score_Yes_Open:.2f}' if isinstance(Matchtime_Both_Teams_To_Score_Yes_Open, (int, float)) else f'A MS KG VAR | {Matchtime_Both_Teams_To_Score_Yes_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_ms_kg_var'])
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_kg_yok'] = streamlit.checkbox(f'A MS KG YOK | {Matchtime_Both_Teams_To_Score_No_Open:.2f}' if isinstance(Matchtime_Both_Teams_To_Score_No_Open, (int, float)) else f'A MS KG YOK | {Matchtime_Both_Teams_To_Score_No_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_ms_kg_yok'])
                    streamlit.session_state.Filters_Dictionary['filter_a_tek'] = streamlit.checkbox(f'A TEK | {Odd_Open:.2f}' if isinstance(Odd_Open, (int, float)) else f'A TEK | {Odd_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_tek'])
                    streamlit.session_state.Filters_Dictionary['filter_a_cift'] = streamlit.checkbox(f'A ÇİFT | {Even_Open:.2f}' if isinstance(Even_Open, (int, float)) else f'A ÇİFT | {Even_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_cift'])

                with streamlit.expander("İlk yarı/Maç sonucu filtreleri"):
                    streamlit.session_state.Filters_Dictionary['filter_a_1den1'] = streamlit.checkbox(f'A 1/1 | {Home_Home_Open:.2f}' if isinstance(Home_Home_Open, (int, float)) else f'A 1/1 | {Home_Home_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_1den1'])
                    streamlit.session_state.Filters_Dictionary['filter_a_1denx'] = streamlit.checkbox(f'A 1/X | {Home_Tie_Open:.2f}' if isinstance(Home_Tie_Open, (int, float)) else f'A 1/X | {Home_Tie_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_1denx'])
                    streamlit.session_state.Filters_Dictionary['filter_a_1den2'] = streamlit.checkbox(f'A 1/2 | {Home_Away_Open:.2f}' if isinstance(Home_Away_Open, (int, float)) else f'A 1/2 | {Home_Away_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_1den2'])
                    streamlit.session_state.Filters_Dictionary['filter_a_xden1'] = streamlit.checkbox(f'A X/1 | {Tie_Home_Open:.2f}' if isinstance(Tie_Home_Open, (int, float)) else f'A X/1 | {Tie_Home_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_xden1'])
                    streamlit.session_state.Filters_Dictionary['filter_a_xdenx'] = streamlit.checkbox(f'A X/X | {Tie_Tie_Open:.2f}' if isinstance(Tie_Tie_Open, (int, float)) else f'A X/X | {Tie_Tie_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_xdenx'])
                    streamlit.session_state.Filters_Dictionary['filter_a_xden2'] = streamlit.checkbox(f'A X/2 | {Tie_Away_Open:.2f}' if isinstance(Tie_Away_Open, (int, float)) else f'A X/2 | {Tie_Away_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_xden2'])
                    streamlit.session_state.Filters_Dictionary['filter_a_2den1'] = streamlit.checkbox(f'A 2/1 | {Away_Home_Open:.2f}' if isinstance(Away_Home_Open, (int, float)) else f'A 2/1 | {Away_Home_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_2den1'])
                    streamlit.session_state.Filters_Dictionary['filter_a_2denx'] = streamlit.checkbox(f'A 2/X | {Away_Tie_Open:.2f}' if isinstance(Away_Tie_Open, (int, float)) else f'A 2/X | {Away_Tie_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_2denx'])
                    streamlit.session_state.Filters_Dictionary['filter_a_2den2'] = streamlit.checkbox(f'A 2/2 | {Away_Away_Open:.2f}' if isinstance(Away_Away_Open, (int, float)) else f'A 2/2 | {Away_Away_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_2den2'])

                with streamlit.expander("İlk yarı skor filtreleri"):
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_10'] = streamlit.checkbox(f'A İY SKOR 1-0 | {Firsthalf_1_0_Open:.2f}' if isinstance(Firsthalf_1_0_Open, (int, float)) else f'A İY SKOR 1-0 | {Firsthalf_1_0_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_10'])
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_20'] = streamlit.checkbox(f'A İY SKOR 2-0 | {Firsthalf_2_0_Open:.2f}' if isinstance(Firsthalf_2_0_Open, (int, float)) else f'A İY SKOR 2-0 | {Firsthalf_2_0_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_20'])
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_21'] = streamlit.checkbox(f'A İY SKOR 2-1 | {Firsthalf_2_1_Open:.2f}' if isinstance(Firsthalf_2_1_Open, (int, float)) else f'A İY SKOR 2-1 | {Firsthalf_2_1_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_21'])
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_30'] = streamlit.checkbox(f'A İY SKOR 3-0 | {Firsthalf_3_0_Open:.2f}' if isinstance(Firsthalf_3_0_Open, (int, float)) else f'A İY SKOR 3-0 | {Firsthalf_3_0_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_30'])
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_31'] = streamlit.checkbox(f'A İY SKOR 3-1 | {Firsthalf_3_1_Open:.2f}' if isinstance(Firsthalf_3_1_Open, (int, float)) else f'A İY SKOR 3-1 | {Firsthalf_3_1_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_31'])
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_32'] = streamlit.checkbox(f'A İY SKOR 3-2 | {Firsthalf_3_2_Open:.2f}' if isinstance(Firsthalf_3_2_Open, (int, float)) else f'A İY SKOR 3-2 | {Firsthalf_3_2_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_32'])
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_40'] = streamlit.checkbox(f'A İY SKOR 4-0 | {Firsthalf_4_0_Open:.2f}' if isinstance(Firsthalf_4_0_Open, (int, float)) else f'A İY SKOR 4-0 | {Firsthalf_4_0_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_40'])
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_41'] = streamlit.checkbox(f'A İY SKOR 4-1 | {Firsthalf_4_1_Open:.2f}' if isinstance(Firsthalf_4_1_Open, (int, float)) else f'A İY SKOR 4-1 | {Firsthalf_4_1_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_41'])
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_42'] = streamlit.checkbox(f'A İY SKOR 4-2 | {Firsthalf_4_2_Open:.2f}' if isinstance(Firsthalf_4_2_Open, (int, float)) else f'A İY SKOR 4-2 | {Firsthalf_4_2_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_42'])
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_43'] = streamlit.checkbox(f'A İY SKOR 4-3 | {Firsthalf_4_3_Open:.2f}' if isinstance(Firsthalf_4_3_Open, (int, float)) else f'A İY SKOR 4-3 | {Firsthalf_4_3_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_43'])
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_00'] = streamlit.checkbox(f'A İY SKOR 0-0 | {Firsthalf_0_0_Open:.2f}' if isinstance(Firsthalf_0_0_Open, (int, float)) else f'A İY SKOR 0-0 | {Firsthalf_0_0_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_00'])
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_11'] = streamlit.checkbox(f'A İY SKOR 1-1 | {Firsthalf_1_1_Open:.2f}' if isinstance(Firsthalf_1_1_Open, (int, float)) else f'A İY SKOR 1-1 | {Firsthalf_1_1_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_11'])
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_22'] = streamlit.checkbox(f'A İY SKOR 2-2 | {Firsthalf_2_2_Open:.2f}' if isinstance(Firsthalf_2_2_Open, (int, float)) else f'A İY SKOR 2-2 | {Firsthalf_2_2_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_22'])
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_33'] = streamlit.checkbox(f'A İY SKOR 3-3 | {Firsthalf_3_3_Open:.2f}' if isinstance(Firsthalf_3_3_Open, (int, float)) else f'A İY SKOR 3-3 | {Firsthalf_3_3_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_33'])
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_44'] = streamlit.checkbox(f'A İY SKOR 4-4 | {Firsthalf_4_4_Open:.2f}' if isinstance(Firsthalf_4_4_Open, (int, float)) else f'A İY SKOR 4-4 | {Firsthalf_4_4_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_44'])
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_01'] = streamlit.checkbox(f'A İY SKOR 0-1 | {Firsthalf_0_1_Open:.2f}' if isinstance(Firsthalf_0_1_Open, (int, float)) else f'A İY SKOR 0-1 | {Firsthalf_0_1_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_01'])
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_02'] = streamlit.checkbox(f'A İY SKOR 0-2 | {Firsthalf_0_2_Open:.2f}' if isinstance(Firsthalf_0_2_Open, (int, float)) else f'A İY SKOR 0-2 | {Firsthalf_0_2_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_02'])
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_12'] = streamlit.checkbox(f'A İY SKOR 1-2 | {Firsthalf_1_2_Open:.2f}' if isinstance(Firsthalf_1_2_Open, (int, float)) else f'A İY SKOR 1-2 | {Firsthalf_1_2_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_12'])
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_03'] = streamlit.checkbox(f'A İY SKOR 0-3 | {Firsthalf_0_3_Open:.2f}' if isinstance(Firsthalf_0_3_Open, (int, float)) else f'A İY SKOR 0-3 | {Firsthalf_0_3_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_03'])
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_13'] = streamlit.checkbox(f'A İY SKOR 1-3 | {Firsthalf_1_3_Open:.2f}' if isinstance(Firsthalf_1_3_Open, (int, float)) else f'A İY SKOR 1-3 | {Firsthalf_1_3_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_13'])
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_23'] = streamlit.checkbox(f'A İY SKOR 2-3 | {Firsthalf_2_3_Open:.2f}' if isinstance(Firsthalf_2_3_Open, (int, float)) else f'A İY SKOR 2-3 | {Firsthalf_2_3_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_23'])
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_04'] = streamlit.checkbox(f'A İY SKOR 0-4 | {Firsthalf_0_4_Open:.2f}' if isinstance(Firsthalf_0_4_Open, (int, float)) else f'A İY SKOR 0-4 | {Firsthalf_0_4_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_04'])
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_14'] = streamlit.checkbox(f'A İY SKOR 1-4 | {Firsthalf_1_4_Open:.2f}' if isinstance(Firsthalf_1_4_Open, (int, float)) else f'A İY SKOR 1-4 | {Firsthalf_1_4_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_14'])
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_24'] = streamlit.checkbox(f'A İY SKOR 2-4 | {Firsthalf_2_4_Open:.2f}' if isinstance(Firsthalf_2_4_Open, (int, float)) else f'A İY SKOR 2-4 | {Firsthalf_2_4_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_24'])
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_34'] = streamlit.checkbox(f'A İY SKOR 3-4 | {Firsthalf_3_4_Open:.2f}' if isinstance(Firsthalf_3_4_Open, (int, float)) else f'A İY SKOR 3-4 | {Firsthalf_3_4_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_34'])

                with streamlit.expander("Maç sonu skor filtreleri"):
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_10'] = streamlit.checkbox(f'A MS SKOR 1-0 | {Matchtime_1_0_Open:.2f}' if isinstance(Matchtime_1_0_Open, (int, float)) else f'A MS SKOR 1-0 | {Matchtime_1_0_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_10'])
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_20'] = streamlit.checkbox(f'A MS SKOR 2-0 | {Matchtime_2_0_Open:.2f}' if isinstance(Matchtime_2_0_Open, (int, float)) else f'A MS SKOR 2-0 | {Matchtime_2_0_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_20'])
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_21'] = streamlit.checkbox(f'A MS SKOR 2-1 | {Matchtime_2_1_Open:.2f}' if isinstance(Matchtime_2_1_Open, (int, float)) else f'A MS SKOR 2-1 | {Matchtime_2_1_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_21'])
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_30'] = streamlit.checkbox(f'A MS SKOR 3-0 | {Matchtime_3_0_Open:.2f}' if isinstance(Matchtime_3_0_Open, (int, float)) else f'A MS SKOR 3-0 | {Matchtime_3_0_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_30'])
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_31'] = streamlit.checkbox(f'A MS SKOR 3-1 | {Matchtime_3_1_Open:.2f}' if isinstance(Matchtime_3_1_Open, (int, float)) else f'A MS SKOR 3-1 | {Matchtime_3_1_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_31'])
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_32'] = streamlit.checkbox(f'A MS SKOR 3-2 | {Matchtime_3_2_Open:.2f}' if isinstance(Matchtime_3_2_Open, (int, float)) else f'A MS SKOR 3-2 | {Matchtime_3_2_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_32'])
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_40'] = streamlit.checkbox(f'A MS SKOR 4-0 | {Matchtime_4_0_Open:.2f}' if isinstance(Matchtime_4_0_Open, (int, float)) else f'A MS SKOR 4-0 | {Matchtime_4_0_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_40'])
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_41'] = streamlit.checkbox(f'A MS SKOR 4-1 | {Matchtime_4_1_Open:.2f}' if isinstance(Matchtime_4_1_Open, (int, float)) else f'A MS SKOR 4-1 | {Matchtime_4_1_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_41'])
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_42'] = streamlit.checkbox(f'A MS SKOR 4-2 | {Matchtime_4_2_Open:.2f}' if isinstance(Matchtime_4_2_Open, (int, float)) else f'A MS SKOR 4-2 | {Matchtime_4_2_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_42'])
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_43'] = streamlit.checkbox(f'A MS SKOR 4-3 | {Matchtime_4_3_Open:.2f}' if isinstance(Matchtime_4_3_Open, (int, float)) else f'A MS SKOR 4-3 | {Matchtime_4_3_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_43'])
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_00'] = streamlit.checkbox(f'A MS SKOR 0-0 | {Matchtime_0_0_Open:.2f}' if isinstance(Matchtime_0_0_Open, (int, float)) else f'A MS SKOR 0-0 | {Matchtime_0_0_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_00'])
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_11'] = streamlit.checkbox(f'A MS SKOR 1-1 | {Matchtime_1_1_Open:.2f}' if isinstance(Matchtime_1_1_Open, (int, float)) else f'A MS SKOR 1-1 | {Matchtime_1_1_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_11'])
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_22'] = streamlit.checkbox(f'A MS SKOR 2-2 | {Matchtime_2_2_Open:.2f}' if isinstance(Matchtime_2_2_Open, (int, float)) else f'A MS SKOR 2-2 | {Matchtime_2_2_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_22'])
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_33'] = streamlit.checkbox(f'A MS SKOR 3-3 | {Matchtime_3_3_Open:.2f}' if isinstance(Matchtime_3_3_Open, (int, float)) else f'A MS SKOR 3-3 | {Matchtime_3_3_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_33'])
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_44'] = streamlit.checkbox(f'A MS SKOR 4-4 | {Matchtime_4_4_Open:.2f}' if isinstance(Matchtime_4_4_Open, (int, float)) else f'A MS SKOR 4-4 | {Matchtime_4_4_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_44'])
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_01'] = streamlit.checkbox(f'A MS SKOR 0-1 | {Matchtime_0_1_Open:.2f}' if isinstance(Matchtime_0_1_Open, (int, float)) else f'A MS SKOR 0-1 | {Matchtime_0_1_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_01'])
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_02'] = streamlit.checkbox(f'A MS SKOR 0-2 | {Matchtime_0_2_Open:.2f}' if isinstance(Matchtime_0_2_Open, (int, float)) else f'A MS SKOR 0-2 | {Matchtime_0_2_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_02'])
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_12'] = streamlit.checkbox(f'A MS SKOR 1-2 | {Matchtime_1_2_Open:.2f}' if isinstance(Matchtime_1_2_Open, (int, float)) else f'A MS SKOR 1-2 | {Matchtime_1_2_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_12'])
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_03'] = streamlit.checkbox(f'A MS SKOR 0-3 | {Matchtime_0_3_Open:.2f}' if isinstance(Matchtime_0_3_Open, (int, float)) else f'A MS SKOR 0-3 | {Matchtime_0_3_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_03'])
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_13'] = streamlit.checkbox(f'A MS SKOR 1-3 | {Matchtime_1_3_Open:.2f}' if isinstance(Matchtime_1_3_Open, (int, float)) else f'A MS SKOR 1-3 | {Matchtime_1_3_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_13'])
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_23'] = streamlit.checkbox(f'A MS SKOR 2-3 | {Matchtime_2_3_Open:.2f}' if isinstance(Matchtime_2_3_Open, (int, float)) else f'A MS SKOR 2-3 | {Matchtime_2_3_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_23'])
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_04'] = streamlit.checkbox(f'A MS SKOR 0-4 | {Matchtime_0_4_Open:.2f}' if isinstance(Matchtime_0_4_Open, (int, float)) else f'A MS SKOR 0-4 | {Matchtime_0_4_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_04'])
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_14'] = streamlit.checkbox(f'A MS SKOR 1-4 | {Matchtime_1_4_Open:.2f}' if isinstance(Matchtime_1_4_Open, (int, float)) else f'A MS SKOR 1-4 | {Matchtime_1_4_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_14'])
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_24'] = streamlit.checkbox(f'A MS SKOR 2-4 | {Matchtime_2_4_Open:.2f}' if isinstance(Matchtime_2_4_Open, (int, float)) else f'A MS SKOR 2-4 | {Matchtime_2_4_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_24'])
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_34'] = streamlit.checkbox(f'A MS SKOR 3-4 | {Matchtime_3_4_Open:.2f}' if isinstance(Matchtime_3_4_Open, (int, float)) else f'A MS SKOR 3-4 | {Matchtime_3_4_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_34'])

                with streamlit.expander("İlk yarı asya handikap filtreleri"):
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_neg_2'] = streamlit.checkbox(f'A İY ASYA EV  -2.0 | {Firsthalf_Asian_Neg2_Home_Open:.2f}' if isinstance(Firsthalf_Asian_Neg2_Home_Open, (int, float)) else f'A İY ASYA EV  -2.0 | {Firsthalf_Asian_Neg2_Home_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_neg_2'])
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_neg_2'] = streamlit.checkbox(f'A İY ASYA DEP -2.0 | {Firsthalf_Asian_Neg2_Away_Open:.2f}' if isinstance(Firsthalf_Asian_Neg2_Away_Open, (int, float)) else f'A İY ASYA DEP -2.0 | {Firsthalf_Asian_Neg2_Away_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_neg_2'])
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_neg_2_neg_1_5'] = streamlit.checkbox(f'A İY ASYA EV  -2, -1.5 | {Firsthalf_Asian_Neg2_Neg1_5_Home_Open:.2f}' if isinstance(Firsthalf_Asian_Neg2_Neg1_5_Home_Open, (int, float)) else f'A İY ASYA EV  -2, -1.5 | {Firsthalf_Asian_Neg2_Neg1_5_Home_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_neg_2_neg_1_5'])
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_neg_2_neg_1_5'] = streamlit.checkbox(f'A İY ASYA DEP -2, -1.5 | {Firsthalf_Asian_Neg2_Neg1_5_Away_Open:.2f}' if isinstance(Firsthalf_Asian_Neg2_Neg1_5_Away_Open, (int, float)) else f'A İY ASYA DEP -2, -1.5 | {Firsthalf_Asian_Neg2_Neg1_5_Away_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_neg_2_neg_1_5'])
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_neg_1_5'] = streamlit.checkbox(f'A İY ASYA EV  -1.5 | {Firsthalf_Asian_Neg1_5_Home_Open:.2f}' if isinstance(Firsthalf_Asian_Neg1_5_Home_Open, (int, float)) else f'A İY ASYA EV  -1.5 | {Firsthalf_Asian_Neg1_5_Home_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_neg_1_5'])
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_neg_1_5'] = streamlit.checkbox(f'A İY ASYA DEP -1.5 | {Firsthalf_Asian_Neg1_5_Away_Open:.2f}' if isinstance(Firsthalf_Asian_Neg1_5_Away_Open, (int, float)) else f'A İY ASYA DEP -1.5 | {Firsthalf_Asian_Neg1_5_Away_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_neg_1_5'])
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_neg_1_5_neg_1'] = streamlit.checkbox(f'A İY ASYA EV  -1.5, -1.0 | {Firsthalf_Asian_Neg1_5_Neg1_Home_Open:.2f}' if isinstance(Firsthalf_Asian_Neg1_5_Neg1_Home_Open, (int, float)) else f'A İY ASYA EV  -1.5, -1.0 | {Firsthalf_Asian_Neg1_5_Neg1_Home_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_neg_1_5_neg_1'])
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_neg_1_5_neg_1'] = streamlit.checkbox(f'A İY ASYA DEP -1.5, -1.0 | {Firsthalf_Asian_Neg1_5_Neg1_Away_Open:.2f}' if isinstance(Firsthalf_Asian_Neg1_5_Neg1_Away_Open, (int, float)) else f'A İY ASYA DEP -1.5, -1.0 | {Firsthalf_Asian_Neg1_5_Neg1_Away_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_neg_1_5_neg_1'])
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_neg_1'] = streamlit.checkbox(f'A İY ASYA EV  -1.0 | {Firsthalf_Asian_Neg1_Home_Open:.2f}' if isinstance(Firsthalf_Asian_Neg1_Home_Open, (int, float)) else f'A İY ASYA EV  -1.0 | {Firsthalf_Asian_Neg1_Home_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_neg_1'])
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_neg_1'] = streamlit.checkbox(f'A İY ASYA DEP -1.0 | {Firsthalf_Asian_Neg1_Away_Open:.2f}' if isinstance(Firsthalf_Asian_Neg1_Away_Open, (int, float)) else f'A İY ASYA DEP -1.0 | {Firsthalf_Asian_Neg1_Away_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_neg_1'])
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_neg_1_neg_0_5'] = streamlit.checkbox(f'A İY ASYA EV  -1.0, -0.5 | {Firsthalf_Asian_Neg1_Neg0_5_Home_Open:.2f}' if isinstance(Firsthalf_Asian_Neg1_Neg0_5_Home_Open, (int, float)) else f'A İY ASYA EV  -1.0, -0.5 | {Firsthalf_Asian_Neg1_Neg0_5_Home_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_neg_1_neg_0_5'])
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_neg_1_neg_0_5'] = streamlit.checkbox(f'A İY ASYA DEP -1.0, -0.5 | {Firsthalf_Asian_Neg1_Neg0_5_Away_Open:.2f}' if isinstance(Firsthalf_Asian_Neg1_Neg0_5_Away_Open, (int, float)) else f'A İY ASYA DEP -1.0, -0.5 | {Firsthalf_Asian_Neg1_Neg0_5_Away_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_neg_1_neg_0_5'])
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_neg_0_5'] = streamlit.checkbox(f'A İY ASYA EV  -0.5 | {Firsthalf_Asian_Neg0_5_Home_Open:.2f}' if isinstance(Firsthalf_Asian_Neg0_5_Home_Open, (int, float)) else f'A İY ASYA EV  -0.5 | {Firsthalf_Asian_Neg0_5_Home_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_neg_0_5'])
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_neg_0_5'] = streamlit.checkbox(f'A İY ASYA DEP -0.5 | {Firsthalf_Asian_Neg0_5_Away_Open:.2f}' if isinstance(Firsthalf_Asian_Neg0_5_Away_Open, (int, float)) else f'A İY ASYA DEP -0.5 | {Firsthalf_Asian_Neg0_5_Away_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_neg_0_5'])
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_neg_0_5_pos_0'] = streamlit.checkbox(f'A İY ASYA EV  -0.5, 0.0 | {Firsthalf_Asian_Neg0_5_Neg0_0_Home_Open:.2f}' if isinstance(Firsthalf_Asian_Neg0_5_Neg0_0_Home_Open, (int, float)) else f'A İY ASYA EV  -0.5, 0.0 | {Firsthalf_Asian_Neg0_5_Neg0_0_Home_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_neg_0_5_pos_0'])
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_neg_0_5_pos_0'] = streamlit.checkbox(f'A İY ASYA DEP -0.5, 0.0 | {Firsthalf_Asian_Neg0_5_Neg0_0_Away_Open:.2f}' if isinstance(Firsthalf_Asian_Neg0_5_Neg0_0_Away_Open, (int, float)) else f'A İY ASYA DEP -0.5, 0.0 | {Firsthalf_Asian_Neg0_5_Neg0_0_Away_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_neg_0_5_pos_0'])
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_pos_0_pos_0_5'] = streamlit.checkbox(f'A İY ASYA EV  0.0, 0.5 | {Firsthalf_Asian_0_0_Pos0_5_Home_Open:.2f}' if isinstance(Firsthalf_Asian_0_0_Pos0_5_Home_Open, (int, float)) else f'A İY ASYA EV  0.0, 0.5 | {Firsthalf_Asian_0_0_Pos0_5_Home_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_pos_0_pos_0_5'])
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_pos_0_pos_0_5'] = streamlit.checkbox(f'A İY ASYA DEP 0.0, 0.5 | {Firsthalf_Asian_0_0_Pos0_5_Away_Open:.2f}' if isinstance(Firsthalf_Asian_0_0_Pos0_5_Away_Open, (int, float)) else f'A İY ASYA DEP 0.0, 0.5 | {Firsthalf_Asian_0_0_Pos0_5_Away_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_pos_0_pos_0_5'])
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_pos_0_5'] = streamlit.checkbox(f'A İY ASYA EV  0.5 | {Firsthalf_Asian_Pos0_5_Home_Open:.2f}' if isinstance(Firsthalf_Asian_Pos0_5_Home_Open, (int, float)) else f'A İY ASYA EV  0.5 | {Firsthalf_Asian_Pos0_5_Home_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_pos_0_5'])
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_pos_0_5'] = streamlit.checkbox(f'A İY ASYA DEP 0.5 | {Firsthalf_Asian_Pos0_5_Away_Open:.2f}' if isinstance(Firsthalf_Asian_Pos0_5_Away_Open, (int, float)) else f'A İY ASYA DEP 0.5 | {Firsthalf_Asian_Pos0_5_Away_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_pos_0_5'])
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_pos_0_5_pos_1'] = streamlit.checkbox(f'A İY ASYA EV  0.5, 1.0 | {Firsthalf_Asian_Pos0_5_Pos1_0_Home_Open:.2f}' if isinstance(Firsthalf_Asian_Pos0_5_Pos1_0_Home_Open, (int, float)) else f'A İY ASYA EV  0.5, 1.0 | {Firsthalf_Asian_Pos0_5_Pos1_0_Home_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_pos_0_5_pos_1'])
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_pos_0_5_pos_1'] = streamlit.checkbox(f'A İY ASYA DEP 0.5, 1.0 | {Firsthalf_Asian_Pos0_5_Pos1_0_Away_Open:.2f}' if isinstance(Firsthalf_Asian_Pos0_5_Pos1_0_Away_Open, (int, float)) else f'A İY ASYA DEP 0.5, 1.0 | {Firsthalf_Asian_Pos0_5_Pos1_0_Away_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_pos_0_5_pos_1'])
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_pos_1'] = streamlit.checkbox(f'A İY ASYA EV  1.0 | {Firsthalf_Asian_Pos1_Home_Open:.2f}' if isinstance(Firsthalf_Asian_Pos1_Home_Open, (int, float)) else f'A İY ASYA EV  1.0 | {Firsthalf_Asian_Pos1_Home_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_pos_1'])
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_pos_1'] = streamlit.checkbox(f'A İY ASYA DEP 1.0 | {Firsthalf_Asian_Pos1_Away_Open:.2f}' if isinstance(Firsthalf_Asian_Pos1_Away_Open, (int, float)) else f'A İY ASYA DEP 1.0 | {Firsthalf_Asian_Pos1_Away_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_pos_1'])
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_pos_1_pos_1_5'] = streamlit.checkbox(f'A İY ASYA EV  1.0, 1.5 | {Firsthalf_Asian_Pos1_Pos1_5_Home_Open:.2f}' if isinstance(Firsthalf_Asian_Pos1_Pos1_5_Home_Open, (int, float)) else f'A İY ASYA EV  1.0, 1.5 | {Firsthalf_Asian_Pos1_Pos1_5_Home_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_pos_1_pos_1_5'])
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_pos_1_pos_1_5'] = streamlit.checkbox(f'A İY ASYA DEP 1.0, 1.5 | {Firsthalf_Asian_Pos1_Pos1_5_Away_Open:.2f}' if isinstance(Firsthalf_Asian_Pos1_Pos1_5_Away_Open, (int, float)) else f'A İY ASYA DEP 1.0, 1.5 | {Firsthalf_Asian_Pos1_Pos1_5_Away_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_pos_1_pos_1_5'])
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_pos_1_5'] = streamlit.checkbox(f'A İY ASYA EV  1.5 | {Firsthalf_Asian_Pos1_5_Home_Open:.2f}' if isinstance(Firsthalf_Asian_Pos1_5_Home_Open, (int, float)) else f'A İY ASYA EV  1.5 | {Firsthalf_Asian_Pos1_5_Home_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_pos_1_5'])
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_pos_1_5'] = streamlit.checkbox(f'A İY ASYA DEP 1.5 | {Firsthalf_Asian_Pos1_5_Away_Open:.2f}' if isinstance(Firsthalf_Asian_Pos1_5_Away_Open, (int, float)) else f'A İY ASYA DEP 1.5 | {Firsthalf_Asian_Pos1_5_Away_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_pos_1_5'])
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_pos_1_5_pos_2'] = streamlit.checkbox(f'A İY ASYA EV  1.5, 2.0 | {Firsthalf_Asian_Pos1_5_Pos2_Home_Open:.2f}' if isinstance(Firsthalf_Asian_Pos1_5_Pos2_Home_Open, (int, float)) else f'A İY ASYA EV  1.5, 2.0 | {Firsthalf_Asian_Pos1_5_Pos2_Home_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_pos_1_5_pos_2'])
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_pos_1_5_pos_2'] = streamlit.checkbox(f'A İY ASYA DEP 1.5, 2.0 | {Firsthalf_Asian_Pos1_5_Pos2_Away_Open:.2f}' if isinstance(Firsthalf_Asian_Pos1_5_Pos2_Away_Open, (int, float)) else f'A İY ASYA DEP 1.5, 2.0 | {Firsthalf_Asian_Pos1_5_Pos2_Away_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_pos_1_5_pos_2'])
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_pos_2'] = streamlit.checkbox(f'A İY ASYA EV  2.0 | {Firsthalf_Asian_Pos2_Home_Open:.2f}' if isinstance(Firsthalf_Asian_Pos2_Home_Open, (int, float)) else f'A İY ASYA EV  2.0 | {Firsthalf_Asian_Pos2_Home_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_pos_2'])
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_pos_2'] = streamlit.checkbox(f'A İY ASYA DEP 2.0 | {Firsthalf_Asian_Pos2_Away_Open:.2f}' if isinstance(Firsthalf_Asian_Pos2_Away_Open, (int, float)) else f'A İY ASYA DEP 2.0 | {Firsthalf_Asian_Pos2_Away_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_pos_2'])

                with streamlit.expander("Maç sonu asya handikap filtreleri"):
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_neg_2'] = streamlit.checkbox(f'A MS ASYA EV  -2.0 | {Matchtime_Asian_Neg2_Home_Open:.2f}' if isinstance(Matchtime_Asian_Neg2_Home_Open, (int, float)) else f'A MS ASYA EV  -2.0 | {Matchtime_Asian_Neg2_Home_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_neg_2'])
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_neg_2'] = streamlit.checkbox(f'A MS ASYA DEP -2.0 | {Matchtime_Asian_Neg2_Away_Open:.2f}' if isinstance(Matchtime_Asian_Neg2_Away_Open, (int, float)) else f'A MS ASYA DEP -2.0 | {Matchtime_Asian_Neg2_Away_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_neg_2'])
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_neg_2_neg_1_5'] = streamlit.checkbox(f'A MS ASYA EV  -2, -1.5 | {Matchtime_Asian_Neg2_Neg1_5_Home_Open:.2f}' if isinstance(Matchtime_Asian_Neg2_Neg1_5_Home_Open, (int, float)) else f'A MS ASYA EV  -2, -1.5 | {Matchtime_Asian_Neg2_Neg1_5_Home_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_neg_2_neg_1_5'])
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_neg_2_neg_1_5'] = streamlit.checkbox(f'A MS ASYA DEP -2, -1.5 | {Matchtime_Asian_Neg2_Neg1_5_Away_Open:.2f}' if isinstance(Matchtime_Asian_Neg2_Neg1_5_Away_Open, (int, float)) else f'A MS ASYA DEP -2, -1.5 | {Matchtime_Asian_Neg2_Neg1_5_Away_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_neg_2_neg_1_5'])
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_neg_1_5'] = streamlit.checkbox(f'A MS ASYA EV  -1.5 | {Matchtime_Asian_Neg1_5_Home_Open:.2f}' if isinstance(Matchtime_Asian_Neg1_5_Home_Open, (int, float)) else f'A MS ASYA EV  -1.5 | {Matchtime_Asian_Neg1_5_Home_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_neg_1_5'])
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_neg_1_5'] = streamlit.checkbox(f'A MS ASYA DEP -1.5 | {Matchtime_Asian_Neg1_5_Away_Open:.2f}' if isinstance(Matchtime_Asian_Neg1_5_Away_Open, (int, float)) else f'A MS ASYA DEP -1.5 | {Matchtime_Asian_Neg1_5_Away_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_neg_1_5'])
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_neg_1_5_neg_1'] = streamlit.checkbox(f'A MS ASYA EV  -1.5, -1.0 | {Matchtime_Asian_Neg1_5_Neg1_Home_Open:.2f}' if isinstance(Matchtime_Asian_Neg1_5_Neg1_Home_Open, (int, float)) else f'A MS ASYA EV  -1.5, -1.0 | {Matchtime_Asian_Neg1_5_Neg1_Home_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_neg_1_5_neg_1'])
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_neg_1_5_neg_1'] = streamlit.checkbox(f'A MS ASYA DEP -1.5, -1.0 | {Matchtime_Asian_Neg1_5_Neg1_Away_Open:.2f}' if isinstance(Matchtime_Asian_Neg1_5_Neg1_Away_Open, (int, float)) else f'A MS ASYA DEP -1.5, -1.0 | {Matchtime_Asian_Neg1_5_Neg1_Away_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_neg_1_5_neg_1'])
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_neg_1'] = streamlit.checkbox(f'A MS ASYA EV  -1.0 | {Matchtime_Asian_Neg1_Home_Open:.2f}' if isinstance(Matchtime_Asian_Neg1_Home_Open, (int, float)) else f'A MS ASYA EV  -1.0 | {Matchtime_Asian_Neg1_Home_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_neg_1'])
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_neg_1'] = streamlit.checkbox(f'A MS ASYA DEP -1.0 | {Matchtime_Asian_Neg1_Away_Open:.2f}' if isinstance(Matchtime_Asian_Neg1_Away_Open, (int, float)) else f'A MS ASYA DEP -1.0 | {Matchtime_Asian_Neg1_Away_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_neg_1'])
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_neg_1_neg_0_5'] = streamlit.checkbox(f'A MS ASYA EV  -1.0, -0.5 | {Matchtime_Asian_Neg1_Neg0_5_Home_Open:.2f}' if isinstance(Matchtime_Asian_Neg1_Neg0_5_Home_Open, (int, float)) else f'A MS ASYA EV  -1.0, -0.5 | {Matchtime_Asian_Neg1_Neg0_5_Home_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_neg_1_neg_0_5'])
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_neg_1_neg_0_5'] = streamlit.checkbox(f'A MS ASYA DEP -1.0, -0.5 | {Matchtime_Asian_Neg1_Neg0_5_Away_Open:.2f}' if isinstance(Matchtime_Asian_Neg1_Neg0_5_Away_Open, (int, float)) else f'A MS ASYA DEP -1.0, -0.5 | {Matchtime_Asian_Neg1_Neg0_5_Away_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_neg_1_neg_0_5'])
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_neg_0_5'] = streamlit.checkbox(f'A MS ASYA EV  -0.5 | {Matchtime_Asian_Neg0_5_Home_Open:.2f}' if isinstance(Matchtime_Asian_Neg0_5_Home_Open, (int, float)) else f'A MS ASYA EV  -0.5 | {Matchtime_Asian_Neg0_5_Home_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_neg_0_5'])
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_neg_0_5'] = streamlit.checkbox(f'A MS ASYA DEP -0.5 | {Matchtime_Asian_Neg0_5_Away_Open:.2f}' if isinstance(Matchtime_Asian_Neg0_5_Away_Open, (int, float)) else f'A MS ASYA DEP -0.5 | {Matchtime_Asian_Neg0_5_Away_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_neg_0_5'])
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_neg_0_5_pos_0'] = streamlit.checkbox(f'A MS ASYA EV  -0.5, 0.0 | {Matchtime_Asian_Neg0_5_Neg0_0_Home_Open:.2f}' if isinstance(Matchtime_Asian_Neg0_5_Neg0_0_Home_Open, (int, float)) else f'A MS ASYA EV  -0.5, 0.0 | {Matchtime_Asian_Neg0_5_Neg0_0_Home_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_neg_0_5_pos_0'])
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_neg_0_5_pos_0'] = streamlit.checkbox(f'A MS ASYA DEP -0.5, 0.0 | {Matchtime_Asian_Neg0_5_Neg0_0_Away_Open:.2f}' if isinstance(Matchtime_Asian_Neg0_5_Neg0_0_Away_Open, (int, float)) else f'A MS ASYA DEP -0.5, 0.0 | {Matchtime_Asian_Neg0_5_Neg0_0_Away_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_neg_0_5_pos_0'])
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_pos_0_pos_0_5'] = streamlit.checkbox(f'A MS ASYA EV  0.0, 0.5 | {Matchtime_Asian_0_0_Pos0_5_Home_Open:.2f}' if isinstance(Matchtime_Asian_0_0_Pos0_5_Home_Open, (int, float)) else f'A MS ASYA EV  0.0, 0.5 | {Matchtime_Asian_0_0_Pos0_5_Home_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_pos_0_pos_0_5'])
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_pos_0_pos_0_5'] = streamlit.checkbox(f'A MS ASYA DEP 0.0, 0.5 | {Matchtime_Asian_0_0_Pos0_5_Away_Open:.2f}' if isinstance(Matchtime_Asian_0_0_Pos0_5_Away_Open, (int, float)) else f'A MS ASYA DEP 0.0, 0.5 | {Matchtime_Asian_0_0_Pos0_5_Away_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_pos_0_pos_0_5'])
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_pos_0_5'] = streamlit.checkbox(f'A MS ASYA EV  0.5 | {Matchtime_Asian_Pos0_5_Home_Open:.2f}' if isinstance(Matchtime_Asian_Pos0_5_Home_Open, (int, float)) else f'A MS ASYA EV  0.5 | {Matchtime_Asian_Pos0_5_Home_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_pos_0_5'])
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_pos_0_5'] = streamlit.checkbox(f'A MS ASYA DEP 0.5 | {Matchtime_Asian_Pos0_5_Away_Open:.2f}' if isinstance(Matchtime_Asian_Pos0_5_Away_Open, (int, float)) else f'A MS ASYA DEP 0.5 | {Matchtime_Asian_Pos0_5_Away_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_pos_0_5'])
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_pos_0_5_pos_1'] = streamlit.checkbox(f'A MS ASYA EV  0.5, 1.0 | {Matchtime_Asian_Pos0_5_Pos1_0_Home_Open:.2f}' if isinstance(Matchtime_Asian_Pos0_5_Pos1_0_Home_Open, (int, float)) else f'A MS ASYA EV  0.5, 1.0 | {Matchtime_Asian_Pos0_5_Pos1_0_Home_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_pos_0_5_pos_1'])
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_pos_0_5_pos_1'] = streamlit.checkbox(f'A MS ASYA DEP 0.5, 1.0 | {Matchtime_Asian_Pos0_5_Pos1_0_Away_Open:.2f}' if isinstance(Matchtime_Asian_Pos0_5_Pos1_0_Away_Open, (int, float)) else f'A MS ASYA DEP 0.5, 1.0 | {Matchtime_Asian_Pos0_5_Pos1_0_Away_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_pos_0_5_pos_1'])
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_pos_1'] = streamlit.checkbox(f'A MS ASYA EV  1.0 | {Matchtime_Asian_Pos1_Home_Open:.2f}' if isinstance(Matchtime_Asian_Pos1_Home_Open, (int, float)) else f'A MS ASYA EV  1.0 | {Matchtime_Asian_Pos1_Home_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_pos_1'])
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_pos_1'] = streamlit.checkbox(f'A MS ASYA DEP 1.0 | {Matchtime_Asian_Pos1_Away_Open:.2f}' if isinstance(Matchtime_Asian_Pos1_Away_Open, (int, float)) else f'A MS ASYA DEP 1.0 | {Matchtime_Asian_Pos1_Away_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_pos_1'])
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_pos_1_pos_1_5'] = streamlit.checkbox(f'A MS ASYA EV  1.0, 1.5 | {Matchtime_Asian_Pos1_Pos1_5_Home_Open:.2f}' if isinstance(Matchtime_Asian_Pos1_Pos1_5_Home_Open, (int, float)) else f'A MS ASYA EV  1.0, 1.5 | {Matchtime_Asian_Pos1_Pos1_5_Home_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_pos_1_pos_1_5'])
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_pos_1_pos_1_5'] = streamlit.checkbox(f'A MS ASYA DEP 1.0, 1.5 | {Matchtime_Asian_Pos1_Pos1_5_Away_Open:.2f}' if isinstance(Matchtime_Asian_Pos1_Pos1_5_Away_Open, (int, float)) else f'A MS ASYA DEP 1.0, 1.5 | {Matchtime_Asian_Pos1_Pos1_5_Away_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_pos_1_pos_1_5'])
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_pos_1_5'] = streamlit.checkbox(f'A MS ASYA EV  1.5 | {Matchtime_Asian_Pos1_5_Home_Open:.2f}' if isinstance(Matchtime_Asian_Pos1_5_Home_Open, (int, float)) else f'A MS ASYA EV  1.5 | {Matchtime_Asian_Pos1_5_Home_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_pos_1_5'])
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_pos_1_5'] = streamlit.checkbox(f'A MS ASYA DEP 1.5 | {Matchtime_Asian_Pos1_5_Away_Open:.2f}' if isinstance(Matchtime_Asian_Pos1_5_Away_Open, (int, float)) else f'A MS ASYA DEP 1.5 | {Matchtime_Asian_Pos1_5_Away_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_pos_1_5'])
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_pos_1_5_pos_2'] = streamlit.checkbox(f'A MS ASYA EV  1.5, 2.0 | {Matchtime_Asian_Pos1_5_Pos2_Home_Open:.2f}' if isinstance(Matchtime_Asian_Pos1_5_Pos2_Home_Open, (int, float)) else f'A MS ASYA EV  1.5, 2.0 | {Matchtime_Asian_Pos1_5_Pos2_Home_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_pos_1_5_pos_2'])
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_pos_1_5_pos_2'] = streamlit.checkbox(f'A MS ASYA DEP 1.5, 2.0 | {Matchtime_Asian_Pos1_5_Pos2_Away_Open:.2f}' if isinstance(Matchtime_Asian_Pos1_5_Pos2_Away_Open, (int, float)) else f'A MS ASYA DEP 1.5, 2.0 | {Matchtime_Asian_Pos1_5_Pos2_Away_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_pos_1_5_pos_2'])
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_pos_2'] = streamlit.checkbox(f'A MS ASYA EV  2.0 | {Matchtime_Asian_Pos2_Home_Open:.2f}' if isinstance(Matchtime_Asian_Pos2_Home_Open, (int, float)) else f'A MS ASYA EV  2.0 | {Matchtime_Asian_Pos2_Home_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_pos_2'])
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_pos_2'] = streamlit.checkbox(f'A MS ASYA DEP 2.0 | {Matchtime_Asian_Pos2_Away_Open:.2f}' if isinstance(Matchtime_Asian_Pos2_Away_Open, (int, float)) else f'A MS ASYA DEP 2.0 | {Matchtime_Asian_Pos2_Away_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_pos_2'])

                with streamlit.expander("İlk yarı avrupa handikap"):
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ev_neg_1'] = streamlit.checkbox(f'A İY AVRUPA EV  -1.0 | {Firsthalf_European_Neg1_Home_Open:.2f}' if isinstance(Firsthalf_European_Neg1_Home_Open, (int, float)) else f'A İY AVRUPA EV  -1.0 | {Firsthalf_European_Neg1_Home_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ev_neg_1'])
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ber_neg_1'] = streamlit.checkbox(f'A İY AVRUPA BER -1.0 | {Firsthalf_European_Neg1_Tie_Open:.2f}' if isinstance(Firsthalf_European_Neg1_Tie_Open, (int, float)) else f'A İY AVRUPA BER -1.0 | {Firsthalf_European_Neg1_Tie_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ber_neg_1'])
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_dep_neg_1'] = streamlit.checkbox(f'A İY AVRUPA DEP -1.0 | {Firsthalf_European_Neg1_Away_Open:.2f}' if isinstance(Firsthalf_European_Neg1_Away_Open, (int, float)) else f'A İY AVRUPA DEP -1.0 | {Firsthalf_European_Neg1_Away_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_dep_neg_1'])
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ev_neg_1_neg_0_5'] = streamlit.checkbox(f'A İY AVRUPA EV  -1.0, -0.5 | {Firsthalf_European_Neg1_Neg0_5_Home_Open:.2f}' if isinstance(Firsthalf_European_Neg1_Neg0_5_Home_Open, (int, float)) else f'A İY AVRUPA EV  -1.0, -0.5 | {Firsthalf_European_Neg1_Neg0_5_Home_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ev_neg_1_neg_0_5'])
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ber_neg_1_neg_0_5'] = streamlit.checkbox(f'A İY AVRUPA BER -1.0, -0.5 | {Firsthalf_European_Neg1_Neg0_5_Tie_Open:.2f}' if isinstance(Firsthalf_European_Neg1_Neg0_5_Tie_Open, (int, float)) else f'A İY AVRUPA BER -1.0, -0.5 | {Firsthalf_European_Neg1_Neg0_5_Tie_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ber_neg_1_neg_0_5'])
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_dep_neg_1_neg_0_5'] = streamlit.checkbox(f'A İY AVRUPA DEP -1.0, -0.5 | {Firsthalf_European_Neg1_Neg0_5_Away_Open:.2f}' if isinstance(Firsthalf_European_Neg1_Neg0_5_Away_Open, (int, float)) else f'A İY AVRUPA DEP -1.0, -0.5 | {Firsthalf_European_Neg1_Neg0_5_Away_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_dep_neg_1_neg_0_5'])
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ev_neg_0_5'] = streamlit.checkbox(f'A İY AVRUPA EV  -0.5 | {Firsthalf_European_Neg0_5_Home_Open:.2f}' if isinstance(Firsthalf_European_Neg0_5_Home_Open, (int, float)) else f'A İY AVRUPA EV  -0.5 | {Firsthalf_European_Neg0_5_Home_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ev_neg_0_5'])
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ber_neg_0_5'] = streamlit.checkbox(f'A İY AVRUPA BER -0.5 | {Firsthalf_European_Neg0_5_Tie_Open:.2f}' if isinstance(Firsthalf_European_Neg0_5_Tie_Open, (int, float)) else f'A İY AVRUPA BER -0.5 | {Firsthalf_European_Neg0_5_Tie_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ber_neg_0_5'])
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_dep_neg_0_5'] = streamlit.checkbox(f'A İY AVRUPA DEP -0.5 | {Firsthalf_European_Neg0_5_Away_Open:.2f}' if isinstance(Firsthalf_European_Neg0_5_Away_Open, (int, float)) else f'A İY AVRUPA DEP -0.5 | {Firsthalf_European_Neg0_5_Away_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_dep_neg_0_5'])
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ev_neg_0_5_pos_0'] = streamlit.checkbox(f'A İY AVRUPA EV  -0.5, 0.0 | {Firsthalf_European_Neg0_5_0_0_Home_Open:.2f}' if isinstance(Firsthalf_European_Neg0_5_0_0_Home_Open, (int, float)) else f'A İY AVRUPA EV  -0.5, 0.0 | {Firsthalf_European_Neg0_5_0_0_Home_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ev_neg_0_5_pos_0'])
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ber_neg_0_5_pos_0'] = streamlit.checkbox(f'A İY AVRUPA BER -0.5, 0.0 | {Firsthalf_European_Neg0_5_0_0_Tie_Open:.2f}' if isinstance(Firsthalf_European_Neg0_5_0_0_Tie_Open, (int, float)) else f'A İY AVRUPA BER -0.5, 0.0 | {Firsthalf_European_Neg0_5_0_0_Tie_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ber_neg_0_5_pos_0'])
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_dep_neg_0_5_pos_0'] = streamlit.checkbox(f'A İY AVRUPA DEP -0.5, 0.0 | {Firsthalf_European_Neg0_5_0_0_Away_Open:.2f}' if isinstance(Firsthalf_European_Neg0_5_0_0_Away_Open, (int, float)) else f'A İY AVRUPA DEP -0.5, 0.0 | {Firsthalf_European_Neg0_5_0_0_Away_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_dep_neg_0_5_pos_0'])
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ev_pos_0_pos_0_5'] = streamlit.checkbox(f'A İY AVRUPA EV  0.0, 0.5 | {Firsthalf_European_0_0_Pos0_5_Home_Open:.2f}' if isinstance(Firsthalf_European_0_0_Pos0_5_Home_Open, (int, float)) else f'A İY AVRUPA EV  0.0, 0.5 | {Firsthalf_European_0_0_Pos0_5_Home_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ev_pos_0_pos_0_5'])
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ber_pos_0_pos_0_5'] = streamlit.checkbox(f'A İY AVRUPA BER 0.0, 0.5 | {Firsthalf_European_0_0_Pos0_5_Tie_Open:.2f}' if isinstance(Firsthalf_European_0_0_Pos0_5_Tie_Open, (int, float)) else f'A İY AVRUPA BER 0.0, 0.5 | {Firsthalf_European_0_0_Pos0_5_Tie_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ber_pos_0_pos_0_5'])
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_dep_pos_0_pos_0_5'] = streamlit.checkbox(f'A İY AVRUPA DEP 0.0, 0.5 | {Firsthalf_European_0_0_Pos0_5_Away_Open:.2f}' if isinstance(Firsthalf_European_0_0_Pos0_5_Away_Open, (int, float)) else f'A İY AVRUPA DEP 0.0, 0.5 | {Firsthalf_European_0_0_Pos0_5_Away_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_dep_pos_0_pos_0_5'])
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ev_pos_0_5'] = streamlit.checkbox(f'A İY AVRUPA EV  0.5 | {Firsthalf_European_Pos0_5_Home_Open:.2f}' if isinstance(Firsthalf_European_Pos0_5_Home_Open, (int, float)) else f'A İY AVRUPA EV  0.5 | {Firsthalf_European_Pos0_5_Home_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ev_pos_0_5'])
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ber_pos_0_5'] = streamlit.checkbox(f'A İY AVRUPA BER 0.5 | {Firsthalf_European_Pos0_5_Tie_Open:.2f}' if isinstance(Firsthalf_European_Pos0_5_Tie_Open, (int, float)) else f'A İY AVRUPA BER 0.5 | {Firsthalf_European_Pos0_5_Tie_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ber_pos_0_5'])
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_dep_pos_0_5'] = streamlit.checkbox(f'A İY AVRUPA DEP 0.5 | {Firsthalf_European_Pos0_5_Away_Open:.2f}' if isinstance(Firsthalf_European_Pos0_5_Away_Open, (int, float)) else f'A İY AVRUPA DEP 0.5 | {Firsthalf_European_Pos0_5_Away_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_dep_pos_0_5'])
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ev_pos_0_5_pos_1'] = streamlit.checkbox(f'A İY AVRUPA EV  0.5, 1.0 | {Firsthalf_European_Pos0_5_Pos1_0_Home_Open:.2f}' if isinstance(Firsthalf_European_Pos0_5_Pos1_0_Home_Open, (int, float)) else f'A İY AVRUPA EV  0.5, 1.0 | {Firsthalf_European_Pos0_5_Pos1_0_Home_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ev_pos_0_5_pos_1'])
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ber_pos_0_5_pos_1'] = streamlit.checkbox(f'A İY AVRUPA BER 0.5, 1.0 | {Firsthalf_European_Pos0_5_Pos1_0_Tie_Open:.2f}' if isinstance(Firsthalf_European_Pos0_5_Pos1_0_Tie_Open, (int, float)) else f'A İY AVRUPA BER 0.5, 1.0 | {Firsthalf_European_Pos0_5_Pos1_0_Tie_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ber_pos_0_5_pos_1'])
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_dep_pos_0_5_pos_1'] = streamlit.checkbox(f'A İY AVRUPA DEP 0.5, 1.0 | {Firsthalf_European_Pos0_5_Pos1_0_Away_Open:.2f}' if isinstance(Firsthalf_European_Pos0_5_Pos1_0_Away_Open, (int, float)) else f'A İY AVRUPA DEP 0.5, 1.0 | {Firsthalf_European_Pos0_5_Pos1_0_Away_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_dep_pos_0_5_pos_1'])
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ev_pos_1'] = streamlit.checkbox(f'A İY AVRUPA EV  1.0 | {Firsthalf_European_Pos1_0_Home_Open:.2f}' if isinstance(Firsthalf_European_Pos1_0_Home_Open, (int, float)) else f'A İY AVRUPA EV  1.0 | {Firsthalf_European_Pos1_0_Home_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ev_pos_1'])
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ber_pos_1'] = streamlit.checkbox(f'A İY AVRUPA BER 1.0 | {Firsthalf_European_Pos1_0_Tie_Open:.2f}' if isinstance(Firsthalf_European_Pos1_0_Tie_Open, (int, float)) else f'A İY AVRUPA BER 1.0 | {Firsthalf_European_Pos1_0_Tie_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ber_pos_1'])
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_dep_pos_1'] = streamlit.checkbox(f'A İY AVRUPA DEP 1.0 | {Firsthalf_European_Pos1_0_Away_Open:.2f}' if isinstance(Firsthalf_European_Pos1_0_Away_Open, (int, float)) else f'A İY AVRUPA DEP 1.0 | {Firsthalf_European_Pos1_0_Away_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_dep_pos_1'])

                with streamlit.expander("Maç sonu avrupa handikap"):
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ev_neg_1'] = streamlit.checkbox(f'A MS AVRUPA EV  -1.0 | {Matchtime_European_Neg1_Home_Open:.2f}' if isinstance(Matchtime_European_Neg1_Home_Open, (int, float)) else f'A MS AVRUPA EV  -1.0 | {Matchtime_European_Neg1_Home_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ev_neg_1'])
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ber_neg_1'] = streamlit.checkbox(f'A MS AVRUPA BER -1.0 | {Matchtime_European_Neg1_Tie_Open:.2f}' if isinstance(Matchtime_European_Neg1_Tie_Open, (int, float)) else f'A MS AVRUPA BER -1.0 | {Matchtime_European_Neg1_Tie_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ber_neg_1'])
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_dep_neg_1'] = streamlit.checkbox(f'A MS AVRUPA DEP -1.0 | {Matchtime_European_Neg1_Away_Open:.2f}' if isinstance(Matchtime_European_Neg1_Away_Open, (int, float)) else f'A MS AVRUPA DEP -1.0 | {Matchtime_European_Neg1_Away_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_dep_neg_1'])
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ev_neg_1_neg_0_5'] = streamlit.checkbox(f'A MS AVRUPA EV  -1.0, -0.5 | {Matchtime_European_Neg1_Neg0_5_Home_Open:.2f}' if isinstance(Matchtime_European_Neg1_Neg0_5_Home_Open, (int, float)) else f'A MS AVRUPA EV  -1.0, -0.5 | {Matchtime_European_Neg1_Neg0_5_Home_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ev_neg_1_neg_0_5'])
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ber_neg_1_neg_0_5'] = streamlit.checkbox(f'A MS AVRUPA BER -1.0, -0.5 | {Matchtime_European_Neg1_Neg0_5_Tie_Open:.2f}' if isinstance(Matchtime_European_Neg1_Neg0_5_Tie_Open, (int, float)) else f'A MS AVRUPA BER -1.0, -0.5 | {Matchtime_European_Neg1_Neg0_5_Tie_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ber_neg_1_neg_0_5'])
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_dep_neg_1_neg_0_5'] = streamlit.checkbox(f'A MS AVRUPA DEP -1.0, -0.5 | {Matchtime_European_Neg1_Neg0_5_Away_Open:.2f}' if isinstance(Matchtime_European_Neg1_Neg0_5_Away_Open, (int, float)) else f'A MS AVRUPA DEP -1.0, -0.5 | {Matchtime_European_Neg1_Neg0_5_Away_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_dep_neg_1_neg_0_5'])
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ev_neg_0_5'] = streamlit.checkbox(f'A MS AVRUPA EV  -0.5 | {Matchtime_European_Neg0_5_Home_Open:.2f}' if isinstance(Matchtime_European_Neg0_5_Home_Open, (int, float)) else f'A MS AVRUPA EV  -0.5 | {Matchtime_European_Neg0_5_Home_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ev_neg_0_5'])
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ber_neg_0_5'] = streamlit.checkbox(f'A MS AVRUPA BER -0.5 | {Matchtime_European_Neg0_5_Tie_Open:.2f}' if isinstance(Matchtime_European_Neg0_5_Tie_Open, (int, float)) else f'A MS AVRUPA BER -0.5 | {Matchtime_European_Neg0_5_Tie_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ber_neg_0_5'])
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_dep_neg_0_5'] = streamlit.checkbox(f'A MS AVRUPA DEP -0.5 | {Matchtime_European_Neg0_5_Away_Open:.2f}' if isinstance(Matchtime_European_Neg0_5_Away_Open, (int, float)) else f'A MS AVRUPA DEP -0.5 | {Matchtime_European_Neg0_5_Away_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_dep_neg_0_5'])
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ev_neg_0_5_pos_0'] = streamlit.checkbox(f'A MS AVRUPA EV  -0.5, 0.0 | {Matchtime_European_Neg0_5_0_0_Home_Open:.2f}' if isinstance(Matchtime_European_Neg0_5_0_0_Home_Open, (int, float)) else f'A MS AVRUPA EV  -0.5, 0.0 | {Matchtime_European_Neg0_5_0_0_Home_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ev_neg_0_5_pos_0'])
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ber_neg_0_5_pos_0'] = streamlit.checkbox(f'A MS AVRUPA BER -0.5, 0.0 | {Matchtime_European_Neg0_5_0_0_Tie_Open:.2f}' if isinstance(Matchtime_European_Neg0_5_0_0_Tie_Open, (int, float)) else f'A MS AVRUPA BER -0.5, 0.0 | {Matchtime_European_Neg0_5_0_0_Tie_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ber_neg_0_5_pos_0'])
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_dep_neg_0_5_pos_0'] = streamlit.checkbox(f'A MS AVRUPA DEP -0.5, 0.0 | {Matchtime_European_Neg0_5_0_0_Away_Open:.2f}' if isinstance(Matchtime_European_Neg0_5_0_0_Away_Open, (int, float)) else f'A MS AVRUPA DEP -0.5, 0.0 | {Matchtime_European_Neg0_5_0_0_Away_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_dep_neg_0_5_pos_0'])
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ev_pos_0_pos_0_5'] = streamlit.checkbox(f'A MS AVRUPA EV  0.0, 0.5 | {Matchtime_European_0_0_Pos0_5_Home_Open:.2f}' if isinstance(Matchtime_European_0_0_Pos0_5_Home_Open, (int, float)) else f'A MS AVRUPA EV  0.0, 0.5 | {Matchtime_European_0_0_Pos0_5_Home_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ev_pos_0_pos_0_5'])
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ber_pos_0_pos_0_5'] = streamlit.checkbox(f'A MS AVRUPA BER 0.0, 0.5 | {Matchtime_European_0_0_Pos0_5_Tie_Open:.2f}' if isinstance(Matchtime_European_0_0_Pos0_5_Tie_Open, (int, float)) else f'A MS AVRUPA BER 0.0, 0.5 | {Matchtime_European_0_0_Pos0_5_Tie_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ber_pos_0_pos_0_5'])
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_dep_pos_0_pos_0_5'] = streamlit.checkbox(f'A MS AVRUPA DEP 0.0, 0.5 | {Matchtime_European_0_0_Pos0_5_Away_Open:.2f}' if isinstance(Matchtime_European_0_0_Pos0_5_Away_Open, (int, float)) else f'A MS AVRUPA DEP 0.0, 0.5 | {Matchtime_European_0_0_Pos0_5_Away_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_dep_pos_0_pos_0_5'])
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ev_pos_0_5'] = streamlit.checkbox(f'A MS AVRUPA EV  0.5 | {Matchtime_European_Pos0_5_Home_Open:.2f}' if isinstance(Matchtime_European_Pos0_5_Home_Open, (int, float)) else f'A MS AVRUPA EV  0.5 | {Matchtime_European_Pos0_5_Home_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ev_pos_0_5'])
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ber_pos_0_5'] = streamlit.checkbox(f'A MS AVRUPA BER 0.5 | {Matchtime_European_Pos0_5_Tie_Open:.2f}' if isinstance(Matchtime_European_Pos0_5_Tie_Open, (int, float)) else f'A MS AVRUPA BER 0.5 | {Matchtime_European_Pos0_5_Tie_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ber_pos_0_5'])
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_dep_pos_0_5'] = streamlit.checkbox(f'A MS AVRUPA DEP 0.5 | {Matchtime_European_Pos0_5_Away_Open:.2f}' if isinstance(Matchtime_European_Pos0_5_Away_Open, (int, float)) else f'A MS AVRUPA DEP 0.5 | {Matchtime_European_Pos0_5_Away_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_dep_pos_0_5'])
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ev_pos_0_5_pos_1'] = streamlit.checkbox(f'A MS AVRUPA EV  0.5, 1.0 | {Matchtime_European_Pos0_5_Pos1_0_Home_Open:.2f}' if isinstance(Matchtime_European_Pos0_5_Pos1_0_Home_Open, (int, float)) else f'A MS AVRUPA EV  0.5, 1.0 | {Matchtime_European_Pos0_5_Pos1_0_Home_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ev_pos_0_5_pos_1'])
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ber_pos_0_5_pos_1'] = streamlit.checkbox(f'A MS AVRUPA BER 0.5, 1.0 | {Matchtime_European_Pos0_5_Pos1_0_Tie_Open:.2f}' if isinstance(Matchtime_European_Pos0_5_Pos1_0_Tie_Open, (int, float)) else f'A MS AVRUPA BER 0.5, 1.0 | {Matchtime_European_Pos0_5_Pos1_0_Tie_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ber_pos_0_5_pos_1'])
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_dep_pos_0_5_pos_1'] = streamlit.checkbox(f'A MS AVRUPA DEP 0.5, 1.0 | {Matchtime_European_Pos0_5_Pos1_0_Away_Open:.2f}' if isinstance(Matchtime_European_Pos0_5_Pos1_0_Away_Open, (int, float)) else f'A MS AVRUPA DEP 0.5, 1.0 | {Matchtime_European_Pos0_5_Pos1_0_Away_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_dep_pos_0_5_pos_1'])
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ev_pos_1'] = streamlit.checkbox(f'A MS AVRUPA EV  1.0 | {Matchtime_European_Pos1_0_Home_Open:.2f}' if isinstance(Matchtime_European_Pos1_0_Home_Open, (int, float)) else f'A MS AVRUPA EV  1.0 | {Matchtime_European_Pos1_0_Home_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ev_pos_1'])
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ber_pos_1'] = streamlit.checkbox(f'A MS AVRUPA BER 1.0 | {Matchtime_European_Pos1_0_Tie_Open:.2f}' if isinstance(Matchtime_European_Pos1_0_Tie_Open, (int, float)) else f'A MS AVRUPA BER 1.0 | {Matchtime_European_Pos1_0_Tie_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ber_pos_1'])
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_dep_pos_1'] = streamlit.checkbox(f'A MS AVRUPA DEP 1.0 | {Matchtime_European_Pos1_0_Away_Open:.2f}' if isinstance(Matchtime_European_Pos1_0_Away_Open, (int, float)) else f'A MS AVRUPA DEP 1.0 | {Matchtime_European_Pos1_0_Away_Open}', value=streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_dep_pos_1'])
            with tab4:
                with streamlit.expander("Maç sonucu filtreleri"):
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_1'] = streamlit.checkbox(f'K MS 1 | {Fulltime_Home_Close:.2f}' if isinstance(Fulltime_Home_Close, (int, float)) else f'K MS 1 | {Fulltime_Home_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_ms_1'])
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_x'] = streamlit.checkbox(f'K MS X | {Fulltime_Tie_Close:.2f}' if isinstance(Fulltime_Tie_Close, (int, float)) else f'K MS X | {Fulltime_Tie_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_ms_x'])
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_2'] = streamlit.checkbox(f'K MS 2 | {Fulltime_Away_Close:.2f}' if isinstance(Fulltime_Away_Close, (int, float)) else f'K MS 2 | {Fulltime_Away_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_ms_2'])
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_0_5_a'] = streamlit.checkbox(f'K MS 0.5 A | {Matchtime_Under_0_5_Close:.2f}' if isinstance(Matchtime_Under_0_5_Close, (int, float)) else f'K MS 0.5 A | {Matchtime_Under_0_5_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_ms_0_5_a'])
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_0_5_ü'] = streamlit.checkbox(f'K MS 0.5 Ü | {Matchtime_Over_0_5_Close:.2f}' if isinstance(Matchtime_Over_0_5_Close, (int, float)) else f'K MS 0.5 Ü | {Matchtime_Over_0_5_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_ms_0_5_ü'])
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_1_5_a'] = streamlit.checkbox(f'K MS 1.5 A | {Matchtime_Under_1_5_Close:.2f}' if isinstance(Matchtime_Under_1_5_Close, (int, float)) else f'K MS 1.5 A | {Matchtime_Under_1_5_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_ms_1_5_a'])
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_1_5_ü'] = streamlit.checkbox(f'K MS 1.5 Ü | {Matchtime_Over_1_5_Close:.2f}' if isinstance(Matchtime_Over_1_5_Close, (int, float)) else f'K MS 1.5 Ü | {Matchtime_Over_1_5_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_ms_1_5_ü'])
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_2_5_a'] = streamlit.checkbox(f'K MS 2.5 A | {Matchtime_Under_2_5_Close:.2f}' if isinstance(Matchtime_Under_2_5_Close, (int, float)) else f'K MS 2.5 A | {Matchtime_Under_2_5_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_ms_2_5_a'])
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_2_5_ü'] = streamlit.checkbox(f'K MS 2.5 Ü | {Matchtime_Over_2_5_Close:.2f}' if isinstance(Matchtime_Over_2_5_Close, (int, float)) else f'K MS 2.5 Ü | {Matchtime_Over_2_5_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_ms_2_5_ü'])
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_3_5_a'] = streamlit.checkbox(f'K MS 3.5 A | {Matchtime_Under_3_5_Close:.2f}' if isinstance(Matchtime_Under_3_5_Close, (int, float)) else f'K MS 3.5 A | {Matchtime_Under_3_5_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_ms_3_5_a'])
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_3_5_ü'] = streamlit.checkbox(f'K MS 3.5 Ü | {Matchtime_Over_3_5_Close:.2f}' if isinstance(Matchtime_Over_3_5_Close, (int, float)) else f'K MS 3.5 Ü | {Matchtime_Over_3_5_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_ms_3_5_ü'])

                with streamlit.expander("İlk yarı filtreleri"):
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_1'] = streamlit.checkbox(f'K İY 1 | {Firsthalf_Home_Close:.2f}' if isinstance(Firsthalf_Home_Close, (int, float)) else f'K İY 1 | {Firsthalf_Home_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_iy_1'])
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_x'] = streamlit.checkbox(f'K İY X | {Firsthalf_Tie_Close:.2f}' if isinstance(Firsthalf_Tie_Close, (int, float)) else f'K İY X | {Firsthalf_Tie_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_iy_x'])
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_2'] = streamlit.checkbox(f'K İY 2 | {Firsthalf_Away_Close:.2f}' if isinstance(Firsthalf_Away_Close, (int, float)) else f'K İY 2 | {Firsthalf_Away_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_iy_2'])
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_0_5_a'] = streamlit.checkbox(f'K İY 0.5 A | {Firsthalf_Under_0_5_Close:.2f}' if isinstance(Firsthalf_Under_0_5_Close, (int, float)) else f'K İY 0.5 A | {Firsthalf_Under_0_5_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_iy_0_5_a'])
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_0_5_ü'] = streamlit.checkbox(f'K İY 0.5 Ü | {Firsthalf_Over_0_5_Close:.2f}' if isinstance(Firsthalf_Over_0_5_Close, (int, float)) else f'K İY 0.5 Ü | {Firsthalf_Over_0_5_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_iy_0_5_ü'])
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_1_5_a'] = streamlit.checkbox(f'K İY 1.5 A | {Firsthalf_Under_1_5_Close:.2f}' if isinstance(Firsthalf_Under_1_5_Close, (int, float)) else f'K İY 1.5 A | {Firsthalf_Under_1_5_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_iy_1_5_a'])
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_1_5_ü'] = streamlit.checkbox(f'K İY 1.5 Ü | {Firsthalf_Over_1_5_Close:.2f}' if isinstance(Firsthalf_Over_1_5_Close, (int, float)) else f'K İY 1.5 Ü | {Firsthalf_Over_1_5_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_iy_1_5_ü'])

                with streamlit.expander("Çifte şans filtreleri"):
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_1_x_cs'] = streamlit.checkbox(f'K MS 1-X | {Matchtime_Home_And_Tie_Close:.2f}' if isinstance(Matchtime_Home_And_Tie_Close, (int, float)) else f'K MS 1-X | {Matchtime_Home_And_Tie_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_ms_1_x_cs'])
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_1_2_cs'] = streamlit.checkbox(f'K MS 1-2 | {Matchtime_Home_And_Away_Close:.2f}' if isinstance(Matchtime_Home_And_Away_Close, (int, float)) else f'K MS 1-2 | {Matchtime_Home_And_Away_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_ms_1_2_cs'])
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_x_2_cs'] = streamlit.checkbox(f'K MS X-2 | {Matchtime_Tie_And_Away_Close:.2f}' if isinstance(Matchtime_Tie_And_Away_Close, (int, float)) else f'K MS X-2 | {Matchtime_Tie_And_Away_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_ms_x_2_cs'])
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_1_x_cs'] = streamlit.checkbox(f'K İY 1-X | {Firsthalf_Home_And_Tie_Close:.2f}' if isinstance(Firsthalf_Home_And_Tie_Close, (int, float)) else f'K İY 1-X | {Firsthalf_Home_And_Tie_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_iy_1_x_cs'])
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_1_2_cs'] = streamlit.checkbox(f'K İY 1-2 | {Firsthalf_Home_And_Away_Close:.2f}' if isinstance(Firsthalf_Home_And_Away_Close, (int, float)) else f'K İY 1-2 | {Firsthalf_Home_And_Away_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_iy_1_2_cs'])
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_x_2_cs'] = streamlit.checkbox(f'K İY X-2 | {Firsthalf_Tie_And_Away_Close:.2f}' if isinstance(Firsthalf_Tie_And_Away_Close, (int, float)) else f'K İY X-2 | {Firsthalf_Tie_And_Away_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_iy_x_2_cs'])

                with streamlit.expander("Karşılıklı filtreleri"):
                    streamlit.session_state.Filters_Dictionary['filter_k_home_win'] = streamlit.checkbox(f'K EV Kazanır | {Draw_No_Bet_Home_Close:.2f}' if isinstance(Draw_No_Bet_Home_Close, (int, float)) else f'K EV Kazanır | {Draw_No_Bet_Home_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_home_win'])
                    streamlit.session_state.Filters_Dictionary['filter_k_away_win'] = streamlit.checkbox(f'K DEP Kazanır | {Draw_No_Bet_Away_Close:.2f}' if isinstance(Draw_No_Bet_Away_Close, (int, float)) else f'K DEP Kazanır | {Draw_No_Bet_Away_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_away_win'])
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_kg_var'] = streamlit.checkbox(f'K İY KG VAR | {Firsthalf_Both_Teams_To_Score_Yes_Close:.2f}' if isinstance(Firsthalf_Both_Teams_To_Score_Yes_Close, (int, float)) else f'K İY KG VAR | {Firsthalf_Both_Teams_To_Score_Yes_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_iy_kg_var'])
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_kg_yok'] = streamlit.checkbox(f'K İY KG YOK | {Firsthalf_Both_Teams_To_Score_No_Close:.2f}' if isinstance(Firsthalf_Both_Teams_To_Score_No_Close, (int, float)) else f'K İY KG YOK | {Firsthalf_Both_Teams_To_Score_No_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_iy_kg_yok'])
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_kg_var'] = streamlit.checkbox(f'K MS KG VAR | {Matchtime_Both_Teams_To_Score_Yes_Close:.2f}' if isinstance(Matchtime_Both_Teams_To_Score_Yes_Close, (int, float)) else f'K MS KG VAR | {Matchtime_Both_Teams_To_Score_Yes_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_ms_kg_var'])
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_kg_yok'] = streamlit.checkbox(f'K MS KG YOK | {Matchtime_Both_Teams_To_Score_No_Close:.2f}' if isinstance(Matchtime_Both_Teams_To_Score_No_Close, (int, float)) else f'K MS KG YOK | {Matchtime_Both_Teams_To_Score_No_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_ms_kg_yok'])
                    streamlit.session_state.Filters_Dictionary['filter_k_tek'] = streamlit.checkbox(f'K TEK | {Odd_Close:.2f}' if isinstance(Odd_Close, (int, float)) else f'K TEK | {Odd_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_tek'])
                    streamlit.session_state.Filters_Dictionary['filter_k_cift'] = streamlit.checkbox(f'K ÇİFT | {Even_Close:.2f}' if isinstance(Even_Close, (int, float)) else f'K ÇİFT | {Even_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_cift'])

                with streamlit.expander("İlk yarı/Maç sonucu filtreleri"):
                    streamlit.session_state.Filters_Dictionary['filter_k_1den1'] = streamlit.checkbox(f'K 1/1 | {Home_Home_Close:.2f}' if isinstance(Home_Home_Close, (int, float)) else f'K 1/1 | {Home_Home_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_1den1'])
                    streamlit.session_state.Filters_Dictionary['filter_k_1denx'] = streamlit.checkbox(f'K 1/X | {Home_Tie_Close:.2f}' if isinstance(Home_Tie_Close, (int, float)) else f'K 1/X | {Home_Tie_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_1denx'])
                    streamlit.session_state.Filters_Dictionary['filter_k_1den2'] = streamlit.checkbox(f'K 1/2 | {Home_Away_Close:.2f}' if isinstance(Home_Away_Close, (int, float)) else f'K 1/2 | {Home_Away_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_1den2'])
                    streamlit.session_state.Filters_Dictionary['filter_k_xden1'] = streamlit.checkbox(f'K X/1 | {Tie_Home_Close:.2f}' if isinstance(Tie_Home_Close, (int, float)) else f'K X/1 | {Tie_Home_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_xden1'])
                    streamlit.session_state.Filters_Dictionary['filter_k_xdenx'] = streamlit.checkbox(f'K X/X | {Tie_Tie_Close:.2f}' if isinstance(Tie_Tie_Close, (int, float)) else f'K X/X | {Tie_Tie_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_xdenx'])
                    streamlit.session_state.Filters_Dictionary['filter_k_xden2'] = streamlit.checkbox(f'K X/2 | {Tie_Away_Close:.2f}' if isinstance(Tie_Away_Close, (int, float)) else f'K X/2 | {Tie_Away_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_xden2'])
                    streamlit.session_state.Filters_Dictionary['filter_k_2den1'] = streamlit.checkbox(f'K 2/1 | {Away_Home_Close:.2f}' if isinstance(Away_Home_Close, (int, float)) else f'K 2/1 | {Away_Home_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_2den1'])
                    streamlit.session_state.Filters_Dictionary['filter_k_2denx'] = streamlit.checkbox(f'K 2/X | {Away_Tie_Close:.2f}' if isinstance(Away_Tie_Close, (int, float)) else f'K 2/X | {Away_Tie_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_2denx'])
                    streamlit.session_state.Filters_Dictionary['filter_k_2den2'] = streamlit.checkbox(f'K 2/2 | {Away_Away_Close:.2f}' if isinstance(Away_Away_Close, (int, float)) else f'K 2/2 | {Away_Away_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_2den2'])

                with streamlit.expander("İlk yarı skor filtreleri"):
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_10'] = streamlit.checkbox(f'K İY SKOR 1-0 | {Firsthalf_1_0_Close:.2f}' if isinstance(Firsthalf_1_0_Close, (int, float)) else f'K İY SKOR 1-0 | {Firsthalf_1_0_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_10'])
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_20'] = streamlit.checkbox(f'K İY SKOR 2-0 | {Firsthalf_2_0_Close:.2f}' if isinstance(Firsthalf_2_0_Close, (int, float)) else f'K İY SKOR 2-0 | {Firsthalf_2_0_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_20'])
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_21'] = streamlit.checkbox(f'K İY SKOR 2-1 | {Firsthalf_2_1_Close:.2f}' if isinstance(Firsthalf_2_1_Close, (int, float)) else f'K İY SKOR 2-1 | {Firsthalf_2_1_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_21'])
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_30'] = streamlit.checkbox(f'K İY SKOR 3-0 | {Firsthalf_3_0_Close:.2f}' if isinstance(Firsthalf_3_0_Close, (int, float)) else f'K İY SKOR 3-0 | {Firsthalf_3_0_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_30'])
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_31'] = streamlit.checkbox(f'K İY SKOR 3-1 | {Firsthalf_3_1_Close:.2f}' if isinstance(Firsthalf_3_1_Close, (int, float)) else f'K İY SKOR 3-1 | {Firsthalf_3_1_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_31'])
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_32'] = streamlit.checkbox(f'K İY SKOR 3-2 | {Firsthalf_3_2_Close:.2f}' if isinstance(Firsthalf_3_2_Close, (int, float)) else f'K İY SKOR 3-2 | {Firsthalf_3_2_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_32'])
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_40'] = streamlit.checkbox(f'K İY SKOR 4-0 | {Firsthalf_4_0_Close:.2f}' if isinstance(Firsthalf_4_0_Close, (int, float)) else f'K İY SKOR 4-0 | {Firsthalf_4_0_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_40'])
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_41'] = streamlit.checkbox(f'K İY SKOR 4-1 | {Firsthalf_4_1_Close:.2f}' if isinstance(Firsthalf_4_1_Close, (int, float)) else f'K İY SKOR 4-1 | {Firsthalf_4_1_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_41'])
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_42'] = streamlit.checkbox(f'K İY SKOR 4-2 | {Firsthalf_4_2_Close:.2f}' if isinstance(Firsthalf_4_2_Close, (int, float)) else f'K İY SKOR 4-2 | {Firsthalf_4_2_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_42'])
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_43'] = streamlit.checkbox(f'K İY SKOR 4-3 | {Firsthalf_4_3_Close:.2f}' if isinstance(Firsthalf_4_3_Close, (int, float)) else f'K İY SKOR 4-3 | {Firsthalf_4_3_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_43'])
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_00'] = streamlit.checkbox(f'K İY SKOR 0-0 | {Firsthalf_0_0_Close:.2f}' if isinstance(Firsthalf_0_0_Close, (int, float)) else f'K İY SKOR 0-0 | {Firsthalf_0_0_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_00'])
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_11'] = streamlit.checkbox(f'K İY SKOR 1-1 | {Firsthalf_1_1_Close:.2f}' if isinstance(Firsthalf_1_1_Close, (int, float)) else f'K İY SKOR 1-1 | {Firsthalf_1_1_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_11'])
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_22'] = streamlit.checkbox(f'K İY SKOR 2-2 | {Firsthalf_2_2_Close:.2f}' if isinstance(Firsthalf_2_2_Close, (int, float)) else f'K İY SKOR 2-2 | {Firsthalf_2_2_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_22'])
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_33'] = streamlit.checkbox(f'K İY SKOR 3-3 | {Firsthalf_3_3_Close:.2f}' if isinstance(Firsthalf_3_3_Close, (int, float)) else f'K İY SKOR 3-3 | {Firsthalf_3_3_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_33'])
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_44'] = streamlit.checkbox(f'K İY SKOR 4-4 | {Firsthalf_4_4_Close:.2f}' if isinstance(Firsthalf_4_4_Close, (int, float)) else f'K İY SKOR 4-4 | {Firsthalf_4_4_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_44'])
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_01'] = streamlit.checkbox(f'K İY SKOR 0-1 | {Firsthalf_0_1_Close:.2f}' if isinstance(Firsthalf_0_1_Close, (int, float)) else f'K İY SKOR 0-1 | {Firsthalf_0_1_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_01'])
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_02'] = streamlit.checkbox(f'K İY SKOR 0-2 | {Firsthalf_0_2_Close:.2f}' if isinstance(Firsthalf_0_2_Close, (int, float)) else f'K İY SKOR 0-2 | {Firsthalf_0_2_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_02'])
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_12'] = streamlit.checkbox(f'K İY SKOR 1-2 | {Firsthalf_1_2_Close:.2f}' if isinstance(Firsthalf_1_2_Close, (int, float)) else f'K İY SKOR 1-2 | {Firsthalf_1_2_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_12'])
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_03'] = streamlit.checkbox(f'K İY SKOR 0-3 | {Firsthalf_0_3_Close:.2f}' if isinstance(Firsthalf_0_3_Close, (int, float)) else f'K İY SKOR 0-3 | {Firsthalf_0_3_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_03'])
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_13'] = streamlit.checkbox(f'K İY SKOR 1-3 | {Firsthalf_1_3_Close:.2f}' if isinstance(Firsthalf_1_3_Close, (int, float)) else f'K İY SKOR 1-3 | {Firsthalf_1_3_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_13'])
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_23'] = streamlit.checkbox(f'K İY SKOR 2-3 | {Firsthalf_2_3_Close:.2f}' if isinstance(Firsthalf_2_3_Close, (int, float)) else f'K İY SKOR 2-3 | {Firsthalf_2_3_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_23'])
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_04'] = streamlit.checkbox(f'K İY SKOR 0-4 | {Firsthalf_0_4_Close:.2f}' if isinstance(Firsthalf_0_4_Close, (int, float)) else f'K İY SKOR 0-4 | {Firsthalf_0_4_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_04'])
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_14'] = streamlit.checkbox(f'K İY SKOR 1-4 | {Firsthalf_1_4_Close:.2f}' if isinstance(Firsthalf_1_4_Close, (int, float)) else f'K İY SKOR 1-4 | {Firsthalf_1_4_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_14'])
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_24'] = streamlit.checkbox(f'K İY SKOR 2-4 | {Firsthalf_2_4_Close:.2f}' if isinstance(Firsthalf_2_4_Close, (int, float)) else f'K İY SKOR 2-4 | {Firsthalf_2_4_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_24'])
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_34'] = streamlit.checkbox(f'K İY SKOR 3-4 | {Firsthalf_3_4_Close:.2f}' if isinstance(Firsthalf_3_4_Close, (int, float)) else f'K İY SKOR 3-4 | {Firsthalf_3_4_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_34'])

                with streamlit.expander("Maç sonu skor filtreleri"):
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_10'] = streamlit.checkbox(f'K MS SKOR 1-0 | {Matchtime_1_0_Close:.2f}' if isinstance(Matchtime_1_0_Close, (int, float)) else f'K MS SKOR 1-0 | {Matchtime_1_0_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_10'])
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_20'] = streamlit.checkbox(f'K MS SKOR 2-0 | {Matchtime_2_0_Close:.2f}' if isinstance(Matchtime_2_0_Close, (int, float)) else f'K MS SKOR 2-0 | {Matchtime_2_0_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_20'])
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_21'] = streamlit.checkbox(f'K MS SKOR 2-1 | {Matchtime_2_1_Close:.2f}' if isinstance(Matchtime_2_1_Close, (int, float)) else f'K MS SKOR 2-1 | {Matchtime_2_1_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_21'])
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_30'] = streamlit.checkbox(f'K MS SKOR 3-0 | {Matchtime_3_0_Close:.2f}' if isinstance(Matchtime_3_0_Close, (int, float)) else f'K MS SKOR 3-0 | {Matchtime_3_0_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_30'])
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_31'] = streamlit.checkbox(f'K MS SKOR 3-1 | {Matchtime_3_1_Close:.2f}' if isinstance(Matchtime_3_1_Close, (int, float)) else f'K MS SKOR 3-1 | {Matchtime_3_1_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_31'])
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_32'] = streamlit.checkbox(f'K MS SKOR 3-2 | {Matchtime_3_2_Close:.2f}' if isinstance(Matchtime_3_2_Close, (int, float)) else f'K MS SKOR 3-2 | {Matchtime_3_2_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_32'])
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_40'] = streamlit.checkbox(f'K MS SKOR 4-0 | {Matchtime_4_0_Close:.2f}' if isinstance(Matchtime_4_0_Close, (int, float)) else f'K MS SKOR 4-0 | {Matchtime_4_0_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_40'])
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_41'] = streamlit.checkbox(f'K MS SKOR 4-1 | {Matchtime_4_1_Close:.2f}' if isinstance(Matchtime_4_1_Close, (int, float)) else f'K MS SKOR 4-1 | {Matchtime_4_1_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_41'])
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_42'] = streamlit.checkbox(f'K MS SKOR 4-2 | {Matchtime_4_2_Close:.2f}' if isinstance(Matchtime_4_2_Close, (int, float)) else f'K MS SKOR 4-2 | {Matchtime_4_2_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_42'])
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_43'] = streamlit.checkbox(f'K MS SKOR 4-3 | {Matchtime_4_3_Close:.2f}' if isinstance(Matchtime_4_3_Close, (int, float)) else f'K MS SKOR 4-3 | {Matchtime_4_3_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_43'])
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_00'] = streamlit.checkbox(f'K MS SKOR 0-0 | {Matchtime_0_0_Close:.2f}' if isinstance(Matchtime_0_0_Close, (int, float)) else f'K MS SKOR 0-0 | {Matchtime_0_0_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_00'])
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_11'] = streamlit.checkbox(f'K MS SKOR 1-1 | {Matchtime_1_1_Close:.2f}' if isinstance(Matchtime_1_1_Close, (int, float)) else f'K MS SKOR 1-1 | {Matchtime_1_1_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_11'])
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_22'] = streamlit.checkbox(f'K MS SKOR 2-2 | {Matchtime_2_2_Close:.2f}' if isinstance(Matchtime_2_2_Close, (int, float)) else f'K MS SKOR 2-2 | {Matchtime_2_2_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_22'])
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_33'] = streamlit.checkbox(f'K MS SKOR 3-3 | {Matchtime_3_3_Close:.2f}' if isinstance(Matchtime_3_3_Close, (int, float)) else f'K MS SKOR 3-3 | {Matchtime_3_3_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_33'])
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_44'] = streamlit.checkbox(f'K MS SKOR 4-4 | {Matchtime_4_4_Close:.2f}' if isinstance(Matchtime_4_4_Close, (int, float)) else f'K MS SKOR 4-4 | {Matchtime_4_4_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_44'])
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_01'] = streamlit.checkbox(f'K MS SKOR 0-1 | {Matchtime_0_1_Close:.2f}' if isinstance(Matchtime_0_1_Close, (int, float)) else f'K MS SKOR 0-1 | {Matchtime_0_1_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_01'])
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_02'] = streamlit.checkbox(f'K MS SKOR 0-2 | {Matchtime_0_2_Close:.2f}' if isinstance(Matchtime_0_2_Close, (int, float)) else f'K MS SKOR 0-2 | {Matchtime_0_2_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_02'])
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_12'] = streamlit.checkbox(f'K MS SKOR 1-2 | {Matchtime_1_2_Close:.2f}' if isinstance(Matchtime_1_2_Close, (int, float)) else f'K MS SKOR 1-2 | {Matchtime_1_2_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_12'])
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_03'] = streamlit.checkbox(f'K MS SKOR 0-3 | {Matchtime_0_3_Close:.2f}' if isinstance(Matchtime_0_3_Close, (int, float)) else f'K MS SKOR 0-3 | {Matchtime_0_3_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_03'])
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_13'] = streamlit.checkbox(f'K MS SKOR 1-3 | {Matchtime_1_3_Close:.2f}' if isinstance(Matchtime_1_3_Close, (int, float)) else f'K MS SKOR 1-3 | {Matchtime_1_3_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_13'])
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_23'] = streamlit.checkbox(f'K MS SKOR 2-3 | {Matchtime_2_3_Close:.2f}' if isinstance(Matchtime_2_3_Close, (int, float)) else f'K MS SKOR 2-3 | {Matchtime_2_3_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_23'])
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_04'] = streamlit.checkbox(f'K MS SKOR 0-4 | {Matchtime_0_4_Close:.2f}' if isinstance(Matchtime_0_4_Close, (int, float)) else f'K MS SKOR 0-4 | {Matchtime_0_4_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_04'])
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_14'] = streamlit.checkbox(f'K MS SKOR 1-4 | {Matchtime_1_4_Close:.2f}' if isinstance(Matchtime_1_4_Close, (int, float)) else f'K MS SKOR 1-4 | {Matchtime_1_4_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_14'])
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_24'] = streamlit.checkbox(f'K MS SKOR 2-4 | {Matchtime_2_4_Close:.2f}' if isinstance(Matchtime_2_4_Close, (int, float)) else f'K MS SKOR 2-4 | {Matchtime_2_4_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_24'])
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_34'] = streamlit.checkbox(f'K MS SKOR 3-4 | {Matchtime_3_4_Close:.2f}' if isinstance(Matchtime_3_4_Close, (int, float)) else f'K MS SKOR 3-4 | {Matchtime_3_4_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_34'])

                with streamlit.expander("İlk yarı asya handikap filtreleri"):
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_neg_2'] = streamlit.checkbox(f'K İY ASYA EV  -2.0 | {Firsthalf_Asian_Neg2_Home_Close:.2f}' if isinstance(Firsthalf_Asian_Neg2_Home_Close, (int, float)) else f'K İY ASYA EV  -2.0 | {Firsthalf_Asian_Neg2_Home_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_neg_2'])
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_neg_2'] = streamlit.checkbox(f'K İY ASYA DEP -2.0 | {Firsthalf_Asian_Neg2_Away_Close:.2f}' if isinstance(Firsthalf_Asian_Neg2_Away_Close, (int, float)) else f'K İY ASYA DEP -2.0 | {Firsthalf_Asian_Neg2_Away_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_neg_2'])
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_neg_2_neg_1_5'] = streamlit.checkbox(f'K İY ASYA EV  -2, -1.5 | {Firsthalf_Asian_Neg2_Neg1_5_Home_Close:.2f}' if isinstance(Firsthalf_Asian_Neg2_Neg1_5_Home_Close, (int, float)) else f'K İY ASYA EV  -2, -1.5 | {Firsthalf_Asian_Neg2_Neg1_5_Home_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_neg_2_neg_1_5'])
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_neg_2_neg_1_5'] = streamlit.checkbox(f'K İY ASYA DEP -2, -1.5 | {Firsthalf_Asian_Neg2_Neg1_5_Away_Close:.2f}' if isinstance(Firsthalf_Asian_Neg2_Neg1_5_Away_Close, (int, float)) else f'K İY ASYA DEP -2, -1.5 | {Firsthalf_Asian_Neg2_Neg1_5_Away_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_neg_2_neg_1_5'])
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_neg_1_5'] = streamlit.checkbox(f'K İY ASYA EV  -1.5 | {Firsthalf_Asian_Neg1_5_Home_Close:.2f}' if isinstance(Firsthalf_Asian_Neg1_5_Home_Close, (int, float)) else f'K İY ASYA EV  -1.5 | {Firsthalf_Asian_Neg1_5_Home_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_neg_1_5'])
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_neg_1_5'] = streamlit.checkbox(f'K İY ASYA DEP -1.5 | {Firsthalf_Asian_Neg1_5_Away_Close:.2f}' if isinstance(Firsthalf_Asian_Neg1_5_Away_Close, (int, float)) else f'K İY ASYA DEP -1.5 | {Firsthalf_Asian_Neg1_5_Away_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_neg_1_5'])
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_neg_1_5_neg_1'] = streamlit.checkbox(f'K İY ASYA EV  -1.5, -1.0 | {Firsthalf_Asian_Neg1_5_Neg1_Home_Close:.2f}' if isinstance(Firsthalf_Asian_Neg1_5_Neg1_Home_Close, (int, float)) else f'K İY ASYA EV  -1.5, -1.0 | {Firsthalf_Asian_Neg1_5_Neg1_Home_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_neg_1_5_neg_1'])
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_neg_1_5_neg_1'] = streamlit.checkbox(f'K İY ASYA DEP -1.5, -1.0 | {Firsthalf_Asian_Neg1_5_Neg1_Away_Close:.2f}' if isinstance(Firsthalf_Asian_Neg1_5_Neg1_Away_Close, (int, float)) else f'K İY ASYA DEP -1.5, -1.0 | {Firsthalf_Asian_Neg1_5_Neg1_Away_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_neg_1_5_neg_1'])
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_neg_1'] = streamlit.checkbox(f'K İY ASYA EV  -1.0 | {Firsthalf_Asian_Neg1_Home_Close:.2f}' if isinstance(Firsthalf_Asian_Neg1_Home_Close, (int, float)) else f'K İY ASYA EV  -1.0 | {Firsthalf_Asian_Neg1_Home_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_neg_1'])
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_neg_1'] = streamlit.checkbox(f'K İY ASYA DEP -1.0 | {Firsthalf_Asian_Neg1_Away_Close:.2f}' if isinstance(Firsthalf_Asian_Neg1_Away_Close, (int, float)) else f'K İY ASYA DEP -1.0 | {Firsthalf_Asian_Neg1_Away_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_neg_1'])
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_neg_1_neg_0_5'] = streamlit.checkbox(f'K İY ASYA EV  -1.0, -0.5 | {Firsthalf_Asian_Neg1_Neg0_5_Home_Close:.2f}' if isinstance(Firsthalf_Asian_Neg1_Neg0_5_Home_Close, (int, float)) else f'K İY ASYA EV  -1.0, -0.5 | {Firsthalf_Asian_Neg1_Neg0_5_Home_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_neg_1_neg_0_5'])
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_neg_1_neg_0_5'] = streamlit.checkbox(f'K İY ASYA DEP -1.0, -0.5 | {Firsthalf_Asian_Neg1_Neg0_5_Away_Close:.2f}' if isinstance(Firsthalf_Asian_Neg1_Neg0_5_Away_Close, (int, float)) else f'K İY ASYA DEP -1.0, -0.5 | {Firsthalf_Asian_Neg1_Neg0_5_Away_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_neg_1_neg_0_5'])
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_neg_0_5'] = streamlit.checkbox(f'K İY ASYA EV  -0.5 | {Firsthalf_Asian_Neg0_5_Home_Close:.2f}' if isinstance(Firsthalf_Asian_Neg0_5_Home_Close, (int, float)) else f'K İY ASYA EV  -0.5 | {Firsthalf_Asian_Neg0_5_Home_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_neg_0_5'])
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_neg_0_5'] = streamlit.checkbox(f'K İY ASYA DEP -0.5 | {Firsthalf_Asian_Neg0_5_Away_Close:.2f}' if isinstance(Firsthalf_Asian_Neg0_5_Away_Close, (int, float)) else f'K İY ASYA DEP -0.5 | {Firsthalf_Asian_Neg0_5_Away_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_neg_0_5'])
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_neg_0_5_pos_0'] = streamlit.checkbox(f'K İY ASYA EV  -0.5, 0.0 | {Firsthalf_Asian_Neg0_5_Neg0_0_Home_Close:.2f}' if isinstance(Firsthalf_Asian_Neg0_5_Neg0_0_Home_Close, (int, float)) else f'K İY ASYA EV  -0.5, 0.0 | {Firsthalf_Asian_Neg0_5_Neg0_0_Home_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_neg_0_5_pos_0'])
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_neg_0_5_pos_0'] = streamlit.checkbox(f'K İY ASYA DEP -0.5, 0.0 | {Firsthalf_Asian_Neg0_5_Neg0_0_Away_Close:.2f}' if isinstance(Firsthalf_Asian_Neg0_5_Neg0_0_Away_Close, (int, float)) else f'K İY ASYA DEP -0.5, 0.0 | {Firsthalf_Asian_Neg0_5_Neg0_0_Away_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_neg_0_5_pos_0'])
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_pos_0_pos_0_5'] = streamlit.checkbox(f'K İY ASYA EV  0.0, 0.5 | {Firsthalf_Asian_0_0_Pos0_5_Home_Close:.2f}' if isinstance(Firsthalf_Asian_0_0_Pos0_5_Home_Close, (int, float)) else f'K İY ASYA EV  0.0, 0.5 | {Firsthalf_Asian_0_0_Pos0_5_Home_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_pos_0_pos_0_5'])
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_pos_0_pos_0_5'] = streamlit.checkbox(f'K İY ASYA DEP 0.0, 0.5 | {Firsthalf_Asian_0_0_Pos0_5_Away_Close:.2f}' if isinstance(Firsthalf_Asian_0_0_Pos0_5_Away_Close, (int, float)) else f'K İY ASYA DEP 0.0, 0.5 | {Firsthalf_Asian_0_0_Pos0_5_Away_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_pos_0_pos_0_5'])
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_pos_0_5'] = streamlit.checkbox(f'K İY ASYA EV  0.5 | {Firsthalf_Asian_Pos0_5_Home_Close:.2f}' if isinstance(Firsthalf_Asian_Pos0_5_Home_Close, (int, float)) else f'K İY ASYA EV  0.5 | {Firsthalf_Asian_Pos0_5_Home_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_pos_0_5'])
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_pos_0_5'] = streamlit.checkbox(f'K İY ASYA DEP 0.5 | {Firsthalf_Asian_Pos0_5_Away_Close:.2f}' if isinstance(Firsthalf_Asian_Pos0_5_Away_Close, (int, float)) else f'K İY ASYA DEP 0.5 | {Firsthalf_Asian_Pos0_5_Away_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_pos_0_5'])
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_pos_0_5_pos_1'] = streamlit.checkbox(f'K İY ASYA EV  0.5, 1.0 | {Firsthalf_Asian_Pos0_5_Pos1_0_Home_Close:.2f}' if isinstance(Firsthalf_Asian_Pos0_5_Pos1_0_Home_Close, (int, float)) else f'K İY ASYA EV  0.5, 1.0 | {Firsthalf_Asian_Pos0_5_Pos1_0_Home_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_pos_0_5_pos_1'])
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_pos_0_5_pos_1'] = streamlit.checkbox(f'K İY ASYA DEP 0.5, 1.0 | {Firsthalf_Asian_Pos0_5_Pos1_0_Away_Close:.2f}' if isinstance(Firsthalf_Asian_Pos0_5_Pos1_0_Away_Close, (int, float)) else f'K İY ASYA DEP 0.5, 1.0 | {Firsthalf_Asian_Pos0_5_Pos1_0_Away_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_pos_0_5_pos_1'])
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_pos_1'] = streamlit.checkbox(f'K İY ASYA EV  1.0 | {Firsthalf_Asian_Pos1_Home_Close:.2f}' if isinstance(Firsthalf_Asian_Pos1_Home_Close, (int, float)) else f'K İY ASYA EV  1.0 | {Firsthalf_Asian_Pos1_Home_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_pos_1'])
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_pos_1'] = streamlit.checkbox(f'K İY ASYA DEP 1.0 | {Firsthalf_Asian_Pos1_Away_Close:.2f}' if isinstance(Firsthalf_Asian_Pos1_Away_Close, (int, float)) else f'K İY ASYA DEP 1.0 | {Firsthalf_Asian_Pos1_Away_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_pos_1'])
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_pos_1_pos_1_5'] = streamlit.checkbox(f'K İY ASYA EV  1.0, 1.5 | {Firsthalf_Asian_Pos1_Pos1_5_Home_Close:.2f}' if isinstance(Firsthalf_Asian_Pos1_Pos1_5_Home_Close, (int, float)) else f'K İY ASYA EV  1.0, 1.5 | {Firsthalf_Asian_Pos1_Pos1_5_Home_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_pos_1_pos_1_5'])
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_pos_1_pos_1_5'] = streamlit.checkbox(f'K İY ASYA DEP 1.0, 1.5 | {Firsthalf_Asian_Pos1_Pos1_5_Away_Close:.2f}' if isinstance(Firsthalf_Asian_Pos1_Pos1_5_Away_Close, (int, float)) else f'K İY ASYA DEP 1.0, 1.5 | {Firsthalf_Asian_Pos1_Pos1_5_Away_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_pos_1_pos_1_5'])
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_pos_1_5'] = streamlit.checkbox(f'K İY ASYA EV  1.5 | {Firsthalf_Asian_Pos1_5_Home_Close:.2f}' if isinstance(Firsthalf_Asian_Pos1_5_Home_Close, (int, float)) else f'K İY ASYA EV  1.5 | {Firsthalf_Asian_Pos1_5_Home_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_pos_1_5'])
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_pos_1_5'] = streamlit.checkbox(f'K İY ASYA DEP 1.5 | {Firsthalf_Asian_Pos1_5_Away_Close:.2f}' if isinstance(Firsthalf_Asian_Pos1_5_Away_Close, (int, float)) else f'K İY ASYA DEP 1.5 | {Firsthalf_Asian_Pos1_5_Away_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_pos_1_5'])
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_pos_1_5_pos_2'] = streamlit.checkbox(f'K İY ASYA EV  1.5, 2.0 | {Firsthalf_Asian_Pos1_5_Pos2_Home_Close:.2f}' if isinstance(Firsthalf_Asian_Pos1_5_Pos2_Home_Close, (int, float)) else f'K İY ASYA EV  1.5, 2.0 | {Firsthalf_Asian_Pos1_5_Pos2_Home_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_pos_1_5_pos_2'])
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_pos_1_5_pos_2'] = streamlit.checkbox(f'K İY ASYA DEP 1.5, 2.0 | {Firsthalf_Asian_Pos1_5_Pos2_Away_Close:.2f}' if isinstance(Firsthalf_Asian_Pos1_5_Pos2_Away_Close, (int, float)) else f'K İY ASYA DEP 1.5, 2.0 | {Firsthalf_Asian_Pos1_5_Pos2_Away_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_pos_1_5_pos_2'])
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_pos_2'] = streamlit.checkbox(f'K İY ASYA EV  2.0 | {Firsthalf_Asian_Pos2_Home_Close:.2f}' if isinstance(Firsthalf_Asian_Pos2_Home_Close, (int, float)) else f'K İY ASYA EV  2.0 | {Firsthalf_Asian_Pos2_Home_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_pos_2'])
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_pos_2'] = streamlit.checkbox(f'K İY ASYA DEP 2.0 | {Firsthalf_Asian_Pos2_Away_Close:.2f}' if isinstance(Firsthalf_Asian_Pos2_Away_Close, (int, float)) else f'K İY ASYA DEP 2.0 | {Firsthalf_Asian_Pos2_Away_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_pos_2'])

                with streamlit.expander("Maç sonu asya handikap filtreleri"):
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_neg_2'] = streamlit.checkbox(f'K MS ASYA EV  -2.0 | {Matchtime_Asian_Neg2_Home_Close:.2f}' if isinstance(Matchtime_Asian_Neg2_Home_Close, (int, float)) else f'K MS ASYA EV  -2.0 | {Matchtime_Asian_Neg2_Home_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_neg_2'])
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_neg_2'] = streamlit.checkbox(f'K MS ASYA DEP -2.0 | {Matchtime_Asian_Neg2_Away_Close:.2f}' if isinstance(Matchtime_Asian_Neg2_Away_Close, (int, float)) else f'K MS ASYA DEP -2.0 | {Matchtime_Asian_Neg2_Away_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_neg_2'])
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_neg_2_neg_1_5'] = streamlit.checkbox(f'K MS ASYA EV  -2, -1.5 | {Matchtime_Asian_Neg2_Neg1_5_Home_Close:.2f}' if isinstance(Matchtime_Asian_Neg2_Neg1_5_Home_Close, (int, float)) else f'K MS ASYA EV  -2, -1.5 | {Matchtime_Asian_Neg2_Neg1_5_Home_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_neg_2_neg_1_5'])
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_neg_2_neg_1_5'] = streamlit.checkbox(f'K MS ASYA DEP -2, -1.5 | {Matchtime_Asian_Neg2_Neg1_5_Away_Close:.2f}' if isinstance(Matchtime_Asian_Neg2_Neg1_5_Away_Close, (int, float)) else f'K MS ASYA DEP -2, -1.5 | {Matchtime_Asian_Neg2_Neg1_5_Away_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_neg_2_neg_1_5'])
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_neg_1_5'] = streamlit.checkbox(f'K MS ASYA EV  -1.5 | {Matchtime_Asian_Neg1_5_Home_Close:.2f}' if isinstance(Matchtime_Asian_Neg1_5_Home_Close, (int, float)) else f'K MS ASYA EV  -1.5 | {Matchtime_Asian_Neg1_5_Home_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_neg_1_5'])
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_neg_1_5'] = streamlit.checkbox(f'K MS ASYA DEP -1.5 | {Matchtime_Asian_Neg1_5_Away_Close:.2f}' if isinstance(Matchtime_Asian_Neg1_5_Away_Close, (int, float)) else f'K MS ASYA DEP -1.5 | {Matchtime_Asian_Neg1_5_Away_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_neg_1_5'])
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_neg_1_5_neg_1'] = streamlit.checkbox(f'K MS ASYA EV  -1.5, -1.0 | {Matchtime_Asian_Neg1_5_Neg1_Home_Close:.2f}' if isinstance(Matchtime_Asian_Neg1_5_Neg1_Home_Close, (int, float)) else f'K MS ASYA EV  -1.5, -1.0 | {Matchtime_Asian_Neg1_5_Neg1_Home_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_neg_1_5_neg_1'])
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_neg_1_5_neg_1'] = streamlit.checkbox(f'K MS ASYA DEP -1.5, -1.0 | {Matchtime_Asian_Neg1_5_Neg1_Away_Close:.2f}' if isinstance(Matchtime_Asian_Neg1_5_Neg1_Away_Close, (int, float)) else f'K MS ASYA DEP -1.5, -1.0 | {Matchtime_Asian_Neg1_5_Neg1_Away_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_neg_1_5_neg_1'])
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_neg_1'] = streamlit.checkbox(f'K MS ASYA EV  -1.0 | {Matchtime_Asian_Neg1_Home_Close:.2f}' if isinstance(Matchtime_Asian_Neg1_Home_Close, (int, float)) else f'K MS ASYA EV  -1.0 | {Matchtime_Asian_Neg1_Home_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_neg_1'])
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_neg_1'] = streamlit.checkbox(f'K MS ASYA DEP -1.0 | {Matchtime_Asian_Neg1_Away_Close:.2f}' if isinstance(Matchtime_Asian_Neg1_Away_Close, (int, float)) else f'K MS ASYA DEP -1.0 | {Matchtime_Asian_Neg1_Away_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_neg_1'])
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_neg_1_neg_0_5'] = streamlit.checkbox(f'K MS ASYA EV  -1.0, -0.5 | {Matchtime_Asian_Neg1_Neg0_5_Home_Close:.2f}' if isinstance(Matchtime_Asian_Neg1_Neg0_5_Home_Close, (int, float)) else f'K MS ASYA EV  -1.0, -0.5 | {Matchtime_Asian_Neg1_Neg0_5_Home_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_neg_1_neg_0_5'])
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_neg_1_neg_0_5'] = streamlit.checkbox(f'K MS ASYA DEP -1.0, -0.5 | {Matchtime_Asian_Neg1_Neg0_5_Away_Close:.2f}' if isinstance(Matchtime_Asian_Neg1_Neg0_5_Away_Close, (int, float)) else f'K MS ASYA DEP -1.0, -0.5 | {Matchtime_Asian_Neg1_Neg0_5_Away_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_neg_1_neg_0_5'])
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_neg_0_5'] = streamlit.checkbox(f'K MS ASYA EV  -0.5 | {Matchtime_Asian_Neg0_5_Home_Close:.2f}' if isinstance(Matchtime_Asian_Neg0_5_Home_Close, (int, float)) else f'K MS ASYA EV  -0.5 | {Matchtime_Asian_Neg0_5_Home_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_neg_0_5'])
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_neg_0_5'] = streamlit.checkbox(f'K MS ASYA DEP -0.5 | {Matchtime_Asian_Neg0_5_Away_Close:.2f}' if isinstance(Matchtime_Asian_Neg0_5_Away_Close, (int, float)) else f'K MS ASYA DEP -0.5 | {Matchtime_Asian_Neg0_5_Away_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_neg_0_5'])
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_neg_0_5_pos_0'] = streamlit.checkbox(f'K MS ASYA EV  -0.5, 0.0 | {Matchtime_Asian_Neg0_5_Neg0_0_Home_Close:.2f}' if isinstance(Matchtime_Asian_Neg0_5_Neg0_0_Home_Close, (int, float)) else f'K MS ASYA EV  -0.5, 0.0 | {Matchtime_Asian_Neg0_5_Neg0_0_Home_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_neg_0_5_pos_0'])
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_neg_0_5_pos_0'] = streamlit.checkbox(f'K MS ASYA DEP -0.5, 0.0 | {Matchtime_Asian_Neg0_5_Neg0_0_Away_Close:.2f}' if isinstance(Matchtime_Asian_Neg0_5_Neg0_0_Away_Close, (int, float)) else f'K MS ASYA DEP -0.5, 0.0 | {Matchtime_Asian_Neg0_5_Neg0_0_Away_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_neg_0_5_pos_0'])
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_pos_0_pos_0_5'] = streamlit.checkbox(f'K MS ASYA EV  0.0, 0.5 | {Matchtime_Asian_0_0_Pos0_5_Home_Close:.2f}' if isinstance(Matchtime_Asian_0_0_Pos0_5_Home_Close, (int, float)) else f'K MS ASYA EV  0.0, 0.5 | {Matchtime_Asian_0_0_Pos0_5_Home_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_pos_0_pos_0_5'])
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_pos_0_pos_0_5'] = streamlit.checkbox(f'K MS ASYA DEP 0.0, 0.5 | {Matchtime_Asian_0_0_Pos0_5_Away_Close:.2f}' if isinstance(Matchtime_Asian_0_0_Pos0_5_Away_Close, (int, float)) else f'K MS ASYA DEP 0.0, 0.5 | {Matchtime_Asian_0_0_Pos0_5_Away_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_pos_0_pos_0_5'])
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_pos_0_5'] = streamlit.checkbox(f'K MS ASYA EV  0.5 | {Matchtime_Asian_Pos0_5_Home_Close:.2f}' if isinstance(Matchtime_Asian_Pos0_5_Home_Close, (int, float)) else f'K MS ASYA EV  0.5 | {Matchtime_Asian_Pos0_5_Home_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_pos_0_5'])
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_pos_0_5'] = streamlit.checkbox(f'K MS ASYA DEP 0.5 | {Matchtime_Asian_Pos0_5_Away_Close:.2f}' if isinstance(Matchtime_Asian_Pos0_5_Away_Close, (int, float)) else f'K MS ASYA DEP 0.5 | {Matchtime_Asian_Pos0_5_Away_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_pos_0_5'])
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_pos_0_5_pos_1'] = streamlit.checkbox(f'K MS ASYA EV  0.5, 1.0 | {Matchtime_Asian_Pos0_5_Pos1_0_Home_Close:.2f}' if isinstance(Matchtime_Asian_Pos0_5_Pos1_0_Home_Close, (int, float)) else f'K MS ASYA EV  0.5, 1.0 | {Matchtime_Asian_Pos0_5_Pos1_0_Home_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_pos_0_5_pos_1'])
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_pos_0_5_pos_1'] = streamlit.checkbox(f'K MS ASYA DEP 0.5, 1.0 | {Matchtime_Asian_Pos0_5_Pos1_0_Away_Close:.2f}' if isinstance(Matchtime_Asian_Pos0_5_Pos1_0_Away_Close, (int, float)) else f'K MS ASYA DEP 0.5, 1.0 | {Matchtime_Asian_Pos0_5_Pos1_0_Away_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_pos_0_5_pos_1'])
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_pos_1'] = streamlit.checkbox(f'K MS ASYA EV  1.0 | {Matchtime_Asian_Pos1_Home_Close:.2f}' if isinstance(Matchtime_Asian_Pos1_Home_Close, (int, float)) else f'K MS ASYA EV  1.0 | {Matchtime_Asian_Pos1_Home_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_pos_1'])
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_pos_1'] = streamlit.checkbox(f'K MS ASYA DEP 1.0 | {Matchtime_Asian_Pos1_Away_Close:.2f}' if isinstance(Matchtime_Asian_Pos1_Away_Close, (int, float)) else f'K MS ASYA DEP 1.0 | {Matchtime_Asian_Pos1_Away_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_pos_1'])
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_pos_1_pos_1_5'] = streamlit.checkbox(f'K MS ASYA EV  1.0, 1.5 | {Matchtime_Asian_Pos1_Pos1_5_Home_Close:.2f}' if isinstance(Matchtime_Asian_Pos1_Pos1_5_Home_Close, (int, float)) else f'K MS ASYA EV  1.0, 1.5 | {Matchtime_Asian_Pos1_Pos1_5_Home_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_pos_1_pos_1_5'])
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_pos_1_pos_1_5'] = streamlit.checkbox(f'K MS ASYA DEP 1.0, 1.5 | {Matchtime_Asian_Pos1_Pos1_5_Away_Close:.2f}' if isinstance(Matchtime_Asian_Pos1_Pos1_5_Away_Close, (int, float)) else f'K MS ASYA DEP 1.0, 1.5 | {Matchtime_Asian_Pos1_Pos1_5_Away_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_pos_1_pos_1_5'])
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_pos_1_5'] = streamlit.checkbox(f'K MS ASYA EV  1.5 | {Matchtime_Asian_Pos1_5_Home_Close:.2f}' if isinstance(Matchtime_Asian_Pos1_5_Home_Close, (int, float)) else f'K MS ASYA EV  1.5 | {Matchtime_Asian_Pos1_5_Home_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_pos_1_5'])
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_pos_1_5'] = streamlit.checkbox(f'K MS ASYA DEP 1.5 | {Matchtime_Asian_Pos1_5_Away_Close:.2f}' if isinstance(Matchtime_Asian_Pos1_5_Away_Close, (int, float)) else f'K MS ASYA DEP 1.5 | {Matchtime_Asian_Pos1_5_Away_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_pos_1_5'])
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_pos_1_5_pos_2'] = streamlit.checkbox(f'K MS ASYA EV  1.5, 2.0 | {Matchtime_Asian_Pos1_5_Pos2_Home_Close:.2f}' if isinstance(Matchtime_Asian_Pos1_5_Pos2_Home_Close, (int, float)) else f'K MS ASYA EV  1.5, 2.0 | {Matchtime_Asian_Pos1_5_Pos2_Home_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_pos_1_5_pos_2'])
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_pos_1_5_pos_2'] = streamlit.checkbox(f'K MS ASYA DEP 1.5, 2.0 | {Matchtime_Asian_Pos1_5_Pos2_Away_Close:.2f}' if isinstance(Matchtime_Asian_Pos1_5_Pos2_Away_Close, (int, float)) else f'K MS ASYA DEP 1.5, 2.0 | {Matchtime_Asian_Pos1_5_Pos2_Away_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_pos_1_5_pos_2'])
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_pos_2'] = streamlit.checkbox(f'K MS ASYA EV  2.0 | {Matchtime_Asian_Pos2_Home_Close:.2f}' if isinstance(Matchtime_Asian_Pos2_Home_Close, (int, float)) else f'K MS ASYA EV  2.0 | {Matchtime_Asian_Pos2_Home_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_pos_2'])
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_pos_2'] = streamlit.checkbox(f'K MS ASYA DEP 2.0 | {Matchtime_Asian_Pos2_Away_Close:.2f}' if isinstance(Matchtime_Asian_Pos2_Away_Close, (int, float)) else f'K MS ASYA DEP 2.0 | {Matchtime_Asian_Pos2_Away_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_pos_2'])

                with streamlit.expander("İlk yarı avrupa handikap filtreleri"):
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ev_neg_1'] = streamlit.checkbox(f'K İY AVRUPA EV  -1.0 | {Firsthalf_European_Neg1_Home_Close:.2f}' if isinstance(Firsthalf_European_Neg1_Home_Close, (int, float)) else f'K İY AVRUPA EV  -1.0 | {Firsthalf_European_Neg1_Home_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ev_neg_1'])
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ber_neg_1'] = streamlit.checkbox(f'K İY AVRUPA BER -1.0 | {Firsthalf_European_Neg1_Tie_Close:.2f}' if isinstance(Firsthalf_European_Neg1_Tie_Close, (int, float)) else f'K İY AVRUPA BER -1.0 | {Firsthalf_European_Neg1_Tie_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ber_neg_1'])
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_dep_neg_1'] = streamlit.checkbox(f'K İY AVRUPA DEP -1.0 | {Firsthalf_European_Neg1_Away_Close:.2f}' if isinstance(Firsthalf_European_Neg1_Away_Close, (int, float)) else f'K İY AVRUPA DEP -1.0 | {Firsthalf_European_Neg1_Away_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_dep_neg_1'])
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ev_neg_1_neg_0_5'] = streamlit.checkbox(f'K İY AVRUPA EV  -1.0, -0.5 | {Firsthalf_European_Neg1_Neg0_5_Home_Close:.2f}' if isinstance(Firsthalf_European_Neg1_Neg0_5_Home_Close, (int, float)) else f'K İY AVRUPA EV  -1.0, -0.5 | {Firsthalf_European_Neg1_Neg0_5_Home_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ev_neg_1_neg_0_5'])
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ber_neg_1_neg_0_5'] = streamlit.checkbox(f'K İY AVRUPA BER -1.0, -0.5 | {Firsthalf_European_Neg1_Neg0_5_Tie_Close:.2f}' if isinstance(Firsthalf_European_Neg1_Neg0_5_Tie_Close, (int, float)) else f'K İY AVRUPA BER -1.0, -0.5 | {Firsthalf_European_Neg1_Neg0_5_Tie_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ber_neg_1_neg_0_5'])
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_dep_neg_1_neg_0_5'] = streamlit.checkbox(f'K İY AVRUPA DEP -1.0, -0.5 | {Firsthalf_European_Neg1_Neg0_5_Away_Close:.2f}' if isinstance(Firsthalf_European_Neg1_Neg0_5_Away_Close, (int, float)) else f'K İY AVRUPA DEP -1.0, -0.5 | {Firsthalf_European_Neg1_Neg0_5_Away_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_dep_neg_1_neg_0_5'])
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ev_neg_0_5'] = streamlit.checkbox(f'K İY AVRUPA EV  -0.5 | {Firsthalf_European_Neg0_5_Home_Close:.2f}' if isinstance(Firsthalf_European_Neg0_5_Home_Close, (int, float)) else f'K İY AVRUPA EV  -0.5 | {Firsthalf_European_Neg0_5_Home_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ev_neg_0_5'])
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ber_neg_0_5'] = streamlit.checkbox(f'K İY AVRUPA BER -0.5 | {Firsthalf_European_Neg0_5_Tie_Close:.2f}' if isinstance(Firsthalf_European_Neg0_5_Tie_Close, (int, float)) else f'K İY AVRUPA BER -0.5 | {Firsthalf_European_Neg0_5_Tie_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ber_neg_0_5'])
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_dep_neg_0_5'] = streamlit.checkbox(f'K İY AVRUPA DEP -0.5 | {Firsthalf_European_Neg0_5_Away_Close:.2f}' if isinstance(Firsthalf_European_Neg0_5_Away_Close, (int, float)) else f'K İY AVRUPA DEP -0.5 | {Firsthalf_European_Neg0_5_Away_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_dep_neg_0_5'])
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ev_neg_0_5_pos_0'] = streamlit.checkbox(f'K İY AVRUPA EV  -0.5, 0.0 | {Firsthalf_European_Neg0_5_0_0_Home_Close:.2f}' if isinstance(Firsthalf_European_Neg0_5_0_0_Home_Close, (int, float)) else f'K İY AVRUPA EV  -0.5, 0.0 | {Firsthalf_European_Neg0_5_0_0_Home_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ev_neg_0_5_pos_0'])
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ber_neg_0_5_pos_0'] = streamlit.checkbox(f'K İY AVRUPA BER -0.5, 0.0 | {Firsthalf_European_Neg0_5_0_0_Tie_Close:.2f}' if isinstance(Firsthalf_European_Neg0_5_0_0_Tie_Close, (int, float)) else f'K İY AVRUPA BER -0.5, 0.0 | {Firsthalf_European_Neg0_5_0_0_Tie_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ber_neg_0_5_pos_0'])
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_dep_neg_0_5_pos_0'] = streamlit.checkbox(f'K İY AVRUPA DEP -0.5, 0.0 | {Firsthalf_European_Neg0_5_0_0_Away_Close:.2f}' if isinstance(Firsthalf_European_Neg0_5_0_0_Away_Close, (int, float)) else f'K İY AVRUPA DEP -0.5, 0.0 | {Firsthalf_European_Neg0_5_0_0_Away_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_dep_neg_0_5_pos_0'])
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ev_pos_0_pos_0_5'] = streamlit.checkbox(f'K İY AVRUPA EV  0.0, 0.5 | {Firsthalf_European_0_0_Pos0_5_Home_Close:.2f}' if isinstance(Firsthalf_European_0_0_Pos0_5_Home_Close, (int, float)) else f'K İY AVRUPA EV  0.0, 0.5 | {Firsthalf_European_0_0_Pos0_5_Home_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ev_pos_0_pos_0_5'])
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ber_pos_0_pos_0_5'] = streamlit.checkbox(f'K İY AVRUPA BER 0.0, 0.5 | {Firsthalf_European_0_0_Pos0_5_Tie_Close:.2f}' if isinstance(Firsthalf_European_0_0_Pos0_5_Tie_Close, (int, float)) else f'K İY AVRUPA BER 0.0, 0.5 | {Firsthalf_European_0_0_Pos0_5_Tie_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ber_pos_0_pos_0_5'])
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_dep_pos_0_pos_0_5'] = streamlit.checkbox(f'K İY AVRUPA DEP 0.0, 0.5 | {Firsthalf_European_0_0_Pos0_5_Away_Close:.2f}' if isinstance(Firsthalf_European_0_0_Pos0_5_Away_Close, (int, float)) else f'K İY AVRUPA DEP 0.0, 0.5 | {Firsthalf_European_0_0_Pos0_5_Away_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_dep_pos_0_pos_0_5'])
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ev_pos_0_5'] = streamlit.checkbox(f'K İY AVRUPA EV  0.5 | {Firsthalf_European_Pos0_5_Home_Close:.2f}' if isinstance(Firsthalf_European_Pos0_5_Home_Close, (int, float)) else f'K İY AVRUPA EV  0.5 | {Firsthalf_European_Pos0_5_Home_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ev_pos_0_5'])
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ber_pos_0_5'] = streamlit.checkbox(f'K İY AVRUPA BER 0.5 | {Firsthalf_European_Pos0_5_Tie_Close:.2f}' if isinstance(Firsthalf_European_Pos0_5_Tie_Close, (int, float)) else f'K İY AVRUPA BER 0.5 | {Firsthalf_European_Pos0_5_Tie_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ber_pos_0_5'])
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_dep_pos_0_5'] = streamlit.checkbox(f'K İY AVRUPA DEP 0.5 | {Firsthalf_European_Pos0_5_Away_Close:.2f}' if isinstance(Firsthalf_European_Pos0_5_Away_Close, (int, float)) else f'K İY AVRUPA DEP 0.5 | {Firsthalf_European_Pos0_5_Away_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_dep_pos_0_5'])
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ev_pos_0_5_pos_1'] = streamlit.checkbox(f'K İY AVRUPA EV  0.5, 1.0 | {Firsthalf_European_Pos0_5_Pos1_0_Home_Close:.2f}' if isinstance(Firsthalf_European_Pos0_5_Pos1_0_Home_Close, (int, float)) else f'K İY AVRUPA EV  0.5, 1.0 | {Firsthalf_European_Pos0_5_Pos1_0_Home_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ev_pos_0_5_pos_1'])
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ber_pos_0_5_pos_1'] = streamlit.checkbox(f'K İY AVRUPA BER 0.5, 1.0 | {Firsthalf_European_Pos0_5_Pos1_0_Tie_Close:.2f}' if isinstance(Firsthalf_European_Pos0_5_Pos1_0_Tie_Close, (int, float)) else f'K İY AVRUPA BER 0.5, 1.0 | {Firsthalf_European_Pos0_5_Pos1_0_Tie_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ber_pos_0_5_pos_1'])
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_dep_pos_0_5_pos_1'] = streamlit.checkbox(f'K İY AVRUPA DEP 0.5, 1.0 | {Firsthalf_European_Pos0_5_Pos1_0_Away_Close:.2f}' if isinstance(Firsthalf_European_Pos0_5_Pos1_0_Away_Close, (int, float)) else f'K İY AVRUPA DEP 0.5, 1.0 | {Firsthalf_European_Pos0_5_Pos1_0_Away_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_dep_pos_0_5_pos_1'])
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ev_pos_1'] = streamlit.checkbox(f'K İY AVRUPA EV  1.0 | {Firsthalf_European_Pos1_0_Home_Close:.2f}' if isinstance(Firsthalf_European_Pos1_0_Home_Close, (int, float)) else f'K İY AVRUPA EV  1.0 | {Firsthalf_European_Pos1_0_Home_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ev_pos_1'])
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ber_pos_1'] = streamlit.checkbox(f'K İY AVRUPA BER 1.0 | {Firsthalf_European_Pos1_0_Tie_Close:.2f}' if isinstance(Firsthalf_European_Pos1_0_Tie_Close, (int, float)) else f'K İY AVRUPA BER 1.0 | {Firsthalf_European_Pos1_0_Tie_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ber_pos_1'])
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_dep_pos_1'] = streamlit.checkbox(f'K İY AVRUPA DEP 1.0 | {Firsthalf_European_Pos1_0_Away_Close:.2f}' if isinstance(Firsthalf_European_Pos1_0_Away_Close, (int, float)) else f'K İY AVRUPA DEP 1.0 | {Firsthalf_European_Pos1_0_Away_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_dep_pos_1'])

                with streamlit.expander("Maç sonu avrupa handikap filtreleri"):
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ev_neg_1'] = streamlit.checkbox(f'K MS AVRUPA EV  -1.0 | {Matchtime_European_Neg1_Home_Close:.2f}' if isinstance(Matchtime_European_Neg1_Home_Close, (int, float)) else f'K MS AVRUPA EV  -1.0 | {Matchtime_European_Neg1_Home_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ev_neg_1'])
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ber_neg_1'] = streamlit.checkbox(f'K MS AVRUPA BER -1.0 | {Matchtime_European_Neg1_Tie_Close:.2f}' if isinstance(Matchtime_European_Neg1_Tie_Close, (int, float)) else f'K MS AVRUPA BER -1.0 | {Matchtime_European_Neg1_Tie_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ber_neg_1'])
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_dep_neg_1'] = streamlit.checkbox(f'K MS AVRUPA DEP -1.0 | {Matchtime_European_Neg1_Away_Close:.2f}' if isinstance(Matchtime_European_Neg1_Away_Close, (int, float)) else f'K MS AVRUPA DEP -1.0 | {Matchtime_European_Neg1_Away_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_dep_neg_1'])
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ev_neg_1_neg_0_5'] = streamlit.checkbox(f'K MS AVRUPA EV  -1.0, -0.5 | {Matchtime_European_Neg1_Neg0_5_Home_Close:.2f}' if isinstance(Matchtime_European_Neg1_Neg0_5_Home_Close, (int, float)) else f'K MS AVRUPA EV  -1.0, -0.5 | {Matchtime_European_Neg1_Neg0_5_Home_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ev_neg_1_neg_0_5'])
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ber_neg_1_neg_0_5'] = streamlit.checkbox(f'K MS AVRUPA BER -1.0, -0.5 | {Matchtime_European_Neg1_Neg0_5_Tie_Close:.2f}' if isinstance(Matchtime_European_Neg1_Neg0_5_Tie_Close, (int, float)) else f'K MS AVRUPA BER -1.0, -0.5 | {Matchtime_European_Neg1_Neg0_5_Tie_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ber_neg_1_neg_0_5'])
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_dep_neg_1_neg_0_5'] = streamlit.checkbox(f'K MS AVRUPA DEP -1.0, -0.5 | {Matchtime_European_Neg1_Neg0_5_Away_Close:.2f}' if isinstance(Matchtime_European_Neg1_Neg0_5_Away_Close, (int, float)) else f'K MS AVRUPA DEP -1.0, -0.5 | {Matchtime_European_Neg1_Neg0_5_Away_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_dep_neg_1_neg_0_5'])
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ev_neg_0_5'] = streamlit.checkbox(f'K MS AVRUPA EV  -0.5 | {Matchtime_European_Neg0_5_Home_Close:.2f}' if isinstance(Matchtime_European_Neg0_5_Home_Close, (int, float)) else f'K MS AVRUPA EV  -0.5 | {Matchtime_European_Neg0_5_Home_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ev_neg_0_5'])
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ber_neg_0_5'] = streamlit.checkbox(f'K MS AVRUPA BER -0.5 | {Matchtime_European_Neg0_5_Tie_Close:.2f}' if isinstance(Matchtime_European_Neg0_5_Tie_Close, (int, float)) else f'K MS AVRUPA BER -0.5 | {Matchtime_European_Neg0_5_Tie_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ber_neg_0_5'])
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_dep_neg_0_5'] = streamlit.checkbox(f'K MS AVRUPA DEP -0.5 | {Matchtime_European_Neg0_5_Away_Close:.2f}' if isinstance(Matchtime_European_Neg0_5_Away_Close, (int, float)) else f'K MS AVRUPA DEP -0.5 | {Matchtime_European_Neg0_5_Away_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_dep_neg_0_5'])
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ev_neg_0_5_pos_0'] = streamlit.checkbox(f'K MS AVRUPA EV  -0.5, 0.0 | {Matchtime_European_Neg0_5_0_0_Home_Close:.2f}' if isinstance(Matchtime_European_Neg0_5_0_0_Home_Close, (int, float)) else f'K MS AVRUPA EV  -0.5, 0.0 | {Matchtime_European_Neg0_5_0_0_Home_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ev_neg_0_5_pos_0'])
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ber_neg_0_5_pos_0'] = streamlit.checkbox(f'K MS AVRUPA BER -0.5, 0.0 | {Matchtime_European_Neg0_5_0_0_Tie_Close:.2f}' if isinstance(Matchtime_European_Neg0_5_0_0_Tie_Close, (int, float)) else f'K MS AVRUPA BER -0.5, 0.0 | {Matchtime_European_Neg0_5_0_0_Tie_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ber_neg_0_5_pos_0'])
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_dep_neg_0_5_pos_0'] = streamlit.checkbox(f'K MS AVRUPA DEP -0.5, 0.0 | {Matchtime_European_Neg0_5_0_0_Away_Close:.2f}' if isinstance(Matchtime_European_Neg0_5_0_0_Away_Close, (int, float)) else f'K MS AVRUPA DEP -0.5, 0.0 | {Matchtime_European_Neg0_5_0_0_Away_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_dep_neg_0_5_pos_0'])
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ev_pos_0_pos_0_5'] = streamlit.checkbox(f'K MS AVRUPA EV  0.0, 0.5 | {Matchtime_European_0_0_Pos0_5_Home_Close:.2f}' if isinstance(Matchtime_European_0_0_Pos0_5_Home_Close, (int, float)) else f'K MS AVRUPA EV  0.0, 0.5 | {Matchtime_European_0_0_Pos0_5_Home_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ev_pos_0_pos_0_5'])
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ber_pos_0_pos_0_5'] = streamlit.checkbox(f'K MS AVRUPA BER 0.0, 0.5 | {Matchtime_European_0_0_Pos0_5_Tie_Close:.2f}' if isinstance(Matchtime_European_0_0_Pos0_5_Tie_Close, (int, float)) else f'K MS AVRUPA BER 0.0, 0.5 | {Matchtime_European_0_0_Pos0_5_Tie_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ber_pos_0_pos_0_5'])
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_dep_pos_0_pos_0_5'] = streamlit.checkbox(f'K MS AVRUPA DEP 0.0, 0.5 | {Matchtime_European_0_0_Pos0_5_Away_Close:.2f}' if isinstance(Matchtime_European_0_0_Pos0_5_Away_Close, (int, float)) else f'K MS AVRUPA DEP 0.0, 0.5 | {Matchtime_European_0_0_Pos0_5_Away_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_dep_pos_0_pos_0_5'])
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ev_pos_0_5'] = streamlit.checkbox(f'K MS AVRUPA EV  0.5 | {Matchtime_European_Pos0_5_Home_Close:.2f}' if isinstance(Matchtime_European_Pos0_5_Home_Close, (int, float)) else f'K MS AVRUPA EV  0.5 | {Matchtime_European_Pos0_5_Home_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ev_pos_0_5'])
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ber_pos_0_5'] = streamlit.checkbox(f'K MS AVRUPA BER 0.5 | {Matchtime_European_Pos0_5_Tie_Close:.2f}' if isinstance(Matchtime_European_Pos0_5_Tie_Close, (int, float)) else f'K MS AVRUPA BER 0.5 | {Matchtime_European_Pos0_5_Tie_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ber_pos_0_5'])
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_dep_pos_0_5'] = streamlit.checkbox(f'K MS AVRUPA DEP 0.5 | {Matchtime_European_Pos0_5_Away_Close:.2f}' if isinstance(Matchtime_European_Pos0_5_Away_Close, (int, float)) else f'K MS AVRUPA DEP 0.5 | {Matchtime_European_Pos0_5_Away_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_dep_pos_0_5'])
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ev_pos_0_5_pos_1'] = streamlit.checkbox(f'K MS AVRUPA EV  0.5, 1.0 | {Matchtime_European_Pos0_5_Pos1_0_Home_Close:.2f}' if isinstance(Matchtime_European_Pos0_5_Pos1_0_Home_Close, (int, float)) else f'K MS AVRUPA EV  0.5, 1.0 | {Matchtime_European_Pos0_5_Pos1_0_Home_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ev_pos_0_5_pos_1'])
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ber_pos_0_5_pos_1'] = streamlit.checkbox(f'K MS AVRUPA BER 0.5, 1.0 | {Matchtime_European_Pos0_5_Pos1_0_Tie_Close:.2f}' if isinstance(Matchtime_European_Pos0_5_Pos1_0_Tie_Close, (int, float)) else f'K MS AVRUPA BER 0.5, 1.0 | {Matchtime_European_Pos0_5_Pos1_0_Tie_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ber_pos_0_5_pos_1'])
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_dep_pos_0_5_pos_1'] = streamlit.checkbox(f'K MS AVRUPA DEP 0.5, 1.0 | {Matchtime_European_Pos0_5_Pos1_0_Away_Close:.2f}' if isinstance(Matchtime_European_Pos0_5_Pos1_0_Away_Close, (int, float)) else f'K MS AVRUPA DEP 0.5, 1.0 | {Matchtime_European_Pos0_5_Pos1_0_Away_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_dep_pos_0_5_pos_1'])
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ev_pos_1'] = streamlit.checkbox(f'K MS AVRUPA EV  1.0 | {Matchtime_European_Pos1_0_Home_Close:.2f}' if isinstance(Matchtime_European_Pos1_0_Home_Close, (int, float)) else f'K MS AVRUPA EV  1.0 | {Matchtime_European_Pos1_0_Home_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ev_pos_1'])
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ber_pos_1'] = streamlit.checkbox(f'K MS AVRUPA BER 1.0 | {Matchtime_European_Pos1_0_Tie_Close:.2f}' if isinstance(Matchtime_European_Pos1_0_Tie_Close, (int, float)) else f'K MS AVRUPA BER 1.0 | {Matchtime_European_Pos1_0_Tie_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ber_pos_1'])
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_dep_pos_1'] = streamlit.checkbox(f'K MS AVRUPA DEP 1.0 | {Matchtime_European_Pos1_0_Away_Close:.2f}' if isinstance(Matchtime_European_Pos1_0_Away_Close, (int, float)) else f'K MS AVRUPA DEP 1.0 | {Matchtime_European_Pos1_0_Away_Close}', value=streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_dep_pos_1'])
with Row_4[2]:
    if len(streamlit.session_state.Bulletin_Match_List) == 0:
        streamlit.button(label='💾', key='button_save_filter', disabled=True)
    else:
        streamlit.button(label='💾', key='button_save_filter', disabled=False)
with Row_4[3]:
    if len(streamlit.session_state.Bulletin_Match_List) == 0 or streamlit.session_state.selectbox_saved_filters is None:
        streamlit.button(label='🗑️', key='button_delete_filter', disabled=True)
    else:
        streamlit.button(label='🗑️', key='button_delete_filter', disabled=False)


### ROW 4 CODE END ###


# FİLTRE KAYDETME
@streamlit.dialog("Filtre İşlemi")
def Save_Dialog(item):
    if item == "Kaydet":  # Filtre kaydetme işlemi yapıyorum
        CloseDialog = False
        streamlit.text_input(label="Filtre ismi:", key="dialog_filter_name")
        streamlit.button(label="Evet", key="button_dialog_yes")
        streamlit.button(label="Hayır", key="button_dialog_no")

        if streamlit.session_state.button_dialog_yes:
            if streamlit.session_state.dialog_filter_name == "":
                streamlit.error("Filtre ismi boş bırakılamaz!", icon="🚨")
            else:
                # Database bağlantısı
                Connection = sqlite3.connect(Database_Path)

                # SQL komutlarını uygulama
                Cursor = Connection.cursor()

                # Eklenecek verinin tanımlanması
                data_to_insert = (
                    streamlit.session_state.dialog_filter_name,
                    streamlit.session_state.database_start_date,
                    streamlit.session_state.database_end_date,
                    streamlit.session_state.Filters_Dictionary['filter_country'],
                    streamlit.session_state.Filters_Dictionary['filter_league'],
                    streamlit.session_state.Filters_Dictionary['filter_week'],
                    streamlit.session_state.Filters_Dictionary['filter_hour'],
                    streamlit.session_state.Filters_Dictionary['filter_home_team'],
                    streamlit.session_state.Filters_Dictionary['filter_away_team'],
                    streamlit.session_state.Filters_Dictionary['filter_referee'],
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_1'],
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_x'],
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_2'],
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_0_5_a'],
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_0_5_ü'],
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_1_5_a'],
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_1_5_ü'],
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_2_5_a'],
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_2_5_ü'],
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_3_5_a'],
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_3_5_ü'],
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_1'],
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_x'],
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_2'],
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_0_5_a'],
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_0_5_ü'],
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_1_5_a'],
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_1_5_ü'],
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_1_x_cs'],
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_1_2_cs'],
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_x_2_cs'],
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_1_x_cs'],
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_1_2_cs'],
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_x_2_cs'],
                    streamlit.session_state.Filters_Dictionary['filter_a_home_win'],
                    streamlit.session_state.Filters_Dictionary['filter_a_away_win'],
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_kg_var'],
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_kg_yok'],
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_kg_var'],
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_kg_yok'],
                    streamlit.session_state.Filters_Dictionary['filter_a_tek'],
                    streamlit.session_state.Filters_Dictionary['filter_a_cift'],
                    streamlit.session_state.Filters_Dictionary['filter_a_1den1'],
                    streamlit.session_state.Filters_Dictionary['filter_a_1denx'],
                    streamlit.session_state.Filters_Dictionary['filter_a_1den2'],
                    streamlit.session_state.Filters_Dictionary['filter_a_xden1'],
                    streamlit.session_state.Filters_Dictionary['filter_a_xdenx'],
                    streamlit.session_state.Filters_Dictionary['filter_a_xden2'],
                    streamlit.session_state.Filters_Dictionary['filter_a_2den1'],
                    streamlit.session_state.Filters_Dictionary['filter_a_2denx'],
                    streamlit.session_state.Filters_Dictionary['filter_a_2den2'],
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_10'],
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_20'],
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_21'],
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_30'],
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_31'],
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_32'],
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_40'],
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_41'],
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_42'],
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_43'],
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_00'],
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_11'],
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_22'],
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_33'],
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_44'],
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_01'],
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_02'],
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_12'],
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_03'],
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_13'],
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_23'],
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_04'],
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_14'],
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_24'],
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_34'],
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_10'],
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_20'],
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_21'],
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_30'],
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_31'],
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_32'],
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_40'],
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_41'],
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_42'],
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_43'],
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_00'],
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_11'],
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_22'],
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_33'],
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_44'],
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_01'],
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_02'],
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_12'],
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_03'],
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_13'],
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_23'],
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_04'],
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_14'],
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_24'],
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_34'],
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_neg_2'],
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_neg_2'],
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_neg_2_neg_1_5'],
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_neg_2_neg_1_5'],
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_neg_1_5'],
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_neg_1_5'],
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_neg_1_5_neg_1'],
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_neg_1_5_neg_1'],
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_neg_1'],
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_neg_1'],
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_neg_1_neg_0_5'],
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_neg_1_neg_0_5'],
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_neg_0_5'],
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_neg_0_5'],
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_neg_0_5_pos_0'],
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_neg_0_5_pos_0'],
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_pos_0_pos_0_5'],
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_pos_0_pos_0_5'],
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_pos_0_5'],
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_pos_0_5'],
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_pos_0_5_pos_1'],
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_pos_0_5_pos_1'],
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_pos_1'],
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_pos_1'],
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_pos_1_pos_1_5'],
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_pos_1_pos_1_5'],
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_pos_1_5'],
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_pos_1_5'],
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_pos_1_5_pos_2'],
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_pos_1_5_pos_2'],
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_pos_2'],
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_pos_2'],
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_neg_2'],
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_neg_2'],
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_neg_2_neg_1_5'],
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_neg_2_neg_1_5'],
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_neg_1_5'],
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_neg_1_5'],
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_neg_1_5_neg_1'],
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_neg_1_5_neg_1'],
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_neg_1'],
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_neg_1'],
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_neg_1_neg_0_5'],
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_neg_1_neg_0_5'],
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_neg_0_5'],
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_neg_0_5'],
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_neg_0_5_pos_0'],
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_neg_0_5_pos_0'],
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_pos_0_pos_0_5'],
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_pos_0_pos_0_5'],
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_pos_0_5'],
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_pos_0_5'],
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_pos_0_5_pos_1'],
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_pos_0_5_pos_1'],
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_pos_1'],
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_pos_1'],
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_pos_1_pos_1_5'],
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_pos_1_pos_1_5'],
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_pos_1_5'],
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_pos_1_5'],
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_pos_1_5_pos_2'],
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_pos_1_5_pos_2'],
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_pos_2'],
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_pos_2'],
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ev_neg_1'],
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ber_neg_1'],
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_dep_neg_1'],
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ev_neg_1_neg_0_5'],
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ber_neg_1_neg_0_5'],
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_dep_neg_1_neg_0_5'],
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ev_neg_0_5'],
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ber_neg_0_5'],
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_dep_neg_0_5'],
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ev_neg_0_5_pos_0'],
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ber_neg_0_5_pos_0'],
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_dep_neg_0_5_pos_0'],
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ev_pos_0_pos_0_5'],
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ber_pos_0_pos_0_5'],
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_dep_pos_0_pos_0_5'],
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ev_pos_0_5'],
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ber_pos_0_5'],
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_dep_pos_0_5'],
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ev_pos_0_5_pos_1'],
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ber_pos_0_5_pos_1'],
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_dep_pos_0_5_pos_1'],
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ev_pos_1'],
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ber_pos_1'],
                    streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_dep_pos_1'],
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ev_neg_1'],
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ber_neg_1'],
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_dep_neg_1'],
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ev_neg_1_neg_0_5'],
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ber_neg_1_neg_0_5'],
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_dep_neg_1_neg_0_5'],
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ev_neg_0_5'],
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ber_neg_0_5'],
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_dep_neg_0_5'],
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ev_neg_0_5_pos_0'],
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ber_neg_0_5_pos_0'],
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_dep_neg_0_5_pos_0'],
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ev_pos_0_pos_0_5'],
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ber_pos_0_pos_0_5'],
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_dep_pos_0_pos_0_5'],
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ev_pos_0_5'],
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ber_pos_0_5'],
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_dep_pos_0_5'],
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ev_pos_0_5_pos_1'],
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ber_pos_0_5_pos_1'],
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_dep_pos_0_5_pos_1'],
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ev_pos_1'],
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ber_pos_1'],
                    streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_dep_pos_1'],
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_1'],
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_x'],
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_2'],
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_0_5_a'],
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_0_5_ü'],
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_1_5_a'],
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_1_5_ü'],
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_2_5_a'],
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_2_5_ü'],
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_3_5_a'],
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_3_5_ü'],
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_1'],
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_x'],
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_2'],
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_0_5_a'],
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_0_5_ü'],
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_1_5_a'],
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_1_5_ü'],
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_1_x_cs'],
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_1_2_cs'],
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_x_2_cs'],
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_1_x_cs'],
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_1_2_cs'],
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_x_2_cs'],
                    streamlit.session_state.Filters_Dictionary['filter_k_home_win'],
                    streamlit.session_state.Filters_Dictionary['filter_k_away_win'],
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_kg_var'],
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_kg_yok'],
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_kg_var'],
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_kg_yok'],
                    streamlit.session_state.Filters_Dictionary['filter_k_tek'],
                    streamlit.session_state.Filters_Dictionary['filter_k_cift'],
                    streamlit.session_state.Filters_Dictionary['filter_k_1den1'],
                    streamlit.session_state.Filters_Dictionary['filter_k_1denx'],
                    streamlit.session_state.Filters_Dictionary['filter_k_1den2'],
                    streamlit.session_state.Filters_Dictionary['filter_k_xden1'],
                    streamlit.session_state.Filters_Dictionary['filter_k_xdenx'],
                    streamlit.session_state.Filters_Dictionary['filter_k_xden2'],
                    streamlit.session_state.Filters_Dictionary['filter_k_2den1'],
                    streamlit.session_state.Filters_Dictionary['filter_k_2denx'],
                    streamlit.session_state.Filters_Dictionary['filter_k_2den2'],
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_10'],
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_20'],
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_21'],
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_30'],
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_31'],
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_32'],
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_40'],
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_41'],
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_42'],
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_43'],
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_00'],
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_11'],
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_22'],
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_33'],
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_44'],
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_01'],
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_02'],
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_12'],
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_03'],
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_13'],
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_23'],
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_04'],
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_14'],
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_24'],
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_34'],
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_10'],
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_20'],
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_21'],
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_30'],
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_31'],
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_32'],
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_40'],
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_41'],
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_42'],
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_43'],
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_00'],
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_11'],
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_22'],
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_33'],
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_44'],
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_01'],
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_02'],
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_12'],
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_03'],
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_13'],
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_23'],
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_04'],
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_14'],
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_24'],
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_34'],
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_neg_2'],
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_neg_2'],
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_neg_2_neg_1_5'],
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_neg_2_neg_1_5'],
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_neg_1_5'],
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_neg_1_5'],
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_neg_1_5_neg_1'],
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_neg_1_5_neg_1'],
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_neg_1'],
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_neg_1'],
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_neg_1_neg_0_5'],
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_neg_1_neg_0_5'],
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_neg_0_5'],
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_neg_0_5'],
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_neg_0_5_pos_0'],
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_neg_0_5_pos_0'],
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_pos_0_pos_0_5'],
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_pos_0_pos_0_5'],
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_pos_0_5'],
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_pos_0_5'],
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_pos_0_5_pos_1'],
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_pos_0_5_pos_1'],
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_pos_1'],
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_pos_1'],
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_pos_1_pos_1_5'],
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_pos_1_pos_1_5'],
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_pos_1_5'],
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_pos_1_5'],
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_pos_1_5_pos_2'],
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_pos_1_5_pos_2'],
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_pos_2'],
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_pos_2'],
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_neg_2'],
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_neg_2'],
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_neg_2_neg_1_5'],
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_neg_2_neg_1_5'],
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_neg_1_5'],
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_neg_1_5'],
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_neg_1_5_neg_1'],
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_neg_1_5_neg_1'],
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_neg_1'],
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_neg_1'],
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_neg_1_neg_0_5'],
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_neg_1_neg_0_5'],
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_neg_0_5'],
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_neg_0_5'],
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_neg_0_5_pos_0'],
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_neg_0_5_pos_0'],
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_pos_0_pos_0_5'],
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_pos_0_pos_0_5'],
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_pos_0_5'],
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_pos_0_5'],
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_pos_0_5_pos_1'],
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_pos_0_5_pos_1'],
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_pos_1'],
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_pos_1'],
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_pos_1_pos_1_5'],
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_pos_1_pos_1_5'],
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_pos_1_5'],
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_pos_1_5'],
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_pos_1_5_pos_2'],
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_pos_1_5_pos_2'],
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_pos_2'],
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_pos_2'],
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ev_neg_1'],
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ber_neg_1'],
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_dep_neg_1'],
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ev_neg_1_neg_0_5'],
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ber_neg_1_neg_0_5'],
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_dep_neg_1_neg_0_5'],
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ev_neg_0_5'],
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ber_neg_0_5'],
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_dep_neg_0_5'],
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ev_neg_0_5_pos_0'],
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ber_neg_0_5_pos_0'],
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_dep_neg_0_5_pos_0'],
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ev_pos_0_pos_0_5'],
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ber_pos_0_pos_0_5'],
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_dep_pos_0_pos_0_5'],
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ev_pos_0_5'],
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ber_pos_0_5'],
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_dep_pos_0_5'],
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ev_pos_0_5_pos_1'],
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ber_pos_0_5_pos_1'],
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_dep_pos_0_5_pos_1'],
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ev_pos_1'],
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ber_pos_1'],
                    streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_dep_pos_1'],
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ev_neg_1'],
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ber_neg_1'],
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_dep_neg_1'],
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ev_neg_1_neg_0_5'],
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ber_neg_1_neg_0_5'],
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_dep_neg_1_neg_0_5'],
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ev_neg_0_5'],
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ber_neg_0_5'],
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_dep_neg_0_5'],
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ev_neg_0_5_pos_0'],
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ber_neg_0_5_pos_0'],
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_dep_neg_0_5_pos_0'],
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ev_pos_0_pos_0_5'],
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ber_pos_0_pos_0_5'],
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_dep_pos_0_pos_0_5'],
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ev_pos_0_5'],
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ber_pos_0_5'],
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_dep_pos_0_5'],
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ev_pos_0_5_pos_1'],
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ber_pos_0_5_pos_1'],
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_dep_pos_0_5_pos_1'],
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ev_pos_1'],
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ber_pos_1'],
                    streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_dep_pos_1']
                )

                # Ekleme metni
                insert_statement = '''
                                INSERT INTO saved_filters (
                                    filter_name,
                                    filter_start_date,
                                    filter_end_date,
                                    filter_country,
                                    filter_league,
                                    filter_week,
                                    filter_hour,
                                    filter_home_team,
                                    filter_away_team,
                                    filter_referee,
                                    filter_a_ms_1,
                                    filter_a_ms_x,
                                    filter_a_ms_2,
                                    filter_a_ms_0_5_a,
                                    filter_a_ms_0_5_ü,
                                    filter_a_ms_1_5_a,
                                    filter_a_ms_1_5_ü,
                                    filter_a_ms_2_5_a,
                                    filter_a_ms_2_5_ü,
                                    filter_a_ms_3_5_a,
                                    filter_a_ms_3_5_ü,
                                    filter_a_iy_1,
                                    filter_a_iy_x,
                                    filter_a_iy_2,
                                    filter_a_iy_0_5_a,
                                    filter_a_iy_0_5_ü,
                                    filter_a_iy_1_5_a,
                                    filter_a_iy_1_5_ü,
                                    filter_a_ms_1_x_cs,
                                    filter_a_ms_1_2_cs,
                                    filter_a_ms_x_2_cs,
                                    filter_a_iy_1_x_cs,
                                    filter_a_iy_1_2_cs,
                                    filter_a_iy_x_2_cs,
                                    filter_a_home_win,
                                    filter_a_away_win,
                                    filter_a_iy_kg_var,
                                    filter_a_iy_kg_yok,
                                    filter_a_ms_kg_var,
                                    filter_a_ms_kg_yok,
                                    filter_a_tek,
                                    filter_a_cift,
                                    filter_a_1den1,
                                    filter_a_1denx,
                                    filter_a_1den2,
                                    filter_a_xden1,
                                    filter_a_xdenx,
                                    filter_a_xden2,
                                    filter_a_2den1,
                                    filter_a_2denx,
                                    filter_a_2den2,
                                    filter_a_iy_skor_10,
                                    filter_a_iy_skor_20,
                                    filter_a_iy_skor_21,
                                    filter_a_iy_skor_30,
                                    filter_a_iy_skor_31,
                                    filter_a_iy_skor_32,
                                    filter_a_iy_skor_40,
                                    filter_a_iy_skor_41,
                                    filter_a_iy_skor_42,
                                    filter_a_iy_skor_43,
                                    filter_a_iy_skor_00,
                                    filter_a_iy_skor_11,
                                    filter_a_iy_skor_22,
                                    filter_a_iy_skor_33,
                                    filter_a_iy_skor_44,
                                    filter_a_iy_skor_01,
                                    filter_a_iy_skor_02,
                                    filter_a_iy_skor_12,
                                    filter_a_iy_skor_03,
                                    filter_a_iy_skor_13,
                                    filter_a_iy_skor_23,
                                    filter_a_iy_skor_04,
                                    filter_a_iy_skor_14,
                                    filter_a_iy_skor_24,
                                    filter_a_iy_skor_34,
                                    filter_a_ms_skor_10,
                                    filter_a_ms_skor_20,
                                    filter_a_ms_skor_21,
                                    filter_a_ms_skor_30,
                                    filter_a_ms_skor_31,
                                    filter_a_ms_skor_32,
                                    filter_a_ms_skor_40,
                                    filter_a_ms_skor_41,
                                    filter_a_ms_skor_42,
                                    filter_a_ms_skor_43,
                                    filter_a_ms_skor_00,
                                    filter_a_ms_skor_11,
                                    filter_a_ms_skor_22,
                                    filter_a_ms_skor_33,
                                    filter_a_ms_skor_44,
                                    filter_a_ms_skor_01,
                                    filter_a_ms_skor_02,
                                    filter_a_ms_skor_12,
                                    filter_a_ms_skor_03,
                                    filter_a_ms_skor_13,
                                    filter_a_ms_skor_23,
                                    filter_a_ms_skor_04,
                                    filter_a_ms_skor_14,
                                    filter_a_ms_skor_24,
                                    filter_a_ms_skor_34,
                                    filter_a_iy_asya_ev_neg_2,
                                    filter_a_iy_asya_dep_neg_2,
                                    filter_a_iy_asya_ev_neg_2_neg_1_5,
                                    filter_a_iy_asya_dep_neg_2_neg_1_5,
                                    filter_a_iy_asya_ev_neg_1_5,
                                    filter_a_iy_asya_dep_neg_1_5,
                                    filter_a_iy_asya_ev_neg_1_5_neg_1,
                                    filter_a_iy_asya_dep_neg_1_5_neg_1,
                                    filter_a_iy_asya_ev_neg_1,
                                    filter_a_iy_asya_dep_neg_1,
                                    filter_a_iy_asya_ev_neg_1_neg_0_5,
                                    filter_a_iy_asya_dep_neg_1_neg_0_5,
                                    filter_a_iy_asya_ev_neg_0_5,
                                    filter_a_iy_asya_dep_neg_0_5,
                                    filter_a_iy_asya_ev_neg_0_5_pos_0,
                                    filter_a_iy_asya_dep_neg_0_5_pos_0,
                                    filter_a_iy_asya_ev_pos_0_pos_0_5,
                                    filter_a_iy_asya_dep_pos_0_pos_0_5,
                                    filter_a_iy_asya_ev_pos_0_5,
                                    filter_a_iy_asya_dep_pos_0_5,
                                    filter_a_iy_asya_ev_pos_0_5_pos_1,
                                    filter_a_iy_asya_dep_pos_0_5_pos_1,
                                    filter_a_iy_asya_ev_pos_1,
                                    filter_a_iy_asya_dep_pos_1,
                                    filter_a_iy_asya_ev_pos_1_pos_1_5,
                                    filter_a_iy_asya_dep_pos_1_pos_1_5,
                                    filter_a_iy_asya_ev_pos_1_5,
                                    filter_a_iy_asya_dep_pos_1_5,
                                    filter_a_iy_asya_ev_pos_1_5_pos_2,
                                    filter_a_iy_asya_dep_pos_1_5_pos_2,
                                    filter_a_iy_asya_ev_pos_2,
                                    filter_a_iy_asya_dep_pos_2,
                                    filter_a_ms_asya_ev_neg_2,
                                    filter_a_ms_asya_dep_neg_2,
                                    filter_a_ms_asya_ev_neg_2_neg_1_5,
                                    filter_a_ms_asya_dep_neg_2_neg_1_5,
                                    filter_a_ms_asya_ev_neg_1_5,
                                    filter_a_ms_asya_dep_neg_1_5,
                                    filter_a_ms_asya_ev_neg_1_5_neg_1,
                                    filter_a_ms_asya_dep_neg_1_5_neg_1,
                                    filter_a_ms_asya_ev_neg_1,
                                    filter_a_ms_asya_dep_neg_1,
                                    filter_a_ms_asya_ev_neg_1_neg_0_5,
                                    filter_a_ms_asya_dep_neg_1_neg_0_5,
                                    filter_a_ms_asya_ev_neg_0_5,
                                    filter_a_ms_asya_dep_neg_0_5,
                                    filter_a_ms_asya_ev_neg_0_5_pos_0,
                                    filter_a_ms_asya_dep_neg_0_5_pos_0,
                                    filter_a_ms_asya_ev_pos_0_pos_0_5,
                                    filter_a_ms_asya_dep_pos_0_pos_0_5,
                                    filter_a_ms_asya_ev_pos_0_5,
                                    filter_a_ms_asya_dep_pos_0_5,
                                    filter_a_ms_asya_ev_pos_0_5_pos_1,
                                    filter_a_ms_asya_dep_pos_0_5_pos_1,
                                    filter_a_ms_asya_ev_pos_1,
                                    filter_a_ms_asya_dep_pos_1,
                                    filter_a_ms_asya_ev_pos_1_pos_1_5,
                                    filter_a_ms_asya_dep_pos_1_pos_1_5,
                                    filter_a_ms_asya_ev_pos_1_5,
                                    filter_a_ms_asya_dep_pos_1_5,
                                    filter_a_ms_asya_ev_pos_1_5_pos_2,
                                    filter_a_ms_asya_dep_pos_1_5_pos_2,
                                    filter_a_ms_asya_ev_pos_2,
                                    filter_a_ms_asya_dep_pos_2,
                                    filter_a_iy_avrupa_ev_neg_1,
                                    filter_a_iy_avrupa_ber_neg_1,
                                    filter_a_iy_avrupa_dep_neg_1,
                                    filter_a_iy_avrupa_ev_neg_1_neg_0_5,
                                    filter_a_iy_avrupa_ber_neg_1_neg_0_5,
                                    filter_a_iy_avrupa_dep_neg_1_neg_0_5,
                                    filter_a_iy_avrupa_ev_neg_0_5,
                                    filter_a_iy_avrupa_ber_neg_0_5,
                                    filter_a_iy_avrupa_dep_neg_0_5,
                                    filter_a_iy_avrupa_ev_neg_0_5_pos_0,
                                    filter_a_iy_avrupa_ber_neg_0_5_pos_0,
                                    filter_a_iy_avrupa_dep_neg_0_5_pos_0,
                                    filter_a_iy_avrupa_ev_pos_0_pos_0_5,
                                    filter_a_iy_avrupa_ber_pos_0_pos_0_5,
                                    filter_a_iy_avrupa_dep_pos_0_pos_0_5,
                                    filter_a_iy_avrupa_ev_pos_0_5,
                                    filter_a_iy_avrupa_ber_pos_0_5,
                                    filter_a_iy_avrupa_dep_pos_0_5,
                                    filter_a_iy_avrupa_ev_pos_0_5_pos_1,
                                    filter_a_iy_avrupa_ber_pos_0_5_pos_1,
                                    filter_a_iy_avrupa_dep_pos_0_5_pos_1,
                                    filter_a_iy_avrupa_ev_pos_1,
                                    filter_a_iy_avrupa_ber_pos_1,
                                    filter_a_iy_avrupa_dep_pos_1,
                                    filter_a_ms_avrupa_ev_neg_1,
                                    filter_a_ms_avrupa_ber_neg_1,
                                    filter_a_ms_avrupa_dep_neg_1,
                                    filter_a_ms_avrupa_ev_neg_1_neg_0_5,
                                    filter_a_ms_avrupa_ber_neg_1_neg_0_5,
                                    filter_a_ms_avrupa_dep_neg_1_neg_0_5,
                                    filter_a_ms_avrupa_ev_neg_0_5,
                                    filter_a_ms_avrupa_ber_neg_0_5,
                                    filter_a_ms_avrupa_dep_neg_0_5,
                                    filter_a_ms_avrupa_ev_neg_0_5_pos_0,
                                    filter_a_ms_avrupa_ber_neg_0_5_pos_0,
                                    filter_a_ms_avrupa_dep_neg_0_5_pos_0,
                                    filter_a_ms_avrupa_ev_pos_0_pos_0_5,
                                    filter_a_ms_avrupa_ber_pos_0_pos_0_5,
                                    filter_a_ms_avrupa_dep_pos_0_pos_0_5,
                                    filter_a_ms_avrupa_ev_pos_0_5,
                                    filter_a_ms_avrupa_ber_pos_0_5,
                                    filter_a_ms_avrupa_dep_pos_0_5,
                                    filter_a_ms_avrupa_ev_pos_0_5_pos_1,
                                    filter_a_ms_avrupa_ber_pos_0_5_pos_1,
                                    filter_a_ms_avrupa_dep_pos_0_5_pos_1,
                                    filter_a_ms_avrupa_ev_pos_1,
                                    filter_a_ms_avrupa_ber_pos_1,
                                    filter_a_ms_avrupa_dep_pos_1,
                                    filter_k_ms_1,
                                    filter_k_ms_x,
                                    filter_k_ms_2,
                                    filter_k_ms_0_5_a,
                                    filter_k_ms_0_5_ü,
                                    filter_k_ms_1_5_a,
                                    filter_k_ms_1_5_ü,
                                    filter_k_ms_2_5_a,
                                    filter_k_ms_2_5_ü,
                                    filter_k_ms_3_5_a,
                                    filter_k_ms_3_5_ü,
                                    filter_k_iy_1,
                                    filter_k_iy_x,
                                    filter_k_iy_2,
                                    filter_k_iy_0_5_a,
                                    filter_k_iy_0_5_ü,
                                    filter_k_iy_1_5_a,
                                    filter_k_iy_1_5_ü,
                                    filter_k_ms_1_x_cs,
                                    filter_k_ms_1_2_cs,
                                    filter_k_ms_x_2_cs,
                                    filter_k_iy_1_x_cs,
                                    filter_k_iy_1_2_cs,
                                    filter_k_iy_x_2_cs,
                                    filter_k_home_win,
                                    filter_k_away_win,
                                    filter_k_iy_kg_var,
                                    filter_k_iy_kg_yok,
                                    filter_k_ms_kg_var,
                                    filter_k_ms_kg_yok,
                                    filter_k_tek,
                                    filter_k_cift,
                                    filter_k_1den1,
                                    filter_k_1denx,
                                    filter_k_1den2,
                                    filter_k_xden1,
                                    filter_k_xdenx,
                                    filter_k_xden2,
                                    filter_k_2den1,
                                    filter_k_2denx,
                                    filter_k_2den2,
                                    filter_k_iy_skor_10,
                                    filter_k_iy_skor_20,
                                    filter_k_iy_skor_21,
                                    filter_k_iy_skor_30,
                                    filter_k_iy_skor_31,
                                    filter_k_iy_skor_32,
                                    filter_k_iy_skor_40,
                                    filter_k_iy_skor_41,
                                    filter_k_iy_skor_42,
                                    filter_k_iy_skor_43,
                                    filter_k_iy_skor_00,
                                    filter_k_iy_skor_11,
                                    filter_k_iy_skor_22,
                                    filter_k_iy_skor_33,
                                    filter_k_iy_skor_44,
                                    filter_k_iy_skor_01,
                                    filter_k_iy_skor_02,
                                    filter_k_iy_skor_12,
                                    filter_k_iy_skor_03,
                                    filter_k_iy_skor_13,
                                    filter_k_iy_skor_23,
                                    filter_k_iy_skor_04,
                                    filter_k_iy_skor_14,
                                    filter_k_iy_skor_24,
                                    filter_k_iy_skor_34,
                                    filter_k_ms_skor_10,
                                    filter_k_ms_skor_20,
                                    filter_k_ms_skor_21,
                                    filter_k_ms_skor_30,
                                    filter_k_ms_skor_31,
                                    filter_k_ms_skor_32,
                                    filter_k_ms_skor_40,
                                    filter_k_ms_skor_41,
                                    filter_k_ms_skor_42,
                                    filter_k_ms_skor_43,
                                    filter_k_ms_skor_00,
                                    filter_k_ms_skor_11,
                                    filter_k_ms_skor_22,
                                    filter_k_ms_skor_33,
                                    filter_k_ms_skor_44,
                                    filter_k_ms_skor_01,
                                    filter_k_ms_skor_02,
                                    filter_k_ms_skor_12,
                                    filter_k_ms_skor_03,
                                    filter_k_ms_skor_13,
                                    filter_k_ms_skor_23,
                                    filter_k_ms_skor_04,
                                    filter_k_ms_skor_14,
                                    filter_k_ms_skor_24,
                                    filter_k_ms_skor_34,
                                    filter_k_iy_asya_ev_neg_2,
                                    filter_k_iy_asya_dep_neg_2,
                                    filter_k_iy_asya_ev_neg_2_neg_1_5,
                                    filter_k_iy_asya_dep_neg_2_neg_1_5,
                                    filter_k_iy_asya_ev_neg_1_5,
                                    filter_k_iy_asya_dep_neg_1_5,
                                    filter_k_iy_asya_ev_neg_1_5_neg_1,
                                    filter_k_iy_asya_dep_neg_1_5_neg_1,
                                    filter_k_iy_asya_ev_neg_1,
                                    filter_k_iy_asya_dep_neg_1,
                                    filter_k_iy_asya_ev_neg_1_neg_0_5,
                                    filter_k_iy_asya_dep_neg_1_neg_0_5,
                                    filter_k_iy_asya_ev_neg_0_5,
                                    filter_k_iy_asya_dep_neg_0_5,
                                    filter_k_iy_asya_ev_neg_0_5_pos_0,
                                    filter_k_iy_asya_dep_neg_0_5_pos_0,
                                    filter_k_iy_asya_ev_pos_0_pos_0_5,
                                    filter_k_iy_asya_dep_pos_0_pos_0_5,
                                    filter_k_iy_asya_ev_pos_0_5,
                                    filter_k_iy_asya_dep_pos_0_5,
                                    filter_k_iy_asya_ev_pos_0_5_pos_1,
                                    filter_k_iy_asya_dep_pos_0_5_pos_1,
                                    filter_k_iy_asya_ev_pos_1,
                                    filter_k_iy_asya_dep_pos_1,
                                    filter_k_iy_asya_ev_pos_1_pos_1_5,
                                    filter_k_iy_asya_dep_pos_1_pos_1_5,
                                    filter_k_iy_asya_ev_pos_1_5,
                                    filter_k_iy_asya_dep_pos_1_5,
                                    filter_k_iy_asya_ev_pos_1_5_pos_2,
                                    filter_k_iy_asya_dep_pos_1_5_pos_2,
                                    filter_k_iy_asya_ev_pos_2,
                                    filter_k_iy_asya_dep_pos_2,
                                    filter_k_ms_asya_ev_neg_2,
                                    filter_k_ms_asya_dep_neg_2,
                                    filter_k_ms_asya_ev_neg_2_neg_1_5,
                                    filter_k_ms_asya_dep_neg_2_neg_1_5,
                                    filter_k_ms_asya_ev_neg_1_5,
                                    filter_k_ms_asya_dep_neg_1_5,
                                    filter_k_ms_asya_ev_neg_1_5_neg_1,
                                    filter_k_ms_asya_dep_neg_1_5_neg_1,
                                    filter_k_ms_asya_ev_neg_1,
                                    filter_k_ms_asya_dep_neg_1,
                                    filter_k_ms_asya_ev_neg_1_neg_0_5,
                                    filter_k_ms_asya_dep_neg_1_neg_0_5,
                                    filter_k_ms_asya_ev_neg_0_5,
                                    filter_k_ms_asya_dep_neg_0_5,
                                    filter_k_ms_asya_ev_neg_0_5_pos_0,
                                    filter_k_ms_asya_dep_neg_0_5_pos_0,
                                    filter_k_ms_asya_ev_pos_0_pos_0_5,
                                    filter_k_ms_asya_dep_pos_0_pos_0_5,
                                    filter_k_ms_asya_ev_pos_0_5,
                                    filter_k_ms_asya_dep_pos_0_5,
                                    filter_k_ms_asya_ev_pos_0_5_pos_1,
                                    filter_k_ms_asya_dep_pos_0_5_pos_1,
                                    filter_k_ms_asya_ev_pos_1,
                                    filter_k_ms_asya_dep_pos_1,
                                    filter_k_ms_asya_ev_pos_1_pos_1_5,
                                    filter_k_ms_asya_dep_pos_1_pos_1_5,
                                    filter_k_ms_asya_ev_pos_1_5,
                                    filter_k_ms_asya_dep_pos_1_5,
                                    filter_k_ms_asya_ev_pos_1_5_pos_2,
                                    filter_k_ms_asya_dep_pos_1_5_pos_2,
                                    filter_k_ms_asya_ev_pos_2,
                                    filter_k_ms_asya_dep_pos_2,
                                    filter_k_iy_avrupa_ev_neg_1,
                                    filter_k_iy_avrupa_ber_neg_1,
                                    filter_k_iy_avrupa_dep_neg_1,
                                    filter_k_iy_avrupa_ev_neg_1_neg_0_5,
                                    filter_k_iy_avrupa_ber_neg_1_neg_0_5,
                                    filter_k_iy_avrupa_dep_neg_1_neg_0_5,
                                    filter_k_iy_avrupa_ev_neg_0_5,
                                    filter_k_iy_avrupa_ber_neg_0_5,
                                    filter_k_iy_avrupa_dep_neg_0_5,
                                    filter_k_iy_avrupa_ev_neg_0_5_pos_0,
                                    filter_k_iy_avrupa_ber_neg_0_5_pos_0,
                                    filter_k_iy_avrupa_dep_neg_0_5_pos_0,
                                    filter_k_iy_avrupa_ev_pos_0_pos_0_5,
                                    filter_k_iy_avrupa_ber_pos_0_pos_0_5,
                                    filter_k_iy_avrupa_dep_pos_0_pos_0_5,
                                    filter_k_iy_avrupa_ev_pos_0_5,
                                    filter_k_iy_avrupa_ber_pos_0_5,
                                    filter_k_iy_avrupa_dep_pos_0_5,
                                    filter_k_iy_avrupa_ev_pos_0_5_pos_1,
                                    filter_k_iy_avrupa_ber_pos_0_5_pos_1,
                                    filter_k_iy_avrupa_dep_pos_0_5_pos_1,
                                    filter_k_iy_avrupa_ev_pos_1,
                                    filter_k_iy_avrupa_ber_pos_1,
                                    filter_k_iy_avrupa_dep_pos_1,
                                    filter_k_ms_avrupa_ev_neg_1,
                                    filter_k_ms_avrupa_ber_neg_1,
                                    filter_k_ms_avrupa_dep_neg_1,
                                    filter_k_ms_avrupa_ev_neg_1_neg_0_5,
                                    filter_k_ms_avrupa_ber_neg_1_neg_0_5,
                                    filter_k_ms_avrupa_dep_neg_1_neg_0_5,
                                    filter_k_ms_avrupa_ev_neg_0_5,
                                    filter_k_ms_avrupa_ber_neg_0_5,
                                    filter_k_ms_avrupa_dep_neg_0_5,
                                    filter_k_ms_avrupa_ev_neg_0_5_pos_0,
                                    filter_k_ms_avrupa_ber_neg_0_5_pos_0,
                                    filter_k_ms_avrupa_dep_neg_0_5_pos_0,
                                    filter_k_ms_avrupa_ev_pos_0_pos_0_5,
                                    filter_k_ms_avrupa_ber_pos_0_pos_0_5,
                                    filter_k_ms_avrupa_dep_pos_0_pos_0_5,
                                    filter_k_ms_avrupa_ev_pos_0_5,
                                    filter_k_ms_avrupa_ber_pos_0_5,
                                    filter_k_ms_avrupa_dep_pos_0_5,
                                    filter_k_ms_avrupa_ev_pos_0_5_pos_1,
                                    filter_k_ms_avrupa_ber_pos_0_5_pos_1,
                                    filter_k_ms_avrupa_dep_pos_0_5_pos_1,
                                    filter_k_ms_avrupa_ev_pos_1,
                                    filter_k_ms_avrupa_ber_pos_1,
                                    filter_k_ms_avrupa_dep_pos_1
                            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)'''

                # Filtreyi ekle
                Cursor.execute(insert_statement, data_to_insert)
                Connection.commit()

                # Bağlantıyı kes
                Connection.close()

                streamlit.info("Kayıt başarılı", icon='🚨')

                streamlit.session_state.Filters_Dictionary['filter_start_date'] = datetime(2019, 1, 1)
                streamlit.session_state.Filters_Dictionary['filter_end_date'] = datetime.today()
                streamlit.session_state.Filters_Dictionary['filter_country'] = 0
                streamlit.session_state.Filters_Dictionary['filter_league'] = 0
                streamlit.session_state.Filters_Dictionary['filter_week'] = 0
                streamlit.session_state.Filters_Dictionary['filter_hour'] = 0
                streamlit.session_state.Filters_Dictionary['filter_home_team'] = 0
                streamlit.session_state.Filters_Dictionary['filter_away_team'] = 0
                streamlit.session_state.Filters_Dictionary['filter_referee'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_ms_1'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_ms_x'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_ms_2'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_ms_0_5_a'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_ms_0_5_ü'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_ms_1_5_a'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_ms_1_5_ü'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_ms_2_5_a'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_ms_2_5_ü'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_ms_3_5_a'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_ms_3_5_ü'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_iy_1'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_iy_x'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_iy_2'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_iy_0_5_a'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_iy_0_5_ü'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_iy_1_5_a'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_iy_1_5_ü'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_ms_1_x_cs'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_ms_1_2_cs'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_ms_x_2_cs'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_iy_1_x_cs'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_iy_1_2_cs'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_iy_x_2_cs'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_home_win'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_away_win'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_iy_kg_var'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_iy_kg_yok'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_ms_kg_var'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_ms_kg_yok'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_tek'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_cift'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_1den1'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_1denx'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_1den2'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_xden1'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_xdenx'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_xden2'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_2den1'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_2denx'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_2den2'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_10'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_20'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_21'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_30'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_31'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_32'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_40'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_41'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_42'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_43'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_00'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_11'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_22'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_33'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_44'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_01'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_02'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_12'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_03'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_13'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_23'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_04'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_14'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_24'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_34'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_10'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_20'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_21'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_30'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_31'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_32'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_40'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_41'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_42'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_43'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_00'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_11'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_22'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_33'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_44'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_01'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_02'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_12'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_03'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_13'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_23'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_04'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_14'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_24'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_34'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_neg_2'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_neg_2'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_neg_2_neg_1_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_neg_2_neg_1_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_neg_1_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_neg_1_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_neg_1_5_neg_1'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_neg_1_5_neg_1'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_neg_1'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_neg_1'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_neg_1_neg_0_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_neg_1_neg_0_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_neg_0_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_neg_0_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_neg_0_5_pos_0'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_neg_0_5_pos_0'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_pos_0_pos_0_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_pos_0_pos_0_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_pos_0_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_pos_0_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_pos_0_5_pos_1'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_pos_0_5_pos_1'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_pos_1'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_pos_1'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_pos_1_pos_1_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_pos_1_pos_1_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_pos_1_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_pos_1_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_pos_1_5_pos_2'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_pos_1_5_pos_2'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_pos_2'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_pos_2'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_neg_2'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_neg_2'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_neg_2_neg_1_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_neg_2_neg_1_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_neg_1_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_neg_1_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_neg_1_5_neg_1'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_neg_1_5_neg_1'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_neg_1'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_neg_1'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_neg_1_neg_0_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_neg_1_neg_0_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_neg_0_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_neg_0_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_neg_0_5_pos_0'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_neg_0_5_pos_0'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_pos_0_pos_0_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_pos_0_pos_0_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_pos_0_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_pos_0_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_pos_0_5_pos_1'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_pos_0_5_pos_1'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_pos_1'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_pos_1'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_pos_1_pos_1_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_pos_1_pos_1_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_pos_1_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_pos_1_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_pos_1_5_pos_2'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_pos_1_5_pos_2'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_pos_2'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_pos_2'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ev_neg_1'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ber_neg_1'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_dep_neg_1'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ev_neg_1_neg_0_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ber_neg_1_neg_0_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_dep_neg_1_neg_0_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ev_neg_0_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ber_neg_0_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_dep_neg_0_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ev_neg_0_5_pos_0'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ber_neg_0_5_pos_0'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_dep_neg_0_5_pos_0'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ev_pos_0_pos_0_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ber_pos_0_pos_0_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_dep_pos_0_pos_0_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ev_pos_0_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ber_pos_0_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_dep_pos_0_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ev_pos_0_5_pos_1'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ber_pos_0_5_pos_1'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_dep_pos_0_5_pos_1'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ev_pos_1'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ber_pos_1'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_dep_pos_1'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ev_neg_1'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ber_neg_1'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_dep_neg_1'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ev_neg_1_neg_0_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ber_neg_1_neg_0_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_dep_neg_1_neg_0_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ev_neg_0_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ber_neg_0_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_dep_neg_0_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ev_neg_0_5_pos_0'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ber_neg_0_5_pos_0'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_dep_neg_0_5_pos_0'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ev_pos_0_pos_0_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ber_pos_0_pos_0_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_dep_pos_0_pos_0_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ev_pos_0_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ber_pos_0_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_dep_pos_0_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ev_pos_0_5_pos_1'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ber_pos_0_5_pos_1'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_dep_pos_0_5_pos_1'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ev_pos_1'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ber_pos_1'] = 0
                streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_dep_pos_1'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_ms_1'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_ms_x'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_ms_2'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_ms_0_5_a'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_ms_0_5_ü'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_ms_1_5_a'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_ms_1_5_ü'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_ms_2_5_a'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_ms_2_5_ü'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_ms_3_5_a'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_ms_3_5_ü'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_iy_1'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_iy_x'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_iy_2'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_iy_0_5_a'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_iy_0_5_ü'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_iy_1_5_a'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_iy_1_5_ü'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_ms_1_x_cs'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_ms_1_2_cs'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_ms_x_2_cs'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_iy_1_x_cs'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_iy_1_2_cs'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_iy_x_2_cs'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_home_win'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_away_win'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_iy_kg_var'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_iy_kg_yok'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_ms_kg_var'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_ms_kg_yok'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_tek'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_cift'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_1den1'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_1denx'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_1den2'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_xden1'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_xdenx'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_xden2'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_2den1'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_2denx'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_2den2'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_10'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_20'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_21'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_30'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_31'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_32'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_40'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_41'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_42'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_43'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_00'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_11'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_22'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_33'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_44'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_01'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_02'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_12'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_03'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_13'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_23'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_04'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_14'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_24'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_34'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_10'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_20'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_21'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_30'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_31'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_32'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_40'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_41'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_42'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_43'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_00'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_11'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_22'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_33'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_44'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_01'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_02'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_12'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_03'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_13'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_23'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_04'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_14'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_24'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_34'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_neg_2'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_neg_2'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_neg_2_neg_1_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_neg_2_neg_1_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_neg_1_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_neg_1_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_neg_1_5_neg_1'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_neg_1_5_neg_1'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_neg_1'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_neg_1'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_neg_1_neg_0_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_neg_1_neg_0_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_neg_0_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_neg_0_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_neg_0_5_pos_0'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_neg_0_5_pos_0'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_pos_0_pos_0_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_pos_0_pos_0_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_pos_0_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_pos_0_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_pos_0_5_pos_1'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_pos_0_5_pos_1'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_pos_1'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_pos_1'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_pos_1_pos_1_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_pos_1_pos_1_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_pos_1_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_pos_1_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_pos_1_5_pos_2'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_pos_1_5_pos_2'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_pos_2'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_pos_2'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_neg_2'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_neg_2'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_neg_2_neg_1_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_neg_2_neg_1_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_neg_1_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_neg_1_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_neg_1_5_neg_1'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_neg_1_5_neg_1'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_neg_1'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_neg_1'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_neg_1_neg_0_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_neg_1_neg_0_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_neg_0_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_neg_0_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_neg_0_5_pos_0'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_neg_0_5_pos_0'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_pos_0_pos_0_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_pos_0_pos_0_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_pos_0_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_pos_0_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_pos_0_5_pos_1'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_pos_0_5_pos_1'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_pos_1'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_pos_1'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_pos_1_pos_1_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_pos_1_pos_1_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_pos_1_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_pos_1_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_pos_1_5_pos_2'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_pos_1_5_pos_2'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_pos_2'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_pos_2'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ev_neg_1'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ber_neg_1'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_dep_neg_1'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ev_neg_1_neg_0_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ber_neg_1_neg_0_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_dep_neg_1_neg_0_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ev_neg_0_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ber_neg_0_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_dep_neg_0_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ev_neg_0_5_pos_0'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ber_neg_0_5_pos_0'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_dep_neg_0_5_pos_0'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ev_pos_0_pos_0_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ber_pos_0_pos_0_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_dep_pos_0_pos_0_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ev_pos_0_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ber_pos_0_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_dep_pos_0_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ev_pos_0_5_pos_1'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ber_pos_0_5_pos_1'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_dep_pos_0_5_pos_1'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ev_pos_1'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ber_pos_1'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_dep_pos_1'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ev_neg_1'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ber_neg_1'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_dep_neg_1'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ev_neg_1_neg_0_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ber_neg_1_neg_0_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_dep_neg_1_neg_0_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ev_neg_0_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ber_neg_0_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_dep_neg_0_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ev_neg_0_5_pos_0'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ber_neg_0_5_pos_0'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_dep_neg_0_5_pos_0'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ev_pos_0_pos_0_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ber_pos_0_pos_0_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_dep_pos_0_pos_0_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ev_pos_0_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ber_pos_0_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_dep_pos_0_5'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ev_pos_0_5_pos_1'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ber_pos_0_5_pos_1'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_dep_pos_0_5_pos_1'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ev_pos_1'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ber_pos_1'] = 0
                streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_dep_pos_1'] = 0

                time.sleep(2)
                CloseDialog = True
        elif streamlit.session_state.button_dialog_no:
            CloseDialog = True

        if CloseDialog:
            streamlit.rerun()

    elif item == "Sil":
        streamlit.write(f"Filtreyi silmek istiyor musunuz?")
        CloseDialog = False
        streamlit.button(label="Evet", key="button_dialog_yes")
        streamlit.button(label="Hayır", key="button_dialog_no")

        if streamlit.session_state.button_dialog_yes:
            # Database bağlantısı
            Connection = sqlite3.connect(Database_Path)

            # SQL komutlarını uygulama
            Cursor = Connection.cursor()

            # Ekleme metni
            remove_statement = "DELETE FROM saved_filters WHERE filter_name = ?"

            # Filtreyi ekle
            Cursor.execute(remove_statement, (streamlit.session_state.selectbox_saved_filters,))
            Connection.commit()

            # Bağlantıyı kes
            Connection.close()

            streamlit.info("Silme başarılı", icon='🚨')

            # Filtreleri Sıfırla
            streamlit.session_state.Filters_Dictionary['filter_start_date'] = datetime(2019, 1, 1)
            streamlit.session_state.Filters_Dictionary['filter_end_date'] = datetime.today()
            streamlit.session_state.Filters_Dictionary['filter_country'] = 0
            streamlit.session_state.Filters_Dictionary['filter_league'] = 0
            streamlit.session_state.Filters_Dictionary['filter_week'] = 0
            streamlit.session_state.Filters_Dictionary['filter_hour'] = 0
            streamlit.session_state.Filters_Dictionary['filter_home_team'] = 0
            streamlit.session_state.Filters_Dictionary['filter_away_team'] = 0
            streamlit.session_state.Filters_Dictionary['filter_referee'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_ms_1'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_ms_x'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_ms_2'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_ms_0_5_a'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_ms_0_5_ü'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_ms_1_5_a'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_ms_1_5_ü'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_ms_2_5_a'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_ms_2_5_ü'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_ms_3_5_a'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_ms_3_5_ü'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_iy_1'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_iy_x'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_iy_2'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_iy_0_5_a'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_iy_0_5_ü'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_iy_1_5_a'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_iy_1_5_ü'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_ms_1_x_cs'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_ms_1_2_cs'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_ms_x_2_cs'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_iy_1_x_cs'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_iy_1_2_cs'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_iy_x_2_cs'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_home_win'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_away_win'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_iy_kg_var'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_iy_kg_yok'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_ms_kg_var'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_ms_kg_yok'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_tek'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_cift'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_1den1'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_1denx'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_1den2'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_xden1'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_xdenx'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_xden2'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_2den1'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_2denx'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_2den2'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_10'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_20'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_21'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_30'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_31'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_32'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_40'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_41'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_42'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_43'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_00'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_11'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_22'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_33'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_44'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_01'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_02'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_12'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_03'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_13'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_23'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_04'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_14'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_24'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_iy_skor_34'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_10'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_20'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_21'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_30'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_31'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_32'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_40'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_41'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_42'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_43'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_00'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_11'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_22'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_33'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_44'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_01'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_02'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_12'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_03'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_13'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_23'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_04'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_14'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_24'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_ms_skor_34'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_neg_2'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_neg_2'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_neg_2_neg_1_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_neg_2_neg_1_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_neg_1_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_neg_1_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_neg_1_5_neg_1'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_neg_1_5_neg_1'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_neg_1'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_neg_1'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_neg_1_neg_0_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_neg_1_neg_0_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_neg_0_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_neg_0_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_neg_0_5_pos_0'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_neg_0_5_pos_0'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_pos_0_pos_0_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_pos_0_pos_0_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_pos_0_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_pos_0_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_pos_0_5_pos_1'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_pos_0_5_pos_1'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_pos_1'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_pos_1'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_pos_1_pos_1_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_pos_1_pos_1_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_pos_1_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_pos_1_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_pos_1_5_pos_2'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_pos_1_5_pos_2'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_ev_pos_2'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_iy_asya_dep_pos_2'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_neg_2'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_neg_2'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_neg_2_neg_1_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_neg_2_neg_1_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_neg_1_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_neg_1_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_neg_1_5_neg_1'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_neg_1_5_neg_1'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_neg_1'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_neg_1'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_neg_1_neg_0_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_neg_1_neg_0_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_neg_0_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_neg_0_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_neg_0_5_pos_0'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_neg_0_5_pos_0'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_pos_0_pos_0_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_pos_0_pos_0_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_pos_0_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_pos_0_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_pos_0_5_pos_1'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_pos_0_5_pos_1'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_pos_1'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_pos_1'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_pos_1_pos_1_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_pos_1_pos_1_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_pos_1_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_pos_1_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_pos_1_5_pos_2'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_pos_1_5_pos_2'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_ev_pos_2'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_ms_asya_dep_pos_2'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ev_neg_1'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ber_neg_1'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_dep_neg_1'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ev_neg_1_neg_0_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ber_neg_1_neg_0_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_dep_neg_1_neg_0_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ev_neg_0_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ber_neg_0_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_dep_neg_0_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ev_neg_0_5_pos_0'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ber_neg_0_5_pos_0'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_dep_neg_0_5_pos_0'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ev_pos_0_pos_0_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ber_pos_0_pos_0_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_dep_pos_0_pos_0_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ev_pos_0_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ber_pos_0_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_dep_pos_0_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ev_pos_0_5_pos_1'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ber_pos_0_5_pos_1'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_dep_pos_0_5_pos_1'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ev_pos_1'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_ber_pos_1'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_iy_avrupa_dep_pos_1'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ev_neg_1'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ber_neg_1'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_dep_neg_1'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ev_neg_1_neg_0_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ber_neg_1_neg_0_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_dep_neg_1_neg_0_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ev_neg_0_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ber_neg_0_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_dep_neg_0_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ev_neg_0_5_pos_0'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ber_neg_0_5_pos_0'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_dep_neg_0_5_pos_0'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ev_pos_0_pos_0_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ber_pos_0_pos_0_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_dep_pos_0_pos_0_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ev_pos_0_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ber_pos_0_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_dep_pos_0_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ev_pos_0_5_pos_1'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ber_pos_0_5_pos_1'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_dep_pos_0_5_pos_1'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ev_pos_1'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_ber_pos_1'] = 0
            streamlit.session_state.Filters_Dictionary['filter_a_ms_avrupa_dep_pos_1'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_ms_1'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_ms_x'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_ms_2'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_ms_0_5_a'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_ms_0_5_ü'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_ms_1_5_a'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_ms_1_5_ü'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_ms_2_5_a'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_ms_2_5_ü'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_ms_3_5_a'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_ms_3_5_ü'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_iy_1'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_iy_x'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_iy_2'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_iy_0_5_a'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_iy_0_5_ü'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_iy_1_5_a'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_iy_1_5_ü'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_ms_1_x_cs'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_ms_1_2_cs'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_ms_x_2_cs'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_iy_1_x_cs'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_iy_1_2_cs'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_iy_x_2_cs'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_home_win'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_away_win'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_iy_kg_var'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_iy_kg_yok'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_ms_kg_var'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_ms_kg_yok'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_tek'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_cift'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_1den1'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_1denx'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_1den2'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_xden1'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_xdenx'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_xden2'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_2den1'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_2denx'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_2den2'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_10'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_20'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_21'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_30'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_31'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_32'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_40'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_41'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_42'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_43'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_00'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_11'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_22'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_33'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_44'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_01'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_02'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_12'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_03'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_13'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_23'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_04'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_14'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_24'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_iy_skor_34'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_10'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_20'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_21'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_30'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_31'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_32'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_40'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_41'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_42'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_43'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_00'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_11'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_22'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_33'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_44'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_01'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_02'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_12'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_03'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_13'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_23'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_04'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_14'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_24'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_ms_skor_34'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_neg_2'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_neg_2'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_neg_2_neg_1_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_neg_2_neg_1_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_neg_1_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_neg_1_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_neg_1_5_neg_1'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_neg_1_5_neg_1'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_neg_1'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_neg_1'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_neg_1_neg_0_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_neg_1_neg_0_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_neg_0_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_neg_0_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_neg_0_5_pos_0'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_neg_0_5_pos_0'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_pos_0_pos_0_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_pos_0_pos_0_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_pos_0_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_pos_0_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_pos_0_5_pos_1'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_pos_0_5_pos_1'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_pos_1'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_pos_1'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_pos_1_pos_1_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_pos_1_pos_1_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_pos_1_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_pos_1_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_pos_1_5_pos_2'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_pos_1_5_pos_2'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_ev_pos_2'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_iy_asya_dep_pos_2'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_neg_2'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_neg_2'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_neg_2_neg_1_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_neg_2_neg_1_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_neg_1_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_neg_1_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_neg_1_5_neg_1'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_neg_1_5_neg_1'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_neg_1'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_neg_1'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_neg_1_neg_0_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_neg_1_neg_0_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_neg_0_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_neg_0_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_neg_0_5_pos_0'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_neg_0_5_pos_0'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_pos_0_pos_0_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_pos_0_pos_0_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_pos_0_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_pos_0_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_pos_0_5_pos_1'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_pos_0_5_pos_1'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_pos_1'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_pos_1'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_pos_1_pos_1_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_pos_1_pos_1_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_pos_1_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_pos_1_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_pos_1_5_pos_2'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_pos_1_5_pos_2'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_ev_pos_2'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_ms_asya_dep_pos_2'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ev_neg_1'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ber_neg_1'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_dep_neg_1'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ev_neg_1_neg_0_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ber_neg_1_neg_0_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_dep_neg_1_neg_0_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ev_neg_0_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ber_neg_0_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_dep_neg_0_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ev_neg_0_5_pos_0'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ber_neg_0_5_pos_0'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_dep_neg_0_5_pos_0'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ev_pos_0_pos_0_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ber_pos_0_pos_0_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_dep_pos_0_pos_0_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ev_pos_0_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ber_pos_0_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_dep_pos_0_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ev_pos_0_5_pos_1'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ber_pos_0_5_pos_1'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_dep_pos_0_5_pos_1'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ev_pos_1'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_ber_pos_1'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_iy_avrupa_dep_pos_1'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ev_neg_1'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ber_neg_1'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_dep_neg_1'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ev_neg_1_neg_0_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ber_neg_1_neg_0_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_dep_neg_1_neg_0_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ev_neg_0_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ber_neg_0_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_dep_neg_0_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ev_neg_0_5_pos_0'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ber_neg_0_5_pos_0'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_dep_neg_0_5_pos_0'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ev_pos_0_pos_0_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ber_pos_0_pos_0_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_dep_pos_0_pos_0_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ev_pos_0_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ber_pos_0_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_dep_pos_0_5'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ev_pos_0_5_pos_1'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ber_pos_0_5_pos_1'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_dep_pos_0_5_pos_1'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ev_pos_1'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_ber_pos_1'] = 0
            streamlit.session_state.Filters_Dictionary['filter_k_ms_avrupa_dep_pos_1'] = 0

            time.sleep(2)
            CloseDialog = True
        elif streamlit.session_state.button_dialog_no:
            CloseDialog = True

        if CloseDialog:
            streamlit.rerun()


# FİLTRE KAYDETME
if streamlit.session_state.button_save_filter:
    Save_Dialog('Kaydet')

# FİLTRE SİLME
if streamlit.session_state.button_delete_filter:
    Save_Dialog('Sil')

# ÜST ANALİZ KISMI
# TODO: BURASI KALDI

# ALT ANALİZ KISMI
if len(streamlit.session_state.Bulletin_Match_List) > 0:
    DateStart = streamlit.session_state.Filters_Dictionary['filter_start_date'].strftime("%d.%m.%Y")
    DateEnd = streamlit.session_state.Filters_Dictionary['filter_end_date'].strftime("%d.%m.%Y")

    # Database bağlantısı
    Connection = sqlite3.connect(Database_Path)

    # SQL komutlarını uygulama
    Cursor = Connection.cursor()

    # value == 1 olan key'leri topla
    selected_keys = [key for key, value in streamlit.session_state.Filters_Dictionary.items() if value == 1]

    Query = f"""
        SELECT * FROM match_bet_data
        WHERE date BETWEEN '{DateStart}' AND '{DateEnd}'
    """

    # Ek filtreleri oluştur
    if selected_keys:
        for key in selected_keys:
            value = streamlit.session_state.Filters_Dictionary[key]
            extra_filters = " AND ".join([f"{Filter_SQLQuery_Cross_Dictionary[key]} = '{MatchData_DataFrame[Filter_SQLQuery_Cross_Dictionary[key]]}'"])
            Query += f" AND {extra_filters}"

    # SQL sorgusunu çalıştır
    cursor.execute(Query)

    # Veri alma
    rows = cursor.fetchall()

    # Sütun adlarını almak için (veya sabit olarak yazabilirsin)
    column_names = [description[0] for description in cursor.description]

    # Pandas DataFrame oluştur
    df = pandas.DataFrame(rows, columns=column_names)

    # Gereksiz sütunları sil
    df = df.drop(columns=['match_id', 'home_team_goal_time_1', 'home_team_goal_time_2', 'home_team_goal_time_3', 'home_team_goal_time_4', 'home_team_goal_time_5',
                          'away_team_goal_time_1', 'away_team_goal_time_2', 'away_team_goal_time_3', 'away_team_goal_time_4', 'away_team_goal_time_5',
                          'home_team_form_1', 'home_team_form_2', 'home_team_form_3', 'home_team_form_4', 'home_team_form_5',
                          'away_team_form_1', 'away_team_form_2', 'away_team_form_3', 'away_team_form_4', 'away_team_form_5',
                          'fulltime_european_handicap_neg_1_neg_0_5_home_open', 'fulltime_european_handicap_neg_1_neg_0_5_home_close', 'fulltime_european_handicap_neg_1_neg_0_5_tie_open', 'fulltime_european_handicap_neg_1_neg_0_5_tie_close', 'fulltime_european_handicap_neg_1_neg_0_5_away_open', 'fulltime_european_handicap_neg_1_neg_0_5_away_close',
                          'fulltime_european_handicap_neg_0_5_home_open', 'fulltime_european_handicap_neg_0_5_home_close', 'fulltime_european_handicap_neg_0_5_tie_open', 'fulltime_european_handicap_neg_0_5_tie_close', 'fulltime_european_handicap_neg_0_5_away_open', 'fulltime_european_handicap_neg_0_5_away_close',
                          'fulltime_european_handicap_neg_0_5_0_0_home_open', 'fulltime_european_handicap_neg_0_5_0_0_home_close', 'fulltime_european_handicap_neg_0_5_0_0_tie_open', 'fulltime_european_handicap_neg_0_5_0_0_tie_close', 'fulltime_european_handicap_neg_0_5_0_0_away_open', 'fulltime_european_handicap_neg_0_5_0_0_away_close',
                          'fulltime_european_handicap_0_0_pos_0_5_home_open', 'fulltime_european_handicap_0_0_pos_0_5_home_close', 'fulltime_european_handicap_0_0_pos_0_5_tie_open', 'fulltime_european_handicap_0_0_pos_0_5_tie_close', 'fulltime_european_handicap_0_0_pos_0_5_away_open', 'fulltime_european_handicap_0_0_pos_0_5_away_close',
                          'fulltime_european_handicap_pos_0_5_home_open', 'fulltime_european_handicap_pos_0_5_home_close', 'fulltime_european_handicap_pos_0_5_tie_open', 'fulltime_european_handicap_pos_0_5_tie_close', 'fulltime_european_handicap_pos_0_5_away_open', 'fulltime_european_handicap_pos_0_5_away_close',
                          'fulltime_european_handicap_pos_0_5_pos_1_0_home_open', 'fulltime_european_handicap_pos_0_5_pos_1_0_home_close', 'fulltime_european_handicap_pos_0_5_pos_1_0_tie_open', 'fulltime_european_handicap_pos_0_5_pos_1_0_tie_close', 'fulltime_european_handicap_pos_0_5_pos_1_0_away_open', 'fulltime_european_handicap_pos_0_5_pos_1_0_away_close',
                          'firsthalf_european_handicap_neg_1_neg_0_5_home_open', 'firsthalf_european_handicap_neg_1_neg_0_5_home_close', 'firsthalf_european_handicap_neg_1_neg_0_5_tie_open', 'firsthalf_european_handicap_neg_1_neg_0_5_tie_close', 'firsthalf_european_handicap_neg_1_neg_0_5_away_open', 'firsthalf_european_handicap_neg_1_neg_0_5_away_close',
                          'firsthalf_european_handicap_neg_0_5_home_open', 'firsthalf_european_handicap_neg_0_5_home_close', 'firsthalf_european_handicap_neg_0_5_tie_open', 'firsthalf_european_handicap_neg_0_5_tie_close', 'firsthalf_european_handicap_neg_0_5_away_open', 'firsthalf_european_handicap_neg_0_5_away_close',
                          'firsthalf_european_handicap_neg_0_5_0_0_home_open', 'firsthalf_european_handicap_neg_0_5_0_0_home_close', 'firsthalf_european_handicap_neg_0_5_0_0_tie_open', 'firsthalf_european_handicap_neg_0_5_0_0_tie_close', 'firsthalf_european_handicap_neg_0_5_0_0_away_open', 'firsthalf_european_handicap_neg_0_5_0_0_away_close',
                          'firsthalf_european_handicap_0_0_pos_0_5_home_open', 'firsthalf_european_handicap_0_0_pos_0_5_home_close', 'firsthalf_european_handicap_0_0_pos_0_5_tie_open', 'firsthalf_european_handicap_0_0_pos_0_5_tie_close', 'firsthalf_european_handicap_0_0_pos_0_5_away_open', 'firsthalf_european_handicap_0_0_pos_0_5_away_close',
                          'firsthalf_european_handicap_pos_0_5_home_open', 'firsthalf_european_handicap_pos_0_5_home_close', 'firsthalf_european_handicap_pos_0_5_tie_open', 'firsthalf_european_handicap_pos_0_5_tie_close', 'firsthalf_european_handicap_pos_0_5_away_open', 'firsthalf_european_handicap_pos_0_5_away_close',
                          'firsthalf_european_handicap_pos_0_5_pos_1_0_home_open', 'firsthalf_european_handicap_pos_0_5_pos_1_0_home_close', 'firsthalf_european_handicap_pos_0_5_pos_1_0_tie_open', 'firsthalf_european_handicap_pos_0_5_pos_1_0_tie_close', 'firsthalf_european_handicap_pos_0_5_pos_1_0_away_open', 'firsthalf_european_handicap_pos_0_5_pos_1_0_away_close'])

    df = df.rename(columns={
        'day': 'Gün',
        'month': 'Ay',
        'year': 'Yıl',
        'weekday': 'Haftanın Günü',
        'date': 'Tarih',
        'hour': 'Saat',
        'country': 'Ülke',
        'league': 'Lig',
        'season': 'Sezon',
        'round': 'Hafta',
        'referee': 'Hakem',
        'home_team_name': 'Ev Sahibi',
        'away_team_name': 'Deplasman',
        'home_team_lineup': 'Ev Sahibi Diziliş',
        'away_team_lineup': 'Deplasman Diziliş',
        'corner_kicks': 'Köşe Vuruşu',
        'first_half_score': 'İlk Yarı Skoru',
        'match_time_score': 'İkinci Yarı Skoru',
        'match_result_score': 'Maç Sonu Skoru',
        'first_half_0_5_over_under': 'İlk Yarı 0.5 Alt/Üst',
        'first_half_1_5_over_under': 'İlk Yarı 1.5 Alt/Üst',
        'match_time_1_5_over_under': 'Maç Sonu 1.5 Alt/Üst',
        'match_time_2_5_over_under': 'Maç Sonu 2.5 Alt/Üst',
        'match_time_3_5_over_under': 'Maç Sonu 3.5 Alt/Üst',
        'both_teams_to_score': 'Karşılıklı Gol Var/Yok',
        'total_goals': 'Toplam Gol',
        'first_half_result': 'İlk Yarı Sonucu',
        'match_time_resul': 'Maç Sonu Sonucu',
        'fh_mt_result': 'İlk Yarı / Maç Sonu',
        'fulltime_home_open': 'MS 1 Açılış',
        'fulltime_home_close': 'MS 1 Kapanış',
        'fulltime_tie_open': 'MS X Açılış',
        'fulltime_tie_close': 'MS X Kapanış',
        'fulltime_away_open': 'MS 2 Açılış',
        'fulltime_away_close': 'MS 2 Kapanış',
        'firsthalf_home_open': '1Y 1 Açılış',
        'firsthalf_home_close': '1Y 1 Kapanış',
        'firsthalf_tie_open': '1Y X Açılış',
        'firsthalf_tie_close': '1Y X Kapanış',
        'firsthalf_away_open': '1Y 2 Açılış',
        'firsthalf_away_close': '1Y 2 Kapanış',
        'secondhalf_home_open': '2Y 1 Açılış',
        'secondhalf_home_close': '2Y 1 Kapanış',
        'secondhalf_tie_open': '2Y X Açılış',
        'secondhalf_tie_close': '2Y X Kapanış',
        'secondhalf_away_open': '2Y 2 Açılış',
        'secondhalf_away_close': '2Y 2 Kapanış',
        'home_home_open': 'İY 1/ MS 1 Açılış',
        'home_home_close': 'İY 1/ MS 1 Kapanış',
        'home_tie_open': 'İY 1 / MS 0 Açılış',
        'home_tie_close': 'İY 1 / MS 0 Kapanış',
        'home_away_open': 'İY 1 / MS 2 Açılış',
        'home_away_close': 'İY 1 / MS 2 Kapanış',
        'tie_home_open': 'İY 0 / MS 1 Açılış',
        'tie_home_close': 'İY 0 / MS 1 Kapanış',
        'tie_tie_open': 'İY 0 / MS 0 Açılış',
        'tie_tie_close': 'İY 0 / MS 0 Kapanış',
        'tie_away_open': 'İY 0 / MS 2 Açılış',
        'tie_away_close': 'İY 0 / MS 2 Kapanış',
        'away_home_open': 'İY 2 / MS 1 Açılış',
        'away_home_close': 'İY 2 / MS 1 Kapanış',
        'away_tie_open': 'İY 2 / MS 0 Açılış',
        'away_tie_close': 'İY 2 / MS 0 Kapanış',
        'away_away_open': 'İY 2 / MS 2 Açılış',
        'away_away_close': 'İY 2 / MS 2 Kapanış',
        'first_half_1_0_open': 'İY 1-0 Skor Açılış',
        'first_half_1_0_close': 'İY 1-0 Skor Kapanış',
        'first_half_2_0_open': 'İY 2-0 Skor Açılış',
        'first_half_2_0_close': 'İY 2-0 Skor Kapanış',
        'first_half_2_1_open': 'İY 2-1 Skor Açılış',
        'first_half_2_1_close': 'İY 2-1 Skor Kapanış',
        'first_half_3_0_open': 'İY 3-0 Skor Açılış',
        'first_half_3_0_close': 'İY 3-0 Skor Kapanış',
        'first_half_3_1_open': 'İY 3-1 Skor Açılış',
        'first_half_3_1_close': 'İY 3-1 Skor Kapanış',
        'first_half_3_2_open': 'İY 3-2 Skor Açılış',
        'first_half_3_2_close': 'İY 3-2 Skor Kapanış',
        'first_half_4_0_open': 'İY 4-0 Skor Açılış',
        'first_half_4_0_close': 'İY 4-0 Skor Kapanış',
        'first_half_4_1_open': 'İY 4-1 Skor Açılış',
        'first_half_4_1_close': 'İY 4-1 Skor Kapanış',
        'first_half_4_2_open': 'İY 4-2 Skor Açılış',
        'first_half_4_2_close': 'İY 4-2 Skor Kapanış',
        'first_half_4_3_open': 'İY 4-3 Skor Açılış',
        'first_half_4_3_close': 'İY 4-3 Skor Kapanış',
        'first_half_0_0_open': 'İY 0-0 Skor Açılış',
        'first_half_0_0_close': 'İY 0-0 Skor Kapanış',
        'first_half_1_1_open': 'İY 1-1 Skor Açılış',
        'first_half_1_1_close': 'İY 1-1 Skor Kapanış',
        'first_half_2_2_open': 'İY 2-2 Skor Açılış',
        'first_half_2_2_close': 'İY 2-2 Skor Kapanış',
        'first_half_3_3_open': 'İY 3-3 Skor Açılış',
        'first_half_3_3_close': 'İY 3-3 Skor Kapanış',
        'first_half_4_4_open': 'İY 4-4 Skor Açılış',
        'first_half_4_4_close': 'İY 4-4 Skor Kapanış',
        'first_half_0_1_open': 'İY 0-1 Skor Açılış',
        'first_half_0_1_close': 'İY 0-1 Skor Kapanış',
        'first_half_0_2_open': 'İY 0-2 Skor Açılış',
        'first_half_0_2_close': 'İY 0-2 Skor Kapanış',
        'first_half_1_2_open': 'İY 1-2 Skor Açılış',
        'first_half_1_2_close': 'İY 1-2 Skor Kapanış',
        'first_half_0_3_open': 'İY 0-3 Skor Açılış',
        'first_half_0_3_close': 'İY 0-3 Skor Kapanış',
        'first_half_1_3_open': 'İY 1-3 Skor Açılış',
        'first_half_1_3_close': 'İY 1-3 Skor Kapanış',
        'first_half_2_3_open': 'İY 2-3 Skor Açılış',
        'first_half_2_3_close': 'İY 2-3 Skor Kapanış',
        'first_half_0_4_open': 'İY 0-4 Skor Açılış',
        'first_half_0_4_close': 'İY 0-4 Skor Kapanış',
        'first_half_1_4_open': 'İY 1-4 Skor Açılış',
        'first_half_1_4_close': 'İY 1-4 Skor Kapanış',
        'first_half_2_4_open': 'İY 2-4 Skor Açılış',
        'first_half_2_4_close': 'İY 2-4 Skor Kapanış',
        'first_half_3_4_open': 'İY 3-4 Skor Açılış',
        'first_half_3_4_close': 'İY 3-4 Skor Kapanış',
        'match_time_1_0_open': 'MS 1-0 Skor Açılış',
        'match_time_1_0_close': 'MS 1-0 Skor Kapanış',
        'match_time_2_0_open': 'MS 2-0 Skor Açılış',
        'match_time_2_0_close': 'MS 2-0 Skor Kapanış',
        'match_time_2_1_open': 'MS 2-1 Skor Açılış',
        'match_time_2_1_close': 'MS 2-1 Skor Kapanış',
        'match_time_3_0_open': 'MS 3-0 Skor Açılış',
        'match_time_3_0_close': 'MS 3-0 Skor Kapanış',
        'match_time_3_1_open': 'MS 3-1 Skor Açılış',
        'match_time_3_1_close': 'MS 3-1 Skor Kapanış',
        'match_time_3_2_open': 'MS 3-2 Skor Açılış',
        'match_time_3_2_close': 'MS 3-2 Skor Kapanış',
        'match_time_4_0_open': 'MS 4-0 Skor Açılış',
        'match_time_4_0_close': 'MS 4-0 Skor Kapanış',
        'match_time_4_1_open': 'MS 4-1 Skor Açılış',
        'match_time_4_1_close': 'MS 4-1 Skor Kapanış',
        'match_time_4_2_open': 'MS 4-2 Skor Açılış',
        'match_time_4_2_close': 'MS 4-2 Skor Kapanış',
        'match_time_4_3_open': 'MS 4-3 Skor Açılış',
        'match_time_4_3_close': 'MS 4-3 Skor Kapanış',
        'match_time_0_0_open': 'MS 0-0 Skor Açılış',
        'match_time_0_0_close': 'MS 0-0 Skor Kapanış',
        'match_time_1_1_open': 'MS 1-1 Skor Açılış',
        'match_time_1_1_close': 'MS 1-1 Skor Kapanış',
        'match_time_2_2_open': 'MS 2-2 Skor Açılış',
        'match_time_2_2_close': 'MS 2-2 Skor Kapanış',
        'match_time_3_3_open': 'MS 3-3 Skor Açılış',
        'match_time_3_3_close': 'MS 3-3 Skor Kapanış',
        'match_time_4_4_open': 'MS 4-4 Skor Açılış',
        'match_time_4_4_close': 'MS 4-4 Skor Kapanış',
        'match_time_0_1_open': 'MS 0-1 Skor Açılış',
        'match_time_0_1_close': 'MS 0-1 Skor Kapanış',
        'match_time_0_2_open': 'MS 0-2 Skor Açılış',
        'match_time_0_2_close': 'MS 0-2 Skor Kapanış',
        'match_time_1_2_open': 'MS 1-2 Skor Açılış',
        'match_time_1_2_close': 'MS 1-2 Skor Kapanış',
        'match_time_0_3_open': 'MS 0-3 Skor Açılış',
        'match_time_0_3_close': 'MS 0-3 Skor Kapanış',
        'match_time_1_3_open': 'MS 1-3 Skor Açılış',
        'match_time_1_3_close': 'MS 1-3 Skor Kapanış',
        'match_time_2_3_open': 'MS 2-3 Skor Açılış',
        'match_time_2_3_close': 'MS 2-3 Skor Kapanış',
        'match_time_0_4_open': 'MS 0-4 Skor Açılış',
        'match_time_0_4_close': 'MS 0-4 Skor Kapanış',
        'match_time_1_4_open': 'MS 1-4 Skor Açılış',
        'match_time_1_4_close': 'MS 1-4 Skor Kapanış',
        'match_time_2_4_open': 'MS 2-4 Skor Açılış',
        'match_time_2_4_close': 'MS 2-4 Skor Kapanış',
        'match_time_3_4_open': 'MS 3-4 Skor Açılış',
        'match_time_3_4_close': 'MS 3-4 Skor Kapanış',
        'first_half_home_and_tie_open': 'İY 1 & 0 Çifte Şans Açılış',
        'first_half_home_and_tie_close': 'İY 1 & 0 Çifte Şans Kapanış',
        'first_half_home_and_away_open': 'İY 1 & 2 Çifte Şans Açılış',
        'first_half_home_and_away_close': 'İY 1 & 2 Çifte Şans Kapanış',
        'first_half_tie_and_away_open': 'İY 0 & 2 Çifte Şans Açılış',
        'first_half_tie_and_away_close': 'İY 0 & 2 Çifte Şans Kapanış',
        'match_time_home_and_tie_open': 'MS 1 & 0 Çifte Şans Açılış',
        'match_time_home_and_tie_close': 'MS 1 & 0 Çifte Şans Kapanış',
        'match_time_home_and_away_open': 'MS 1 & 2 Çifte Şans Açılış',
        'match_time_home_and_away_close': 'MS 1 & 2 Çifte Şans Kapanış',
        'match_time_tie_and_away_open': 'MS 0 & 2 Çifte Şans Açılış',
        'match_time_tie_and_away_close': 'MS 0 & 2 Çifte Şans Kapanış',
        'first_half_over_0_5_open': 'İY 0.5 Üst Açılış',
        'first_half_over_0_5_close': 'İY 0.5 Üst Kapanış',
        'first_half_under_0_5_open': 'İY 0.5 Alt Açılış',
        'first_half_under_0_5_close': 'İY 0.5 Alt Kapanış',
        'first_half_over_1_5_open': 'İY 1.5 Üst Açılış',
        'first_half_over_1_5_close': 'İY 1.5 Üst Kapanış',
        'first_half_under_1_5_open': 'İY 1.5 Alt Açılış',
        'first_half_under_1_5_close': 'İY 1.5 Alt Kapanış',
        'match_time_over_0_5_open': 'MS 0.5 Üst Açılış',
        'match_time_over_0_5_close': 'MS 0.5 Üst Kapanış',
        'match_time_under_0_5_open': 'MS 0.5 Alt Açılış',
        'match_time_under_0_5_close': 'MS 0.5 Alt Kapanış',
        'match_time_over_1_5_open': 'MS 1.5 Üst Açılış',
        'match_time_over_1_5_close': 'MS 1.5 Üst Kapanış',
        'match_time_under_1_5_open': 'MS 1.5 Alt Açılış',
        'match_time_under_1_5_close': 'MS 1.5 Alt Kapanış',
        'match_time_over_2_5_open': 'MS 2.5 Üst Açılış',
        'match_time_over_2_5_close': 'MS 2.5 Üst Kapanış',
        'match_time_under_2_5_open': 'MS 2.5 Alt Açılış',
        'match_time_under_2_5_close': 'MS 2.5 Alt Kapanış',
        'match_time_over_3_5_open': 'MS 3.5 Üst Açılış',
        'match_time_over_3_5_close': 'MS 3.5 Üst Kapanış',
        'match_time_under_3_5_open': 'MS 3.5 Alt Açılış',
        'match_time_under_3_5_close': 'MS 3.5 Alt Kapanış',
        'draw_no_bet_home_open': 'Ev Kazanır Açılış',
        'draw_no_bet_home_close': 'Ev Kazanır Kapanış',
        'draw_no_bet_away_open': 'Dep Kazanır Açılış',
        'draw_no_bet_away_close': 'Dep Kazanır Kapanış',
        'firsthalf_both_teams_to_score_yes_open': 'İY KG Var Açılış',
        'firsthalf_both_teams_to_score_yes_close': 'İY KG Var Kapanış',
        'firsthalf_both_teams_to_score_no_open': 'İY KG Yok Açılış',
        'firsthalf_both_teams_to_score_no_close': 'İY KG Yok Kapanış',
        'both_teams_to_score_yes_open': 'MS KG Var Açılış',
        'both_teams_to_score_yes_close': 'MS KG Var Kapanış',
        'both_teams_to_score_no_open': 'MS KG Yok Açılış',
        'both_teams_to_score_no_close': 'MS KG Yok Kapanış',
        'odd_open': 'Tek Açılış',
        'odd_close': 'Tek Kapanış',
        'even_open': 'Çift Açılış',
        'even_close': 'Çift Kapanış',
        'fulltime_asian_handicap_neg2_home_open': 'MS Asya Handikap -2 Ev Açılış',
        'fulltime_asian_handicap_neg2_home_close': 'MS Asya Handikap -2 Ev Kapanış',
        'fulltime_asian_handicap_neg2_away_open': 'MS Asya Handikap -2 Dep Açılış',
        'fulltime_asian_handicap_neg2_away_close': 'MS Asya Handikap -2 Dep Kapanış',
        'fulltime_asian_handicap_neg2_neg1_5_home_open': 'MS Asya Handikap -2, -1.5 Ev Açılış',
        'fulltime_asian_handicap_neg2_neg1_5_home_close': 'MS Asya Handikap -2, -1.5 Ev Kapanış',
        'fulltime_asian_handicap_neg2_neg1_5_away_open': 'MS Asya Handikap -2, -1.5 Dep Açılış',
        'fulltime_asian_handicap_neg2_neg1_5_away_close': 'MS Asya Handikap -2, -1.5 Dep Kapanış',
        'fulltime_asian_handicap_neg1_5_home_open': 'MS Asya Handikap -1.5 Ev Açılış',
        'fulltime_asian_handicap_neg1_5_home_close': 'MS Asya Handikap -1.5 Ev Kapanış',
        'fulltime_asian_handicap_neg1_5_away_open': 'MS Asya Handikap -1.5 Dep Açılış',
        'fulltime_asian_handicap_neg1_5_away_close': 'MS Asya Handikap -1.5 Dep Kapanış',
        'fulltime_asian_handicap_neg1_5_neg_1_home_open': 'MS Asya Handikap -1.5, -1 Ev Açılış',
        'fulltime_asian_handicap_neg1_5_neg_1_home_close': 'MS Asya Handikap -1.5, -1 Ev Kapanış',
        'fulltime_asian_handicap_neg1_5_neg_1_away_open': 'MS Asya Handikap -1.5, -1 Dep Açılış',
        'fulltime_asian_handicap_neg1_5_neg_1_away_close': 'MS Asya Handikap -1.5, -1 Dep Kapanış',
        'fulltime_asian_handicap_neg_1_home_open': 'MS Asya Handikap -1 Ev Açılış',
        'fulltime_asian_handicap_neg_1_home_close': 'MS Asya Handikap -1 Ev Kapanış',
        'fulltime_asian_handicap_neg_1_away_open': 'MS Asya Handikap -1 Dep Açılış',
        'fulltime_asian_handicap_neg_1_away_close': 'MS Asya Handikap -1 Dep Kapanış',
        'fulltime_asian_handicap_neg_1_neg_0_5_home_open': 'MS Asya Handikap -1, -0.5 Ev Açılış',
        'fulltime_asian_handicap_neg_1_neg_0_5_home_close': 'MS Asya Handikap -1, -0.5 Ev Kapanış',
        'fulltime_asian_handicap_neg_1_neg_0_5_away_open': 'MS Asya Handikap -1, -0.5 Dep Açılış',
        'fulltime_asian_handicap_neg_1_neg_0_5_away_close': 'MS Asya Handikap -1, -0.5 Dep Kapanış',
        'fulltime_asian_handicap_neg_0_5_home_open': 'MS Asya Handikap -0.5 Ev Açılış',
        'fulltime_asian_handicap_neg_0_5_home_close': 'MS Asya Handikap -0.5 Ev Kapanış',
        'fulltime_asian_handicap_neg_0_5_away_open': 'MS Asya Handikap -0.5 Dep Açılış',
        'fulltime_asian_handicap_neg_0_5_away_close': 'MS Asya Handikap -0.5 Dep Kapanış',
        'fulltime_asian_handicap_neg_0_5_0_0_home_open': 'MS Asya Handikap -0.5, 0 Ev Açılış',
        'fulltime_asian_handicap_neg_0_5_0_0_home_close': 'MS Asya Handikap -0.5, 0 Ev Kapanış',
        'fulltime_asian_handicap_neg_0_5_0_0_away_open': 'MS Asya Handikap -0.5, 0 Dep Açılış',
        'fulltime_asian_handicap_neg_0_5_0_0_away_close': 'MS Asya Handikap -0.5, 0 Dep Kapanış',
        'fulltime_asian_handicap_0_0_pos_0_5_home_open': 'MS Asya Handikap 0, 0.5 Ev Açılış',
        'fulltime_asian_handicap_0_0_pos_0_5_home_close': 'MS Asya Handikap 0, 0.5 Ev Kapanış',
        'fulltime_asian_handicap_0_0_pos_0_5_away_open': 'MS Asya Handikap 0, 0.5 Dep Açılış',
        'fulltime_asian_handicap_0_0_pos_0_5_away_close': 'MS Asya Handikap 0, 0.5 Dep Kapanış',
        'fulltime_asian_handicap_pos_0_5_home_open': 'MS Asya Handikap 0.5 Ev Açılış',
        'fulltime_asian_handicap_pos_0_5_home_close': 'MS Asya Handikap 0.5 Ev Kapanış',
        'fulltime_asian_handicap_pos_0_5_away_open': 'MS Asya Handikap 0.5 Dep Açılış',
        'fulltime_asian_handicap_pos_0_5_away_close': 'MS Asya Handikap 0.5 Dep Kapanış',
        'fulltime_asian_handicap_pos_0_5_pos_1_0_home_open': 'MS Asya Handikap 0.5, 1 Ev Açılış',
        'fulltime_asian_handicap_pos_0_5_pos_1_0_home_close': 'MS Asya Handikap 0.5, 1 Ev Kapanış',
        'fulltime_asian_handicap_pos_0_5_pos_1_0_away_open': 'MS Asya Handikap 0.5, 1 Dep Açılış',
        'fulltime_asian_handicap_pos_0_5_pos_1_0_away_close': 'MS Asya Handikap 0.5, 1 Dep Kapanış',
        'fulltime_asian_handicap_pos_1_home_open': 'MS Asya Handikap 1 Ev Açılış',
        'fulltime_asian_handicap_pos_1_home_close': 'MS Asya Handikap 1 Ev Kapanış',
        'fulltime_asian_handicap_pos_1_away_open': 'MS Asya Handikap 1 Dep Açılış',
        'fulltime_asian_handicap_pos_1_away_close': 'MS Asya Handikap 1 Dep Kapanış',
        'fulltime_asian_handicap_pos_1_0_pos_1_5_home_open': 'MS Asya Handikap 1, 1.5 Ev Açılış',
        'fulltime_asian_handicap_pos_1_0_pos_1_5_home_close': 'MS Asya Handikap 1, 1.5 Ev Kapanış',
        'fulltime_asian_handicap_pos_1_0_pos_1_5_away_open': 'MS Asya Handikap 1, 1.5 Dep Açılış',
        'fulltime_asian_handicap_pos_1_0_pos_1_5_away_close': 'MS Asya Handikap 1, 1.5 Dep Kapanış',
        'fulltime_asian_handicap_pos_1_5_home_open': 'MS Asya Handikap 1.5 Ev Açılış',
        'fulltime_asian_handicap_pos_1_5_home_close': 'MS Asya Handikap 1.5 Ev Kapanış',
        'fulltime_asian_handicap_pos_1_5_away_open': 'MS Asya Handikap 1.5 Dep Açılış',
        'fulltime_asian_handicap_pos_1_5_away_close': 'MS Asya Handikap 1.5 Dep Kapanış',
        'fulltime_asian_handicap_pos_1_5_pos_2_0_home_open': 'MS Asya Handikap 1.5, 2 Ev Açılış',
        'fulltime_asian_handicap_pos_1_5_pos_2_0_home_close': 'MS Asya Handikap 1.5, 2 Ev Kapanış',
        'fulltime_asian_handicap_pos_1_5_pos_2_0_away_open': 'MS Asya Handikap 1.5, 2 Dep Açılış',
        'fulltime_asian_handicap_pos_1_5_pos_2_0_away_close': 'MS Asya Handikap 1.5, 2 Dep Kapanış',
        'fulltime_asian_handicap_pos_2_home_open': 'MS Asya Handikap 2 Ev Açılış',
        'fulltime_asian_handicap_pos_2_home_close': 'MS Asya Handikap 2 Ev Kapanış',
        'fulltime_asian_handicap_pos_2_away_open': 'MS Asya Handikap 2 Dep Açılış',
        'fulltime_asian_handicap_pos_2_away_close': 'MS Asya Handikap 2 Dep Kapanış',
        'halftime_asian_handicap_neg2_home_open': 'İY Asya Handikap -2 Ev Açılış',
        'halftime_asian_handicap_neg2_home_close': 'İY Asya Handikap -2 Ev Kapanış',
        'halftime_asian_handicap_neg2_away_open': 'İY Asya Handikap -2 Dep Açılış',
        'halftime_asian_handicap_neg2_away_close': 'İY Asya Handikap -2 Dep Kapanış',
        'halftime_asian_handicap_neg2_neg1_5_home_open': 'İY Asya Handikap -2, -1.5 Ev Açılış',
        'halftime_asian_handicap_neg2_neg1_5_home_close': 'İY Asya Handikap -2, -1.5 Ev Kapanış',
        'halftime_asian_handicap_neg2_neg1_5_away_open': 'İY Asya Handikap -2, -1.5 Dep Açılış',
        'halftime_asian_handicap_neg2_neg1_5_away_close': 'İY Asya Handikap -2, -1.5 Dep Kapanış',
        'halftime_asian_handicap_neg1_5_home_open': 'İY Asya Handikap -1.5 Ev Açılış',
        'halftime_asian_handicap_neg1_5_home_close': 'İY Asya Handikap -1.5 Ev Kapanış',
        'halftime_asian_handicap_neg1_5_away_open': 'İY Asya Handikap -1.5 Dep Açılış',
        'halftime_asian_handicap_neg1_5_away_close': 'İY Asya Handikap -1.5 Dep Kapanış',
        'halftime_asian_handicap_neg1_5_neg_1_home_open': 'İY Asya Handikap -1.5, -1 Ev Açılış',
        'halftime_asian_handicap_neg1_5_neg_1_home_close': 'İY Asya Handikap -1.5, -1 Ev Kapanış',
        'halftime_asian_handicap_neg1_5_neg_1_away_open': 'İY Asya Handikap -1.5, -1 Dep Açılış',
        'halftime_asian_handicap_neg1_5_neg_1_away_close': 'İY Asya Handikap -1.5, -1 Dep Kapanış',
        'halftime_asian_handicap_neg_1_home_open': 'İY Asya Handikap -1 Ev Açılış',
        'halftime_asian_handicap_neg_1_home_close': 'İY Asya Handikap -1 Ev Kapanış',
        'halftime_asian_handicap_neg_1_away_open': 'İY Asya Handikap -1 Dep Açılış',
        'halftime_asian_handicap_neg_1_away_close': 'İY Asya Handikap -1 Dep Kapanış',
        'halftime_asian_handicap_neg_1_neg_0_5_home_open': 'İY Asya Handikap -1, -0.5 Ev Açılış',
        'halftime_asian_handicap_neg_1_neg_0_5_home_close': 'İY Asya Handikap -1, -0.5 Ev Kapanış',
        'halftime_asian_handicap_neg_1_neg_0_5_away_open': 'İY Asya Handikap -1, -0.5 Dep Açılış',
        'halftime_asian_handicap_neg_1_neg_0_5_away_close': 'İY Asya Handikap -1, -0.5 Dep Kapanış',
        'halftime_asian_handicap_neg_0_5_home_open': 'İY Asya Handikap -0.5 Ev Açılış',
        'halftime_asian_handicap_neg_0_5_home_close': 'İY Asya Handikap -0.5 Ev Kapanış',
        'halftime_asian_handicap_neg_0_5_away_open': 'İY Asya Handikap -0.5 Dep Açılış',
        'halftime_asian_handicap_neg_0_5_away_close': 'İY Asya Handikap -0.5 Dep Kapanış',
        'halftime_asian_handicap_neg_0_5_0_0_home_open': 'İY Asya Handikap -0.5, 0 Ev Açılış',
        'halftime_asian_handicap_neg_0_5_0_0_home_close': 'İY Asya Handikap -0.5, 0 Ev Kapanış',
        'halftime_asian_handicap_neg_0_5_0_0_away_open': 'İY Asya Handikap -0.5, 0 Dep Açılış',
        'halftime_asian_handicap_neg_0_5_0_0_away_close': 'İY Asya Handikap -0.5, 0 Dep Kapanış',
        'halftime_asian_handicap_0_0_pos_0_5_home_open': 'İY Asya Handikap 0, 0.5 Ev Açılış',
        'halftime_asian_handicap_0_0_pos_0_5_home_close': 'İY Asya Handikap 0, 0.5 Ev Kapanış',
        'halftime_asian_handicap_0_0_pos_0_5_away_open': 'İY Asya Handikap 0, 0.5 Dep Açılış',
        'halftime_asian_handicap_0_0_pos_0_5_away_close': 'İY Asya Handikap 0, 0.5 Dep Kapanış',
        'halftime_asian_handicap_pos_0_5_home_open': 'İY Asya Handikap 0.5 Ev Açılış',
        'halftime_asian_handicap_pos_0_5_home_close': 'İY Asya Handikap 0.5 Ev Kapanış',
        'halftime_asian_handicap_pos_0_5_away_open': 'İY Asya Handikap 0.5 Dep Açılış',
        'halftime_asian_handicap_pos_0_5_away_close': 'İY Asya Handikap 0.5 Dep Kapanış',
        'halftime_asian_handicap_pos_0_5_pos_1_0_home_open': 'İY Asya Handikap 0.5, 1 Ev Açılış',
        'halftime_asian_handicap_pos_0_5_pos_1_0_home_close': 'İY Asya Handikap 0.5, 1 Ev Kapanış',
        'halftime_asian_handicap_pos_0_5_pos_1_0_away_open': 'İY Asya Handikap 0.5, 1 Dep Açılış',
        'halftime_asian_handicap_pos_0_5_pos_1_0_away_close': 'İY Asya Handikap 0.5, 1 Dep Kapanış',
        'halftime_asian_handicap_pos_1_home_open': 'İY Asya Handikap 1 Ev Açılış',
        'halftime_asian_handicap_pos_1_home_close': 'İY Asya Handikap 1 Ev Kapanış',
        'halftime_asian_handicap_pos_1_away_open': 'İY Asya Handikap 1 Dep Açılış',
        'halftime_asian_handicap_pos_1_away_close': 'İY Asya Handikap 1 Dep Kapanış',
        'halftime_asian_handicap_pos_1_0_pos_1_5_home_open': 'İY Asya Handikap 1, 1.5 Ev Açılış',
        'halftime_asian_handicap_pos_1_0_pos_1_5_home_close': 'İY Asya Handikap 1, 1.5 Ev Kapanış',
        'halftime_asian_handicap_pos_1_0_pos_1_5_away_open': 'İY Asya Handikap 1, 1.5 Dep Açılış',
        'halftime_asian_handicap_pos_1_0_pos_1_5_away_close': 'İY Asya Handikap 1, 1.5 Dep Kapanış',
        'halftime_asian_handicap_pos_1_5_home_open': 'İY Asya Handikap 1.5 Ev Açılış',
        'halftime_asian_handicap_pos_1_5_home_close': 'İY Asya Handikap 1.5 Ev Kapanış',
        'halftime_asian_handicap_pos_1_5_away_open': 'İY Asya Handikap 1.5 Dep Açılış',
        'halftime_asian_handicap_pos_1_5_away_close': 'İY Asya Handikap 1.5 Dep Kapanış',
        'halftime_asian_handicap_pos_1_5_pos_2_0_home_open': 'İY Asya Handikap 1.5, 2 Ev Açılış',
        'halftime_asian_handicap_pos_1_5_pos_2_0_home_close': 'İY Asya Handikap 1.5, 2 Ev Kapanış',
        'halftime_asian_handicap_pos_1_5_pos_2_0_away_open': 'İY Asya Handikap 1.5, 2 Dep Açılış',
        'halftime_asian_handicap_pos_1_5_pos_2_0_away_close': 'İY Asya Handikap 1.5, 2 Dep Kapanış',
        'halftime_asian_handicap_pos_2_home_open': 'İY Asya Handikap 2 Ev Açılış',
        'halftime_asian_handicap_pos_2_home_close': 'İY Asya Handikap 2 Ev Kapanış',
        'halftime_asian_handicap_pos_2_away_open': 'İY Asya Handikap 2 Dep Açılış',
        'halftime_asian_handicap_pos_2_away_close': 'İY Asya Handikap 2 Dep Kapanış',
        'fulltime_european_handicap_neg_1_home_open': 'MS Avrupa Handikap -1 Ev Açılış',
        'fulltime_european_handicap_neg_1_home_close': 'MS Avrupa Handikap -1 Ev Kapanış',
        'fulltime_european_handicap_neg_1_tie_open': 'MS Avrupa Handikap -1 Ber Açılış',
        'fulltime_european_handicap_neg_1_tie_close': 'MS Avrupa Handikap -1 Ber Kapanış',
        'fulltime_european_handicap_neg_1_away_open': 'MS Avrupa Handikap -1 Dep Açılış',
        'fulltime_european_handicap_neg_1_away_close': 'MS Avrupa Handikap -1 Dep Kapanış',
        'fulltime_european_handicap_pos_1_home_open': 'MS Avrupa Handikap 1 Ev Açılış',
        'fulltime_european_handicap_pos_1_home_close': 'MS Avrupa Handikap 1 Ev Kapanış',
        'fulltime_european_handicap_pos_1_tie_open': 'MS Avrupa Handikap 1 Ber Açılış',
        'fulltime_european_handicap_pos_1_tie_close': 'MS Avrupa Handikap 1 Ber Kapanış',
        'fulltime_european_handicap_pos_1_away_open': 'MS Avrupa Handikap 1 Dep Açılış',
        'fulltime_european_handicap_pos_1_away_close': 'MS Avrupa Handikap 1 Dep Kapanış',
        'firsthalf_european_handicap_neg_1_home_open': 'İY Avrupa Handikap -1 Ev Açılış',
        'firsthalf_european_handicap_neg_1_home_close': 'İY Avrupa Handikap -1 Ev Kapanış',
        'firsthalf_european_handicap_neg_1_tie_open': 'İY Avrupa Handikap -1 Ber Açılış',
        'firsthalf_european_handicap_neg_1_tie_close': 'İY Avrupa Handikap -1 Ber Kapanış',
        'firsthalf_european_handicap_neg_1_away_open': 'İY Avrupa Handikap -1 Dep Açılış',
        'firsthalf_european_handicap_neg_1_away_close': 'İY Avrupa Handikap -1 Dep Kapanış',
        'firsthalf_european_handicap_pos_1_home_open': 'İY Avrupa Handikap 1 Ev Açılış',
        'firsthalf_european_handicap_pos_1_home_close': 'İY Avrupa Handikap 1 Ev Kapanış',
        'firsthalf_european_handicap_pos_1_tie_open': 'İY Avrupa Handikap 1 Ber Açılış',
        'firsthalf_european_handicap_pos_1_tie_close': 'İY Avrupa Handikap 1 Ber Kapanış',
        'firsthalf_european_handicap_pos_1_away_open': 'İY Avrupa Handikap 1 Dep Açılış',
        'firsthalf_european_handicap_pos_1_away_close': 'İY Avrupa Handikap 1 Dep Kapanış'
    })

    # Numerik değişim hariç tutulacak kolonlar
    exclude_cols = ["Gün", "Ay", "Yıl", "Hafta", "Köşe Vuruşu"]

    # Koşullu renklendirme yapılacak kolonlar
    highlight_cols = ["İlk Yarı 0.5 Alt/Üst", "İlk Yarı 1.5 Alt/Üst",
                      'Maç Sonu 1.5 Alt/Üst', 'Maç Sonu 2.5 Alt/Üst', 'Maç Sonu 3.5 Alt/Üst',
                      'Karşılıklı Gol Var/Yok', 'Toplam Gol', 'İlk Yarı Sonucu', 'Maç Sonu Sonucu', 'İlk Yarı / Maç Sonu']

    # Sayısal sütunları seç, ama exclude_col hariç
    numeric_cols = df.select_dtypes(include=["float", "int"]).columns
    format_cols = [col for col in numeric_cols if col not in exclude_cols]

    # Streamlit için biçimlendirme (.2f, None -> "-")
    mac_siniri = 500
    if len(df) > mac_siniri:
        df = df.iloc[:mac_siniri, :]

    display_df = df.copy()
    for col in format_cols:
        display_df[col] = display_df[col].apply(lambda x: f"{x:.2f}" if pandas.notnull(x) else "-")


    # Renkleme fonksiyonu
    def highlight_logic(val):
        if isinstance(val, str):
            if "ÜST" in val or "VAR" in val:
                return "background-color: lightgreen;"
            elif "ALT" in val or "YOK" in val:
                return "background-color: salmon;"
            elif "0-1" in val:
                return "background-color: #ffe5b4;"
            elif "2-3" in val:
                return "background-color: #ffcc99;"
            elif "4-5" in val:
                return "background-color: #ff9966;"
            elif "6+" in val:
                return "background-color: #e65c00;"
            elif "MS 1" in val:
                return "background-color: lightgreen;"
            elif "İY 1" in val:
                return "background-color: lightgreen;"
            elif "MS 0" in val:
                return "background-color: #ffcc00;"
            elif "İY 0" in val:
                return "background-color: #ffcc00;"
            elif "MS 2" in val:
                return "background-color: #e60000;"
            elif "İY 2" in val:
                return "background-color: #e60000;"
        return ""


    # Stil objesi
    styled_df = display_df.style
    for col in highlight_cols:
        styled_df = styled_df.applymap(highlight_logic, subset=[col])

    # Streamlit'te göster
    streamlit.dataframe(styled_df)

streamlit.button("Export To Excel", key="button_export_excel")

if streamlit.session_state.button_export_excel:
    # TODO: BURADA KAYDEDERKEN BİR DİALOG DAHA AÇALIM VE İSİM YAZILACAK
    # EK OLARAK DİALOG GELECEĞİ İÇİN SPİNNER VS YAPIP BEKLETİRİZ İŞ BİTİNCE DONE YAPAR KAPANIR
    # Excel'e ham df yazılır (gerçek sayılar, noktayı bozmadan)
    excel_path = "Output.xlsx"
    styled_df.to_excel(excel_path, index=False)

    # Excel dosyasını aç, finansal sayı formatı uygula
    wb = load_workbook(excel_path)
    ws = wb.active

    # Biçim uygulanacak sütunların index'leri
    col_indexes = [df.columns.get_loc(col) + 1 for col in format_cols]

    # Hücre biçimlerini ayarla (virgüllü ama sayısal)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for i in col_indexes:
            cell = row[i - 1]

            if cell.value is not None:
                # Sayısal formata uygun olup olmadığını kontrol et
                try:
                    # Eğer metinse sayıya dönüştür
                    if isinstance(cell.value, str) and cell.value.replace(".", "", 1).isdigit():
                        value = float(cell.value)
                    else:
                        value = cell.value

                    # Sayıya dönüştürülmüşse formatla
                    if isinstance(value, (int, float)):
                        # Sayıyı virgüllü formatta kaydediyoruz
                        cell.value = value  # Sayıyı direkt sakla
                        # Finansal format uygula (Excel formatı)
                        cell.number_format = '#,##0.00'  # Türkçe format için
                except ValueError:
                    # Eğer sayıya dönüştürülemiyorsa (metin vs.)
                    pass

    wb.save(excel_path)
